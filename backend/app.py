from __future__ import annotations

import datetime as dt
import enum
import io
import json
import uuid
import os
import re
import tempfile
from decimal import Decimal
from urllib import error as url_error
from urllib import request as url_request
from typing import Literal, Optional
from urllib.parse import urlparse

import pandas as pd
import jwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from fastapi import Depends, FastAPI, File, Header, HTTPException, Query, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import (
    Date,
    DateTime,
    Enum,
    ForeignKey,
    Numeric,
    String,
    UniqueConstraint,
    create_engine,
    delete,
    func,
    inspect,
    select,
    text,
)
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker


def resolve_database_url() -> str:
    db_url = os.getenv("DATABASE_URL", "").strip()
    app_env = os.getenv("APP_ENV", "").strip().lower()
    on_vercel = bool(os.getenv("VERCEL"))
    local_mode = app_env in {"local", "development", "dev"} or not on_vercel
    if db_url:
        return db_url
    if local_mode:
        return "sqlite:///./reconciliation.db"
    raise RuntimeError("DATABASE_URL is required in non-local environment.")


def normalize_database_url(db_url: str) -> str:
    # Neon/Heroku style URL may use postgres://, SQLAlchemy expects postgresql://
    if db_url.startswith("postgres://"):
        return db_url.replace("postgres://", "postgresql://", 1)
    return db_url


def is_sqlite_url(db_url: str) -> bool:
    return urlparse(db_url).scheme.startswith("sqlite")


DATABASE_URL = normalize_database_url(resolve_database_url())
engine_kwargs = {"echo": False, "future": True, "pool_pre_ping": True}
if is_sqlite_url(DATABASE_URL):
    engine_kwargs["connect_args"] = {"check_same_thread": False}
engine = create_engine(DATABASE_URL, **engine_kwargs)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
JWT_SECRET = os.getenv("JWT_SECRET", "change-this-secret")
JWT_EXPIRE_DAYS = int(os.getenv("JWT_EXPIRE_DAYS", "7"))
SUPABASE_URL = os.getenv("SUPABASE_URL", "").strip().rstrip("/")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "").strip()
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
bearer_scheme = HTTPBearer(auto_error=False)


class Base(DeclarativeBase):
    pass


class Role(str, enum.Enum):
    admin = "admin"
    finance_manager = "finance_manager"
    ops_manager = "ops_manager"
    tech = "tech"
    # 兼容历史角色值，避免老 token / 老数据直接失效
    finance = "finance"
    biz = "biz"
    ops = "ops"


class ReconStatus(str, enum.Enum):
    pending = "待确认"
    confirmed = "已确认"
    issue = "异常待处理"


class BillType(str, enum.Enum):
    channel = "channel"
    rd = "rd"


class BillStatus(str, enum.Enum):
    draft = "待发送"
    sent = "已发送"
    acknowledged = "对方确认"
    disputed = "有异议"


class InvoiceStatus(str, enum.Enum):
    pending = "待开票"
    issued = "已开票"
    voided = "已作废"


class CollectionStatus(str, enum.Enum):
    pending = "待回款"
    partial = "部分回款"
    paid = "已回款"


class ProjectStatus(str, enum.Enum):
    active = "active"
    paused = "paused"


class VariantStatus(str, enum.Enum):
    active = "active"
    paused = "paused"


class DiscountType(str, enum.Enum):
    none = "none"
    rate_01 = "0.1"
    rate_005 = "0.05"


class VersionType(str, enum.Enum):
    regular = "常规版"
    joint = "联运版"
    self_operated = "自运营版"
    discount = "折扣版"


class ServerType(str, enum.Enum):
    mixed = "混服"
    dedicated = "专服"


class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(primary_key=True)
    username: Mapped[str] = mapped_column(String(50), unique=True, index=True)
    role: Mapped[Role] = mapped_column(Enum(Role))


ADMIN_USERNAME = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "123456")

LOCAL_USERS = {
    ADMIN_USERNAME: {"password": ADMIN_PASSWORD, "role": Role.admin},
    "finance_manager": {"password": "123456", "role": Role.finance_manager},
    "ops_manager": {"password": "123456", "role": Role.ops_manager},
    "tech": {"password": "123456", "role": Role.tech},
    "finance": {"password": "123456", "role": Role.finance_manager},
    "biz": {"password": "123456", "role": Role.tech},
    "ops": {"password": "123456", "role": Role.ops_manager},
}


class Channel(Base):
    __tablename__ = "channels"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(100), unique=True, index=True)
    active: Mapped[bool] = mapped_column(default=True)


class Game(Base):
    __tablename__ = "games"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(150), unique=True, index=True)
    # 可选业务编码；导入时可按编码匹配游戏
    game_code: Mapped[Optional[str]] = mapped_column(String(80), nullable=True, index=True)
    rd_company: Mapped[str] = mapped_column(String(150))
    # 研发分成（百分比 0~100），作为游戏级固定值；渠道-游戏映射页将优先使用该值
    rd_share_percent: Mapped[Decimal] = mapped_column(Numeric(6, 2), default=Decimal("0"))
    active: Mapped[bool] = mapped_column(default=True)


class Project(Base):
    __tablename__ = "projects"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(150), unique=True, index=True)
    status: Mapped[ProjectStatus] = mapped_column(Enum(ProjectStatus), default=ProjectStatus.active)
    remark: Mapped[str] = mapped_column(String(500), default="")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class GameVariant(Base):
    __tablename__ = "game_variants"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    variant_name: Mapped[str] = mapped_column(String(100))
    raw_game_name: Mapped[str] = mapped_column(String(200), unique=True, index=True)
    discount_type: Mapped[DiscountType] = mapped_column(Enum(DiscountType), default=DiscountType.none)
    version_type: Mapped[VersionType] = mapped_column(Enum(VersionType), default=VersionType.regular)
    server_type: Mapped[ServerType] = mapped_column(Enum(ServerType), default=ServerType.mixed)
    status: Mapped[VariantStatus] = mapped_column(Enum(VariantStatus), default=VariantStatus.active)
    remark: Mapped[str] = mapped_column(String(500), default="")
    rd_company: Mapped[Optional[str]] = mapped_column(String(200), nullable=True)
    publish_company: Mapped[str] = mapped_column(String(200), default="广州熊动科技有限公司")
    rd_share_percent: Mapped[Optional[Decimal]] = mapped_column(Numeric(6, 2), nullable=True)
    publish_share_percent: Mapped[Optional[Decimal]] = mapped_column(Numeric(6, 2), nullable=True)
    settlement_remark: Mapped[Optional[str]] = mapped_column(String(500), nullable=True)
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    project: Mapped[Project] = relationship()


class ChannelGameMap(Base):
    __tablename__ = "channel_game_map"
    id: Mapped[int] = mapped_column(primary_key=True)
    channel_id: Mapped[int] = mapped_column(ForeignKey("channels.id"), index=True)
    game_id: Mapped[int] = mapped_column(ForeignKey("games.id"), index=True)
    revenue_share_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0.3000"))
    rd_settlement_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0.5000"))
    channel: Mapped[Channel] = relationship()
    game: Mapped[Game] = relationship()


class ReconTask(Base):
    __tablename__ = "recon_tasks"
    id: Mapped[int] = mapped_column(primary_key=True)
    period: Mapped[str] = mapped_column(String(20), index=True)
    status: Mapped[ReconStatus] = mapped_column(Enum(ReconStatus), default=ReconStatus.pending)
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class RawStatement(Base):
    __tablename__ = "raw_statements"
    id: Mapped[int] = mapped_column(primary_key=True)
    recon_task_id: Mapped[int] = mapped_column(ForeignKey("recon_tasks.id"), index=True)
    channel_name: Mapped[str] = mapped_column(String(100))
    game_name: Mapped[str] = mapped_column(String(150))
    period: Mapped[str] = mapped_column(String(20), index=True)
    gross_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2))
    channel_id: Mapped[Optional[int]] = mapped_column(index=True, nullable=True)
    game_id: Mapped[Optional[int]] = mapped_column(index=True, nullable=True)
    mapping_id: Mapped[Optional[int]] = mapped_column(index=True, nullable=True)
    channel_share_percent: Mapped[Optional[Decimal]] = mapped_column(Numeric(6, 2), nullable=True)
    project_id: Mapped[Optional[int]] = mapped_column(index=True, nullable=True)
    project_name: Mapped[Optional[str]] = mapped_column(String(150), nullable=True)
    variant_id: Mapped[Optional[int]] = mapped_column(index=True, nullable=True)
    variant_name: Mapped[Optional[str]] = mapped_column(String(100), nullable=True)
    rd_company: Mapped[Optional[str]] = mapped_column(String(200), nullable=True)
    publish_company: Mapped[Optional[str]] = mapped_column(String(200), nullable=True)
    rd_share_percent: Mapped[Optional[Decimal]] = mapped_column(Numeric(6, 2), nullable=True)
    publish_share_percent: Mapped[Optional[Decimal]] = mapped_column(Numeric(6, 2), nullable=True)
    match_status: Mapped[str] = mapped_column(String(20), default="未匹配")
    variant_match_status: Mapped[str] = mapped_column(String(20), default="未匹配版本")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class ReconIssue(Base):
    __tablename__ = "recon_issues"
    id: Mapped[int] = mapped_column(primary_key=True)
    recon_task_id: Mapped[int] = mapped_column(ForeignKey("recon_tasks.id"), index=True)
    issue_type: Mapped[str] = mapped_column(String(50))
    detail: Mapped[str] = mapped_column(String(300))
    resolved: Mapped[bool] = mapped_column(default=False)


class ReconIssueMeta(Base):
    __tablename__ = "recon_issue_meta"
    id: Mapped[int] = mapped_column(primary_key=True)
    issue_id: Mapped[int] = mapped_column(ForeignKey("recon_issues.id"), unique=True, index=True)
    remark: Mapped[str] = mapped_column(String(300), default="")
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class ReconIssueTimeline(Base):
    __tablename__ = "recon_issue_timeline"
    id: Mapped[int] = mapped_column(primary_key=True)
    issue_id: Mapped[int] = mapped_column(ForeignKey("recon_issues.id"), index=True)
    action: Mapped[str] = mapped_column(String(30), default="resolve")
    from_status: Mapped[str] = mapped_column(String(30), default="未处理")
    to_status: Mapped[str] = mapped_column(String(30), default="已处理")
    remark: Mapped[str] = mapped_column(String(300), default="")
    operator: Mapped[str] = mapped_column(String(50), default="system")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class ImportHistory(Base):
    __tablename__ = "import_history"
    id: Mapped[int] = mapped_column(primary_key=True)
    import_type: Mapped[str] = mapped_column(String(20), index=True)
    period: Mapped[str] = mapped_column(String(20), index=True)
    file_name: Mapped[str] = mapped_column(String(200), default="")
    task_id: Mapped[int] = mapped_column(index=True)
    total_count: Mapped[int] = mapped_column(default=0)
    valid_count: Mapped[int] = mapped_column(default=0)
    invalid_count: Mapped[int] = mapped_column(default=0)
    amount_sum: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    status: Mapped[str] = mapped_column(String(30), default="待确认")
    lifecycle_status: Mapped[str] = mapped_column(String(20), default="active", index=True)
    matched_variant_count: Mapped[int] = mapped_column(default=0)
    unmatched_variant_count: Mapped[int] = mapped_column(default=0)
    summary: Mapped[str] = mapped_column(String(500), default="")
    created_by: Mapped[str] = mapped_column(String(50), default="system")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class InvoiceMeta(Base):
    __tablename__ = "invoice_meta"
    id: Mapped[int] = mapped_column(primary_key=True)
    invoice_id: Mapped[int] = mapped_column(ForeignKey("invoices.id"), unique=True, index=True)
    remark: Mapped[str] = mapped_column(String(200), default="")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class ReceiptMeta(Base):
    __tablename__ = "receipt_meta"
    id: Mapped[int] = mapped_column(primary_key=True)
    receipt_id: Mapped[int] = mapped_column(ForeignKey("receipts.id"), unique=True, index=True)
    remark: Mapped[str] = mapped_column(String(200), default="")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class BillingRule(Base):
    __tablename__ = "billing_rules"
    id: Mapped[int] = mapped_column(primary_key=True)
    name: Mapped[str] = mapped_column(String(100), unique=True)
    bill_type: Mapped[BillType] = mapped_column(Enum(BillType))
    default_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4))
    active: Mapped[bool] = mapped_column(default=True)


class Bill(Base):
    __tablename__ = "bills"
    id: Mapped[int] = mapped_column(primary_key=True)
    bill_type: Mapped[BillType] = mapped_column(Enum(BillType), index=True)
    period: Mapped[str] = mapped_column(String(20), index=True)
    target_name: Mapped[str] = mapped_column(String(150), index=True)
    amount: Mapped[Decimal] = mapped_column(Numeric(18, 2))
    status: Mapped[BillStatus] = mapped_column(Enum(BillStatus), default=BillStatus.draft)
    version: Mapped[int] = mapped_column(default=1)
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    collection_status: Mapped[CollectionStatus] = mapped_column(Enum(CollectionStatus), default=CollectionStatus.pending)
    lifecycle_status: Mapped[str] = mapped_column(String(20), default="active", index=True)


class BillDeliveryLog(Base):
    __tablename__ = "bill_delivery_logs"
    id: Mapped[int] = mapped_column(primary_key=True)
    bill_id: Mapped[int] = mapped_column(ForeignKey("bills.id"), index=True)
    delivery_channel: Mapped[str] = mapped_column(String(50), default="internal")
    delivered_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    note: Mapped[str] = mapped_column(String(300), default="")


class Invoice(Base):
    __tablename__ = "invoices"
    id: Mapped[int] = mapped_column(primary_key=True)
    invoice_no: Mapped[str] = mapped_column(String(100), unique=True, index=True)
    bill_id: Mapped[int] = mapped_column(ForeignKey("bills.id"), index=True)
    issue_date: Mapped[dt.date] = mapped_column(Date)
    amount_without_tax: Mapped[Decimal] = mapped_column(Numeric(18, 2))
    tax_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2))
    total_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2))
    status: Mapped[InvoiceStatus] = mapped_column(Enum(InvoiceStatus), default=InvoiceStatus.pending)


class Receipt(Base):
    __tablename__ = "receipts"
    id: Mapped[int] = mapped_column(primary_key=True)
    bill_id: Mapped[int] = mapped_column(ForeignKey("bills.id"), index=True)
    received_at: Mapped[dt.date] = mapped_column(Date)
    amount: Mapped[Decimal] = mapped_column(Numeric(18, 2))
    bank_ref: Mapped[str] = mapped_column(String(100))
    account_name: Mapped[str] = mapped_column(String(100))


class AuditLog(Base):
    __tablename__ = "audit_logs"
    id: Mapped[int] = mapped_column(primary_key=True)
    actor: Mapped[str] = mapped_column(String(50), index=True)
    action: Mapped[str] = mapped_column(String(100))
    target: Mapped[str] = mapped_column(String(100))
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class SystemAuditLog(Base):
    __tablename__ = "system_audit_logs"
    id: Mapped[int] = mapped_column(primary_key=True)
    operator: Mapped[str] = mapped_column(String(50), index=True)
    action: Mapped[str] = mapped_column(String(100), index=True)
    target_type: Mapped[str] = mapped_column(String(50), index=True)
    target_id: Mapped[str] = mapped_column(String(100), default="")
    summary: Mapped[str] = mapped_column(String(500), default="")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), index=True)


class UserProfile(Base):
    __tablename__ = "user_profiles"
    id: Mapped[int] = mapped_column(primary_key=True)
    email: Mapped[str] = mapped_column(String(255), unique=True, index=True)
    # 使用 VARCHAR 存角色字符串，避免与 Postgres 上复用的旧 enum(role) 类型冲突（旧值不含 finance_manager 等）
    role: Mapped[str] = mapped_column(String(50), index=True)
    is_active: Mapped[bool] = mapped_column(default=True, index=True)
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())


class ChannelSettlementStatement(Base):
    __tablename__ = "channel_settlement_statements"
    __table_args__ = (UniqueConstraint("period", "channel_id", name="uq_settlement_period_channel"),)
    id: Mapped[int] = mapped_column(primary_key=True)
    period: Mapped[str] = mapped_column(String(20), index=True)
    channel_id: Mapped[int] = mapped_column(ForeignKey("channels.id"), index=True)
    total_gross_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_discount_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_settlement_base_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_channel_fee_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_settlement_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    # Phase 1 固定 generated；保留字段便于后续扩展多状态
    status: Mapped[str] = mapped_column(String(20), default="generated", index=True)
    note: Mapped[str] = mapped_column(String(500), default="")
    party_platform_name: Mapped[str] = mapped_column(String(200), default="广州熊动科技有限公司")
    party_channel_name: Mapped[str] = mapped_column(String(200), default="")
    # 对账进度（最小版）：pending=待对账，confirmed=已确认，exported=已导出
    reconciliation_status: Mapped[str] = mapped_column(String(20), default="pending", index=True)
    created_by: Mapped[str] = mapped_column(String(50), default="system")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), index=True)
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), index=True)
    channel: Mapped[Channel] = relationship()


class ChannelSettlementStatementItem(Base):
    __tablename__ = "channel_settlement_statement_items"
    id: Mapped[int] = mapped_column(primary_key=True)
    statement_id: Mapped[int] = mapped_column(ForeignKey("channel_settlement_statements.id"), index=True)
    game_id: Mapped[int] = mapped_column(ForeignKey("games.id"), index=True)
    raw_game_name_snapshot: Mapped[str] = mapped_column(String(150), default="")
    game_name_snapshot: Mapped[str] = mapped_column(String(150), default="")
    gross_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    discount_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    test_fee_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    coupon_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    settlement_base_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    share_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0"))
    channel_fee_rate: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0"))
    channel_fee_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    settlement_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    sort_order: Mapped[int] = mapped_column(default=0)
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    game: Mapped[Game] = relationship()


class ChannelGameContract(Base):
    """渠道-游戏月度结算合同（分成与签章信息）；主数据仍关联 channels / games。"""

    __tablename__ = "channel_game_contracts"
    __table_args__ = (
        UniqueConstraint("channel_id", "game_id", "effective_start_date", name="uq_cgc_channel_game_start"),
    )

    id: Mapped[int] = mapped_column(primary_key=True)
    channel_id: Mapped[int] = mapped_column(ForeignKey("channels.id"), index=True)
    game_id: Mapped[int] = mapped_column(ForeignKey("games.id"), index=True)
    effective_start_date: Mapped[dt.date] = mapped_column(Date, index=True)
    effective_end_date: Mapped[Optional[dt.date]] = mapped_column(Date, nullable=True)
    revenue_share_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0.3000"))
    channel_fee_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0"))
    deduct_test_fee: Mapped[bool] = mapped_column(default=True)
    deduct_coupon_fee: Mapped[bool] = mapped_column(default=True)
    our_company_name: Mapped[str] = mapped_column(String(200), default="")
    our_company_address: Mapped[str] = mapped_column(String(400), default="")
    our_company_phone: Mapped[str] = mapped_column(String(100), default="")
    our_tax_no: Mapped[str] = mapped_column(String(80), default="")
    our_bank_name: Mapped[str] = mapped_column(String(200), default="")
    our_bank_account: Mapped[str] = mapped_column(String(80), default="")
    opposite_company_name: Mapped[str] = mapped_column(String(200), default="")
    opposite_tax_no: Mapped[str] = mapped_column(String(80), default="")
    opposite_bank_name: Mapped[str] = mapped_column(String(200), default="")
    opposite_bank_account: Mapped[str] = mapped_column(String(80), default="")
    statement_title_template: Mapped[str] = mapped_column(String(300), default="")
    remark: Mapped[str] = mapped_column(String(500), default="")
    status: Mapped[str] = mapped_column(String(20), default="active", index=True)
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), onupdate=func.now())


class SettlementImportBatch(Base):
    __tablename__ = "import_batches"

    id: Mapped[int] = mapped_column(primary_key=True)
    batch_name: Mapped[str] = mapped_column(String(200), default="")
    source_file_name: Mapped[str] = mapped_column(String(500), default="")
    settlement_month: Mapped[str] = mapped_column(String(20), index=True)
    channel_id: Mapped[int] = mapped_column(ForeignKey("channels.id"), index=True)
    import_type: Mapped[str] = mapped_column(String(20), default="excel")
    total_rows: Mapped[int] = mapped_column(default=0)
    success_rows: Mapped[int] = mapped_column(default=0)
    failed_rows: Mapped[int] = mapped_column(default=0)
    import_status: Mapped[str] = mapped_column(String(30), default="completed", index=True)
    error_summary: Mapped[str] = mapped_column(String(2000), default="")
    created_by: Mapped[str] = mapped_column(String(100), default="system")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), onupdate=func.now())


class SettlementDetailRow(Base):
    __tablename__ = "settlement_details"

    id: Mapped[int] = mapped_column(primary_key=True)
    batch_id: Mapped[int] = mapped_column(ForeignKey("import_batches.id"), index=True)
    settlement_month: Mapped[str] = mapped_column(String(20), index=True)
    channel_id: Mapped[int] = mapped_column(ForeignKey("channels.id"), index=True)
    game_id: Mapped[Optional[int]] = mapped_column(ForeignKey("games.id"), index=True, nullable=True)
    raw_game_name: Mapped[str] = mapped_column(String(200), default="")
    game_name_snapshot: Mapped[str] = mapped_column(String(200), default="")
    gross_revenue: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    test_fee: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    coupon_fee: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    participation_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    revenue_share_ratio: Mapped[Optional[Decimal]] = mapped_column(Numeric(8, 4), nullable=True)
    channel_fee_ratio: Mapped[Optional[Decimal]] = mapped_column(Numeric(8, 4), nullable=True)
    settlement_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    row_status: Mapped[str] = mapped_column(String(30), default="normal", index=True)
    error_message: Mapped[str] = mapped_column(String(500), default="")
    remark: Mapped[str] = mapped_column(String(500), default="")
    monthly_statement_id: Mapped[Optional[int]] = mapped_column(
        ForeignKey("settlement_statements.id"), index=True, nullable=True
    )
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), onupdate=func.now())


class ChannelMonthlySettlementStatement(Base):
    __tablename__ = "settlement_statements"
    __table_args__ = (UniqueConstraint("settlement_month", "channel_id", name="uq_monthly_stmt_month_channel"),)

    id: Mapped[int] = mapped_column(primary_key=True)
    statement_no: Mapped[str] = mapped_column(String(80), unique=True, index=True)
    settlement_month: Mapped[str] = mapped_column(String(20), index=True)
    channel_id: Mapped[int] = mapped_column(ForeignKey("channels.id"), index=True)
    statement_title: Mapped[str] = mapped_column(String(500), default="")
    our_company_name: Mapped[str] = mapped_column(String(200), default="")
    opposite_company_name: Mapped[str] = mapped_column(String(200), default="")
    total_gross_revenue: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_test_fee: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_coupon_fee: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_participation_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_settlement_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    total_settlement_amount_cn: Mapped[str] = mapped_column(String(300), default="")
    statement_status: Mapped[str] = mapped_column(String(30), default="draft", index=True)
    confirmed_at: Mapped[Optional[dt.datetime]] = mapped_column(DateTime, nullable=True)
    exported_at: Mapped[Optional[dt.datetime]] = mapped_column(DateTime, nullable=True)
    paid_at: Mapped[Optional[dt.datetime]] = mapped_column(DateTime, nullable=True)
    remark: Mapped[str] = mapped_column(String(1000), default="")
    our_company_address: Mapped[str] = mapped_column(String(400), default="")
    our_company_phone: Mapped[str] = mapped_column(String(100), default="")
    our_tax_no: Mapped[str] = mapped_column(String(80), default="")
    our_bank_name: Mapped[str] = mapped_column(String(200), default="")
    our_bank_account: Mapped[str] = mapped_column(String(80), default="")
    opposite_tax_no: Mapped[str] = mapped_column(String(80), default="")
    opposite_bank_name: Mapped[str] = mapped_column(String(200), default="")
    opposite_bank_account: Mapped[str] = mapped_column(String(80), default="")
    created_by: Mapped[str] = mapped_column(String(100), default="system")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), onupdate=func.now())


class ChannelMonthlySettlementItem(Base):
    __tablename__ = "settlement_statement_items"

    id: Mapped[int] = mapped_column(primary_key=True)
    statement_id: Mapped[int] = mapped_column(ForeignKey("settlement_statements.id"), index=True)
    sort_no: Mapped[int] = mapped_column(default=0)
    settlement_month: Mapped[str] = mapped_column(String(20), default="")
    game_name: Mapped[str] = mapped_column(String(200), default="")
    gross_revenue: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    test_fee: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    coupon_fee: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    participation_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    revenue_share_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0"))
    channel_fee_ratio: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0"))
    settlement_amount: Mapped[Decimal] = mapped_column(Numeric(18, 2), default=Decimal("0"))
    remark: Mapped[str] = mapped_column(String(500), default="")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), onupdate=func.now())


class ExceptionHandleStatus(str, enum.Enum):
    pending = "pending"
    ignored = "ignored"
    resolved = "resolved"


class ExceptionHandleRecord(Base):
    __tablename__ = "exception_handle_records"
    id: Mapped[int] = mapped_column(primary_key=True)
    exception_type: Mapped[str] = mapped_column(String(20), index=True)
    exception_id: Mapped[str] = mapped_column(String(100), index=True)
    status: Mapped[ExceptionHandleStatus] = mapped_column(Enum(ExceptionHandleStatus), default=ExceptionHandleStatus.pending, index=True)
    remark: Mapped[str] = mapped_column(String(500), default="")
    updated_by: Mapped[str] = mapped_column(String(50), default="system")
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now(), index=True)


class ContractStatus(str, enum.Enum):
    draft = "draft"
    active = "active"
    terminated = "terminated"
    archived = "archived"


class ContractHeader(Base):
    __tablename__ = "contract_headers"
    id: Mapped[int] = mapped_column(primary_key=True)
    contract_no: Mapped[str] = mapped_column(String(80), unique=True, index=True)
    contract_name: Mapped[str] = mapped_column(String(200))
    channel_name: Mapped[str] = mapped_column(String(150), index=True)
    platform_party_name: Mapped[str] = mapped_column(String(200), default="广州熊动科技有限公司")
    platform_party_address: Mapped[str] = mapped_column(String(500), default="")
    developer_party_name: Mapped[str] = mapped_column(String(200), default="")
    developer_party_address: Mapped[str] = mapped_column(String(500), default="")
    start_date: Mapped[dt.date] = mapped_column(Date, index=True)
    end_date: Mapped[dt.date] = mapped_column(Date, index=True)
    status: Mapped[ContractStatus] = mapped_column(Enum(ContractStatus), default=ContractStatus.draft, index=True)
    remark: Mapped[str] = mapped_column(String(1000), default="")
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    items: Mapped[list["ContractItem"]] = relationship(back_populates="contract", cascade="all, delete-orphan")


class ContractItem(Base):
    __tablename__ = "contract_items"
    id: Mapped[int] = mapped_column(primary_key=True)
    contract_id: Mapped[int] = mapped_column(ForeignKey("contract_headers.id", ondelete="CASCADE"), index=True)
    game_name: Mapped[str] = mapped_column(String(150), index=True)
    channel_name: Mapped[str] = mapped_column(String(150), default="", index=True)
    discount_label: Mapped[str] = mapped_column(String(100), default="")
    discount_rate: Mapped[Decimal] = mapped_column(Numeric(8, 4), default=Decimal("0"))
    channel_share_percent: Mapped[Decimal] = mapped_column(Numeric(6, 2), default=Decimal("0"))
    channel_fee_percent: Mapped[Decimal] = mapped_column(Numeric(6, 2), default=Decimal("0"))
    tax_percent: Mapped[Decimal] = mapped_column(Numeric(6, 2), default=Decimal("0"))
    private_percent: Mapped[Decimal] = mapped_column(Numeric(6, 2), default=Decimal("0"))
    item_remark: Mapped[str] = mapped_column(String(500), default="")
    rd_share_note: Mapped[str] = mapped_column(String(500), default="")
    is_active: Mapped[bool] = mapped_column(default=True, index=True)
    created_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    updated_at: Mapped[dt.datetime] = mapped_column(DateTime, default=func.now())
    contract: Mapped["ContractHeader"] = relationship(back_populates="items")


class ChannelIn(BaseModel):
    name: str


class ChannelBulkCreateIn(BaseModel):
    names: list[str]


class GameBulkCreateIn(BaseModel):
    names: list[str]


class GameIn(BaseModel):
    name: str
    rd_company: str
    rd_share_percent: Decimal = Decimal("0")
    game_code: Optional[str] = None


class ProjectIn(BaseModel):
    name: str
    status: ProjectStatus = ProjectStatus.active
    remark: str = ""


class ProjectStatusPatch(BaseModel):
    status: ProjectStatus


class GameVariantIn(BaseModel):
    project_id: int
    variant_name: str
    raw_game_name: str
    discount_type: DiscountType = DiscountType.none
    version_type: VersionType = VersionType.regular
    server_type: ServerType = ServerType.mixed
    status: VariantStatus = VariantStatus.active
    remark: str = ""
    rd_company: Optional[str] = None
    publish_company: Optional[str] = "广州熊动科技有限公司"
    rd_share_percent: Optional[Decimal] = None
    publish_share_percent: Optional[Decimal] = None
    settlement_remark: Optional[str] = None


class VariantStatusPatch(BaseModel):
    status: VariantStatus


class MapIn(BaseModel):
    channel_id: int
    game_id: int
    revenue_share_ratio: Decimal = Field(default=Decimal("0.3000"))
    rd_settlement_ratio: Decimal = Field(default=Decimal("0.5000"))


class MapBulkCreateItem(BaseModel):
    channel_name: str
    game_name: str


class MapBulkCreateIn(BaseModel):
    items: list[MapBulkCreateItem]


class RuleIn(BaseModel):
    name: str
    bill_type: BillType
    default_ratio: Decimal


class ExceptionStatusPatchIn(BaseModel):
    type: str
    id: str
    status: ExceptionHandleStatus
    remark: str = ""


class ContractHeaderIn(BaseModel):
    contract_no: str
    contract_name: str
    channel_name: str
    platform_party_name: str = "广州熊动科技有限公司"
    platform_party_address: str = ""
    developer_party_name: str = ""
    developer_party_address: str = ""
    start_date: dt.date
    end_date: dt.date
    status: ContractStatus = ContractStatus.draft
    remark: str = ""


class ContractLifecycleIn(BaseModel):
    """activate | terminate | archive | restore_active | restore_draft"""

    action: str


class ContractItemIn(BaseModel):
    game_name: str
    channel_name: str = ""
    discount_label: str = ""
    discount_rate: Decimal = Decimal("0")
    channel_share_percent: Decimal = Decimal("0")
    channel_fee_percent: Decimal = Decimal("0")
    tax_percent: Decimal = Decimal("0")
    private_percent: Decimal = Decimal("0")
    item_remark: str = ""
    rd_share_note: str = ""
    is_active: bool = True


class ContractDraftItemOut(BaseModel):
    game_name: str = ""
    channel_name: str = ""
    discount_label: str = ""
    discount_rate: Decimal = Decimal("0")
    channel_share_percent: Decimal = Decimal("0")
    channel_fee_percent: Decimal = Decimal("0")
    tax_percent: Decimal = Decimal("0")
    private_percent: Decimal = Decimal("0")
    item_remark: str = ""
    rd_share_note: str = ""
    is_active: bool = True


class ContractDraftParseOut(BaseModel):
    contract_no: str = ""
    contract_name: str = ""
    channel_name: str = ""
    platform_party_name: str = ""
    platform_party_address: str = ""
    developer_party_name: str = ""
    developer_party_address: str = ""
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: str = "draft"
    remark: str = ""
    items: list[ContractDraftItemOut] = Field(default_factory=list)


class ContractExcelPreviewRow(BaseModel):
    excel_row: int
    contract_no: str
    contract_name: str
    channel_name: str
    developer_party_name: str = ""
    platform_party_name: str = "广州熊动科技有限公司"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    attachment_preview: str = ""
    remark: str = ""
    status: str = "ready"
    issues: list[str] = Field(default_factory=list)
    duplicate_hint: bool = False


class ContractExcelPreviewOut(BaseModel):
    rows: list[ContractExcelPreviewRow]
    file_label: str = ""


class ContractExcelCommitItem(BaseModel):
    contract_no: str
    contract_name: str
    channel_name: str
    developer_party_name: str = ""
    platform_party_name: str = "广州熊动科技有限公司"
    platform_party_address: str = ""
    developer_party_address: str = ""
    start_date: str
    end_date: str
    remark: str = ""


class ContractExcelCommitIn(BaseModel):
    items: list[ContractExcelCommitItem]


class ContractExcelCommitOut(BaseModel):
    created: int = 0
    skipped: int = 0
    created_ids: list[int] = Field(default_factory=list)
    skip_reasons: list[str] = Field(default_factory=list)


class RuleBulkRow(BaseModel):
    row_no: Optional[int] = None
    game: str
    channel: str
    discount_type: str = "无"
    channel_fee: Decimal = Decimal("0")
    tax_rate: Decimal = Decimal("0")
    rd_share: Decimal = Decimal("0.5")
    private_rate: Decimal = Decimal("0")
    ip_license: Decimal = Decimal("0")
    chaofan_channel: Decimal = Decimal("0")
    chaofan_rd: Decimal = Decimal("0")
    status: str = "启用"
    remark: str = ""


class RuleBulkIn(BaseModel):
    rows: list[RuleBulkRow]


class ResolveIssueIn(BaseModel):
    status: str = "已处理"
    remark: str = ""


class BulkResolveIssuesIn(BaseModel):
    issue_ids: list[int]
    remark: str = ""


class LoginIn(BaseModel):
    username: str
    password: str


class AuthLoginIn(BaseModel):
    email: str
    password: str


class AdminResetPasswordIn(BaseModel):
    email: str
    new_password: str


class SettlementStatementGenerateIn(BaseModel):
    period: str
    channel_id: int
    overwrite: bool = False


class SettlementGenerateAllForPeriodIn(BaseModel):
    period: str
    overwrite: bool = False


class SettlementReconciliationStatusIn(BaseModel):
    reconciliation_status: Literal["pending", "confirmed", "exported"]


class MonthlySettlementGenerateIn(BaseModel):
    settlement_month: str
    channel_id: int
    overwrite: bool = False


class MonthlySettlementStatusPatchIn(BaseModel):
    statement_status: Literal["draft", "pending_confirm", "confirmed", "exported", "paid"]


class SettlementDetailPatchIn(BaseModel):
    game_id: Optional[int] = None
    test_fee: Optional[Decimal] = None
    coupon_fee: Optional[Decimal] = None
    revenue_share_ratio: Optional[Decimal] = None
    channel_fee_ratio: Optional[Decimal] = None
    remark: Optional[str] = None
    row_status: Optional[Literal["normal", "error", "pending_confirm"]] = None


class SettlementContractIn(BaseModel):
    channel_id: int
    game_id: int
    effective_start_date: dt.date
    effective_end_date: Optional[dt.date] = None
    revenue_share_ratio: Decimal = Decimal("0.3")
    channel_fee_ratio: Decimal = Decimal("0")
    deduct_test_fee: bool = True
    deduct_coupon_fee: bool = True
    our_company_name: str = ""
    our_company_address: str = ""
    our_company_phone: str = ""
    our_tax_no: str = ""
    our_bank_name: str = ""
    our_bank_account: str = ""
    opposite_company_name: str = ""
    opposite_tax_no: str = ""
    opposite_bank_name: str = ""
    opposite_bank_account: str = ""
    statement_title_template: str = ""
    remark: str = ""
    status: str = "active"


class BillStatusIn(BaseModel):
    status: BillStatus
    note: str = ""


class CleanupDuplicateBillsIn(BaseModel):
    # 安全起见默认预览；确认无误后传 false 执行真实删除
    dry_run: bool = True


class InvoiceIn(BaseModel):
    invoice_no: str
    bill_id: int
    issue_date: dt.date
    amount_without_tax: Decimal
    tax_amount: Decimal
    total_amount: Decimal
    status: InvoiceStatus = InvoiceStatus.issued
    remark: str = ""


class ReceiptIn(BaseModel):
    bill_id: int
    received_at: dt.date
    amount: Decimal
    bank_ref: str
    account_name: str
    remark: str = ""
    status: Optional[CollectionStatus] = None


class Out(BaseModel):
    model_config = ConfigDict(from_attributes=True)


class BillOut(Out):
    id: int
    bill_type: BillType
    period: str
    target_name: str
    amount: Decimal
    status: BillStatus
    version: int
    collection_status: CollectionStatus


def calc_bill_flow_status(
    bill_amount: Decimal,
    bill_status: BillStatus,
    has_invoice: bool,
    received_total: Decimal,
) -> str:
    if received_total >= bill_amount and bill_amount > 0:
        return "已回款"
    if received_total > 0:
        return "部分回款"
    if has_invoice:
        return "已开票"
    if bill_status == BillStatus.sent:
        return "已发送"
    if bill_status in (BillStatus.acknowledged, BillStatus.disputed):
        return "已生成"
    return "草稿"


class ImportHistoryOut(Out):
    id: int
    import_type: str
    period: str
    file_name: str
    task_id: int
    total_count: int
    valid_count: int
    invalid_count: int
    amount_sum: Decimal
    status: str
    lifecycle_status: str = "active"
    summary: str
    created_by: str
    created_at: dt.datetime
    matched_variant_count: int = 0
    unmatched_variant_count: int = 0
    unresolved_issue_count: int = 0
    resolved_issue_count: int = 0
    task_status: str = ""


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def build_token(subject: str, role: Role) -> str:
    expire_at = dt.datetime.now(dt.timezone.utc) + dt.timedelta(days=JWT_EXPIRE_DAYS)
    payload = {"sub": subject, "role": role.value, "exp": expire_at}
    return jwt.encode(payload, JWT_SECRET, algorithm="HS256")


def unauthorized(message: str = "未登录或登录已过期"):
    return HTTPException(status_code=401, detail={"code": 401, "message": message})


def normalize_role(role: Role) -> Role:
    if role == Role.finance:
        return Role.finance_manager
    if role == Role.ops:
        return Role.ops_manager
    if role == Role.biz:
        return Role.tech
    return role


def write_system_audit(
    db: Session,
    operator: str,
    action: str,
    target_type: str,
    target_id: str = "",
    summary: str = "",
):
    db.add(
        SystemAuditLog(
            operator=operator or "system",
            action=action,
            target_type=target_type,
            target_id=target_id,
            summary=summary,
            created_at=dt.datetime.now(),
        )
    )


def supabase_sign_in(email: str, password: str) -> dict:
    if not SUPABASE_URL or not SUPABASE_ANON_KEY:
        raise HTTPException(status_code=500, detail="Supabase 配置缺失，请设置 SUPABASE_URL / SUPABASE_ANON_KEY")
    url = f"{SUPABASE_URL}/auth/v1/token?grant_type=password"
    payload = json.dumps({"email": email, "password": password}).encode("utf-8")
    req = url_request.Request(
        url,
        method="POST",
        data=payload,
        headers={
            "apikey": SUPABASE_ANON_KEY,
            "Content-Type": "application/json",
        },
    )
    try:
        with url_request.urlopen(req, timeout=15) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except url_error.HTTPError as e:
        body = e.read().decode("utf-8") if hasattr(e, "read") else ""
        try:
            parsed = json.loads(body) if body else {}
        except Exception:
            parsed = {}
        if e.code in (400, 401):
            raise HTTPException(status_code=401, detail="邮箱或密码错误") from e
        raise HTTPException(status_code=502, detail=f"Supabase 登录失败: {parsed.get('error_description') or parsed.get('msg') or 'unknown'}") from e
    except Exception as e:
        raise HTTPException(status_code=502, detail="认证服务连接失败") from e


def validate_new_password(password: str):
    raw = (password or "").strip()
    if len(raw) < 8:
        raise HTTPException(status_code=400, detail="新密码不符合规则：至少 8 位")
    if not any(ch.isalpha() for ch in raw) or not any(ch.isdigit() for ch in raw):
        raise HTTPException(status_code=400, detail="新密码不符合规则：需包含字母和数字")


def supabase_find_user_by_email(email: str) -> Optional[dict]:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase 配置缺失，请设置 SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")
    url = f"{SUPABASE_URL}/auth/v1/admin/users?email={email}"
    req = url_request.Request(
        url,
        method="GET",
        headers={
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Content-Type": "application/json",
        },
    )
    try:
        with url_request.urlopen(req, timeout=15) as resp:
            raw = resp.read().decode("utf-8")
            payload = json.loads(raw) if raw else {}
            users = payload.get("users") or []
            return users[0] if users else None
    except url_error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore") if hasattr(e, "read") else ""
        raise HTTPException(status_code=502, detail=f"Supabase 查询用户失败: {body[:200] or e.reason}") from e
    except Exception as e:
        raise HTTPException(status_code=502, detail="Supabase 查询用户失败，请稍后再试") from e


def supabase_admin_reset_password(user_id: str, new_password: str):
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise HTTPException(status_code=500, detail="Supabase 配置缺失，请设置 SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY")
    url = f"{SUPABASE_URL}/auth/v1/admin/users/{user_id}"
    payload = json.dumps({"password": new_password}).encode("utf-8")
    req = url_request.Request(
        url,
        method="PUT",
        data=payload,
        headers={
            "apikey": SUPABASE_SERVICE_ROLE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
            "Content-Type": "application/json",
        },
    )
    try:
        with url_request.urlopen(req, timeout=15):
            return
    except url_error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore") if hasattr(e, "read") else ""
        if e.code == 400:
            raise HTTPException(status_code=400, detail="新密码不符合要求") from e
        raise HTTPException(status_code=502, detail=f"Supabase 重置失败，请稍后再试: {body[:200] or e.reason}") from e
    except Exception as e:
        raise HTTPException(status_code=502, detail="Supabase 重置失败，请稍后再试") from e


def require_role(roles: list[Role]):
    normalized_roles = {normalize_role(x) for x in roles}

    def checker(
        credentials: HTTPAuthorizationCredentials | None = Depends(bearer_scheme),
        x_user: str = Header(default="system"),
        db: Session = Depends(get_db),
    ):
        if not credentials or credentials.scheme.lower() != "bearer":
            raise unauthorized()
        try:
            payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=["HS256"])
            current_role = normalize_role(Role(payload.get("role", "ops_manager")))
            token_user = payload.get("sub", x_user)
        except Exception as e:
            raise unauthorized() from e
        if current_role not in normalized_roles:
            raise HTTPException(status_code=403, detail="无权限")
        db.add(AuditLog(actor=token_user, action="api_call", target=f"role={current_role.value}"))
        db.commit()
        return {"user": token_user, "role": current_role}

    return checker


def require_roles(roles: list[Role]):
    # RBAC 别名，便于语义化调用
    return require_role(roles)


def parse_profile_role(value: object) -> Role:
    """将 user_profiles.role（字符串或历史 Enum 实例）解析为 Role，供 RBAC 使用。"""
    if isinstance(value, Role):
        return normalize_role(value)
    if isinstance(value, str):
        try:
            return normalize_role(Role(value))
        except ValueError as e:
            raise HTTPException(status_code=500, detail=f"用户角色配置无效: {value}") from e
    raise HTTPException(status_code=500, detail="用户角色配置无效")


def ensure_user_profiles_role_string():
    """Postgres：若 role 列为旧 enum，迁移为 VARCHAR，避免新角色值写入失败。"""
    inspector = inspect(engine)
    if "user_profiles" not in inspector.get_table_names():
        return
    if is_sqlite_url(DATABASE_URL):
        return
    try:
        with engine.connect() as conn:
            row = conn.execute(
                text(
                    "SELECT data_type FROM information_schema.columns "
                    "WHERE table_schema = 'public' AND table_name = 'user_profiles' AND column_name = 'role'"
                )
            ).mappings().first()
        if not row or row["data_type"] != "USER-DEFINED":
            return
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE user_profiles ALTER COLUMN role TYPE VARCHAR(50) USING role::text"))
    except Exception:
        # 迁移失败不阻塞启动；后续插入仍可能失败，需人工处理库表
        pass


def create_default_data(db: Session):
    if db.scalar(select(func.count(User.id))) == 0:
        db.add_all(
            [
                User(username="admin", role=Role.admin),
                User(username="finance", role=Role.finance),
                User(username="ops", role=Role.ops),
                User(username="biz", role=Role.biz),
            ]
        )
    if db.scalar(select(func.count(BillingRule.id))) == 0:
        db.add_all(
            [
                BillingRule(name="默认渠道分成", bill_type=BillType.channel, default_ratio=Decimal("0.3000")),
                BillingRule(name="默认研发结算", bill_type=BillType.rd, default_ratio=Decimal("0.5000")),
            ]
        )
    if db.scalar(select(func.count(UserProfile.id))) == 0:
        db.add_all(
            [
                UserProfile(email="wangmiao@dxyx6888.com", role=Role.admin.value, is_active=True),
                UserProfile(email="caiwu@dxyx6888.com", role=Role.finance_manager.value, is_active=True),
                UserProfile(email="pingce@dxyx6888.com", role=Role.ops_manager.value, is_active=True),
                UserProfile(email="515658123@qq.com", role=Role.tech.value, is_active=True),
            ]
        )
    db.commit()


def ensure_game_variant_settlement_columns():
    inspector = inspect(engine)
    if "game_variants" not in inspector.get_table_names():
        return
    existing = {c["name"] for c in inspector.get_columns("game_variants")}
    statements: list[str] = []
    if "rd_company" not in existing:
        statements.append("ALTER TABLE game_variants ADD COLUMN rd_company VARCHAR(200)")
    if "publish_company" not in existing:
        statements.append("ALTER TABLE game_variants ADD COLUMN publish_company VARCHAR(200) DEFAULT '广州熊动科技有限公司'")
    if "rd_share_percent" not in existing:
        statements.append("ALTER TABLE game_variants ADD COLUMN rd_share_percent NUMERIC(6,2)")
    if "publish_share_percent" not in existing:
        statements.append("ALTER TABLE game_variants ADD COLUMN publish_share_percent NUMERIC(6,2)")
    if "settlement_remark" not in existing:
        statements.append("ALTER TABLE game_variants ADD COLUMN settlement_remark VARCHAR(500)")
    if not statements:
        return
    with engine.begin() as conn:
        for stmt in statements:
            conn.execute(text(stmt))


def ensure_game_share_percent_column():
    inspector = inspect(engine)
    if "games" not in inspector.get_table_names():
        return
    existing = {c["name"] for c in inspector.get_columns("games")}
    if "rd_share_percent" in existing:
        return
    try:
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE games ADD COLUMN rd_share_percent NUMERIC(6,2) DEFAULT 0"))
    except Exception:
        pass


def ensure_bill_lifecycle_status_column():
    inspector = inspect(engine)
    if "bills" not in inspector.get_table_names():
        return
    existing = {c["name"] for c in inspector.get_columns("bills")}
    if "lifecycle_status" in existing:
        return
    try:
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE bills ADD COLUMN lifecycle_status VARCHAR(20) DEFAULT 'active'"))
    except Exception:
        pass


def ensure_import_enrichment_columns():
    inspector = inspect(engine)
    if "raw_statements" in inspector.get_table_names():
        existing = {c["name"] for c in inspector.get_columns("raw_statements")}
        statements: list[str] = []
        if "channel_id" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN channel_id INTEGER")
        if "game_id" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN game_id INTEGER")
        if "mapping_id" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN mapping_id INTEGER")
        if "channel_share_percent" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN channel_share_percent NUMERIC(6,2)")
        if "project_id" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN project_id INTEGER")
        if "project_name" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN project_name VARCHAR(150)")
        if "variant_id" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN variant_id INTEGER")
        if "variant_name" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN variant_name VARCHAR(100)")
        if "rd_company" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN rd_company VARCHAR(200)")
        if "publish_company" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN publish_company VARCHAR(200)")
        if "rd_share_percent" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN rd_share_percent NUMERIC(6,2)")
        if "publish_share_percent" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN publish_share_percent NUMERIC(6,2)")
        if "match_status" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN match_status VARCHAR(20) DEFAULT '未匹配'")
        if "variant_match_status" not in existing:
            statements.append("ALTER TABLE raw_statements ADD COLUMN variant_match_status VARCHAR(20) DEFAULT '未匹配版本'")
        if statements:
            with engine.begin() as conn:
                for stmt in statements:
                    conn.execute(text(stmt))
    if "import_history" in inspector.get_table_names():
        existing = {c["name"] for c in inspector.get_columns("import_history")}
        statements: list[str] = []
        if "matched_variant_count" not in existing:
            statements.append("ALTER TABLE import_history ADD COLUMN matched_variant_count INTEGER DEFAULT 0")
        if "unmatched_variant_count" not in existing:
            statements.append("ALTER TABLE import_history ADD COLUMN unmatched_variant_count INTEGER DEFAULT 0")
        if "lifecycle_status" not in existing:
            statements.append("ALTER TABLE import_history ADD COLUMN lifecycle_status VARCHAR(20) DEFAULT 'active'")
        if statements:
            with engine.begin() as conn:
                for stmt in statements:
                    conn.execute(text(stmt))


def normalize_variant_shares(payload: GameVariantIn) -> tuple[Optional[Decimal], Optional[Decimal]]:
    rd_share = payload.rd_share_percent
    publish_share = payload.publish_share_percent
    if rd_share is not None and (rd_share < 0 or rd_share > 100):
        raise HTTPException(status_code=400, detail="研发分成需在 0~100 之间")
    if publish_share is not None and (publish_share < 0 or publish_share > 100):
        raise HTTPException(status_code=400, detail="发行分成需在 0~100 之间")
    if rd_share is not None and publish_share is None:
        publish_share = Decimal("100") - rd_share
    return rd_share, publish_share


def ensure_contract_item_extra_columns():
    """旧库 contract_items 补列：明细渠道名、行备注。"""
    inspector = inspect(engine)
    if "contract_items" not in inspector.get_table_names():
        return
    existing = {c["name"] for c in inspector.get_columns("contract_items")}
    statements: list[str] = []
    if "channel_name" not in existing:
        statements.append("ALTER TABLE contract_items ADD COLUMN channel_name VARCHAR(150) DEFAULT ''")
    if "item_remark" not in existing:
        statements.append("ALTER TABLE contract_items ADD COLUMN item_remark VARCHAR(500) DEFAULT ''")
    if not statements:
        return
    try:
        with engine.begin() as conn:
            for stmt in statements:
                conn.execute(text(stmt))
    except Exception:
        pass


def ensure_contract_lifecycle_status_migration():
    """旧版 contract_headers.status：void→archived，expired→terminated。"""
    inspector = inspect(engine)
    if "contract_headers" not in inspector.get_table_names():
        return
    try:
        with engine.begin() as conn:
            conn.execute(text("UPDATE contract_headers SET status = 'archived' WHERE status IN ('void','VOID')"))
            conn.execute(text("UPDATE contract_headers SET status = 'terminated' WHERE status IN ('expired','EXPIRED')"))
    except Exception:
        pass


def ensure_contract_tables():
    """兼容旧库：补建 channel 合同表（create_all 通常为幂等，此处再保底一次）。"""
    try:
        ContractHeader.__table__.create(bind=engine, checkfirst=True)
        ContractItem.__table__.create(bind=engine, checkfirst=True)
    except Exception:
        pass
    ensure_contract_item_extra_columns()
    ensure_contract_lifecycle_status_migration()


def ensure_settlement_statement_v2_columns():
    """渠道月结单：甲方/乙方快照；行级测试费/代金券/分成比例。"""
    inspector = inspect(engine)
    try:
        if "channel_settlement_statements" in inspector.get_table_names():
            existing_h = {c["name"] for c in inspector.get_columns("channel_settlement_statements")}
            stmts_h: list[str] = []
            if "party_platform_name" not in existing_h:
                stmts_h.append(
                    "ALTER TABLE channel_settlement_statements ADD COLUMN party_platform_name VARCHAR(200) DEFAULT '广州熊动科技有限公司'"
                )
            if "party_channel_name" not in existing_h:
                stmts_h.append("ALTER TABLE channel_settlement_statements ADD COLUMN party_channel_name VARCHAR(200) DEFAULT ''")
            if stmts_h:
                with engine.begin() as conn:
                    for s in stmts_h:
                        conn.execute(text(s))
            if "reconciliation_status" not in existing_h:
                with engine.begin() as conn:
                    conn.execute(
                        text(
                            "ALTER TABLE channel_settlement_statements ADD COLUMN reconciliation_status VARCHAR(20) DEFAULT 'pending'"
                        )
                    )
        if "channel_settlement_statement_items" in inspector.get_table_names():
            existing_i = {c["name"] for c in inspector.get_columns("channel_settlement_statement_items")}
            stmts_i: list[str] = []
            if "test_fee_amount" not in existing_i:
                stmts_i.append(
                    "ALTER TABLE channel_settlement_statement_items ADD COLUMN test_fee_amount NUMERIC(18,2) DEFAULT 0"
                )
            if "coupon_amount" not in existing_i:
                stmts_i.append("ALTER TABLE channel_settlement_statement_items ADD COLUMN coupon_amount NUMERIC(18,2) DEFAULT 0")
            if "share_ratio" not in existing_i:
                stmts_i.append("ALTER TABLE channel_settlement_statement_items ADD COLUMN share_ratio NUMERIC(8,4) DEFAULT 0")
            if stmts_i:
                with engine.begin() as conn:
                    for s in stmts_i:
                        conn.execute(text(s))
    except Exception:
        pass


def ensure_game_code_column():
    try:
        inspector = inspect(engine)
        if "games" not in inspector.get_table_names():
            return
        cols = {c["name"] for c in inspector.get_columns("games")}
        if "game_code" not in cols:
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE games ADD COLUMN game_code VARCHAR(80)"))
    except Exception:
        pass


app = FastAPI(title="内部对账系统", version="1.0.0")

# 浏览器直连后端（如合同 PDF 识别）需跨域；默认 * + 无 credentials，与 Bearer 头兼容。生产可用 CORS_ALLOW_ORIGINS 收紧。
_cors_raw = os.getenv("CORS_ALLOW_ORIGINS", "*").strip()
_cors_origins = [x.strip() for x in _cors_raw.split(",") if x.strip()] or ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    ensure_user_profiles_role_string()
    ensure_game_variant_settlement_columns()
    ensure_game_share_percent_column()
    ensure_bill_lifecycle_status_column()
    ensure_import_enrichment_columns()
    ensure_contract_tables()
    ensure_settlement_statement_v2_columns()
    ensure_game_code_column()
    with SessionLocal() as db:
        create_default_data(db)


@app.get("/")
def root():
    return {"name": "内部对账系统", "phase": "1-3 已实现基础闭环", "docs": "/docs"}


def _perform_auth_login(email: str, password: str, db: Session):
    email = (email or "").strip().lower()
    if not email or not password:
        raise HTTPException(status_code=400, detail="邮箱和密码不能为空")
    supabase_resp = supabase_sign_in(email, password)
    supabase_user = supabase_resp.get("user") or {}
    auth_email = (supabase_user.get("email") or email).strip().lower()
    profile = db.scalar(select(UserProfile).where(UserProfile.email == auth_email))
    if not profile:
        raise HTTPException(status_code=403, detail="未分配系统角色")
    if not profile.is_active:
        raise HTTPException(status_code=403, detail="账号已停用")
    role = parse_profile_role(profile.role)
    token = build_token(auth_email, role)
    write_system_audit(db, auth_email, "login_success", "auth", auth_email, "Supabase 邮箱登录成功")
    db.commit()
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {
            "email": auth_email,
            "role": role.value,
            "is_active": profile.is_active,
        },
    }


@app.post("/auth/login")
def auth_login(payload: AuthLoginIn, db: Session = Depends(get_db)):
    return _perform_auth_login(payload.email, payload.password, db)


@app.post("/login")
def login_compat(payload: LoginIn, db: Session = Depends(get_db)):
    # 兼容旧前端字段（username），底层统一走 Supabase 邮箱登录
    return _perform_auth_login(payload.username, payload.password, db)


@app.get("/auth/me")
def auth_me(ctx: dict = Depends(require_roles([Role.admin, Role.finance_manager, Role.ops_manager, Role.tech])), db: Session = Depends(get_db)):
    email = (ctx["user"] or "").strip().lower()
    profile = db.scalar(select(UserProfile).where(UserProfile.email == email))
    if not profile:
        raise HTTPException(status_code=403, detail="未分配系统角色")
    if not profile.is_active:
        raise HTTPException(status_code=403, detail="账号已停用")
    role = parse_profile_role(profile.role)
    return {"email": email, "role": role.value, "is_active": profile.is_active}


@app.post("/auth/admin/reset-password")
def auth_admin_reset_password(
    payload: AdminResetPasswordIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_roles([Role.admin])),
):
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="邮箱不能为空")
    validate_new_password(payload.new_password)
    user = supabase_find_user_by_email(email)
    if not user:
        raise HTTPException(status_code=404, detail="未找到该邮箱对应的用户")
    user_id = str(user.get("id") or "").strip()
    if not user_id:
        raise HTTPException(status_code=502, detail="Supabase 用户数据异常")
    supabase_admin_reset_password(user_id, payload.new_password)
    write_system_audit(
        db,
        ctx["user"],
        "reset_user_password",
        "auth_user",
        email,
        f"operator_role=admin; 管理员重置用户密码：{email}",
    )
    db.commit()
    return {"ok": True, "message": "密码已重置"}


@app.post("/channels")
def create_channel(payload: ChannelIn, db: Session = Depends(get_db), ctx: dict = Depends(require_role([Role.admin, Role.biz, Role.ops]))):
    channel = Channel(name=payload.name)
    db.add(channel)
    db.flush()
    write_system_audit(db, ctx["user"], "create_channel", "channel", str(channel.id), f"新增渠道: {payload.name}")
    db.commit()
    return {"id": channel.id, "name": channel.name}


@app.post("/channels/bulk-create")
def bulk_create_channels(
    payload: ChannelBulkCreateIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.biz, Role.ops])),
):
    success_count = 0
    failed_names: list[str] = []
    cleaned_names: list[str] = []
    seen = set()
    for raw in payload.names:
        name = (raw or "").strip()
        if not name:
            continue
        if name in seen:
            continue
        seen.add(name)
        cleaned_names.append(name)
    exists = {x.name for x in db.scalars(select(Channel).where(Channel.name.in_(cleaned_names))).all()} if cleaned_names else set()
    for name in cleaned_names:
        if name in exists:
            failed_names.append(name)
            continue
        db.add(Channel(name=name))
        success_count += 1
    write_system_audit(
        db,
        ctx["user"],
        "bulk_create_channels",
        "channel",
        "",
        f"批量新增渠道: 成功{success_count}, 跳过{len(failed_names)}",
    )
    db.commit()
    return {"success_count": success_count, "failed_count": len(failed_names), "failed_names": failed_names}


@app.get("/channels")
def list_channels(
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.finance, Role.tech, Role.biz, Role.ops])),
):
    return db.scalars(select(Channel).order_by(Channel.id.desc())).all()


@app.put("/channels/{channel_id}")
def update_channel(channel_id: int, payload: ChannelIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.biz]))):
    row = db.get(Channel, channel_id)
    if not row:
        raise HTTPException(status_code=404, detail="渠道不存在")
    row.name = payload.name
    write_system_audit(db, _["user"], "update_channel", "channel", str(row.id), f"编辑渠道: {payload.name}")
    db.commit()
    return {"id": row.id, "name": row.name}


@app.delete("/channels/{channel_id}")
def delete_channel(channel_id: int, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin]))):
    row = db.get(Channel, channel_id)
    if not row:
        raise HTTPException(status_code=404, detail="渠道不存在")
    write_system_audit(db, _["user"], "delete_channel", "channel", str(row.id), f"删除渠道: {row.name}")
    db.delete(row)
    db.commit()
    return {"id": channel_id, "deleted": True}


def _contract_status_value(s: ContractStatus | str) -> str:
    if isinstance(s, ContractStatus):
        return s.value
    return str(s)


def _contract_effective_meta(h: ContractHeader, today: Optional[dt.date] = None) -> dict:
    """基于存储状态 + 自然日计算对外展示状态与到期提醒（以后端口径为准）。"""
    today = today or dt.date.today()
    st = h.status
    sd, ed = h.start_date, h.end_date
    days_to_end: Optional[int] = int((ed - today).days) if ed else None

    if st == ContractStatus.archived:
        return {"effective_status": "archived", "days_to_end": days_to_end, "expiry_reminder": ""}
    if st == ContractStatus.terminated:
        return {"effective_status": "terminated", "days_to_end": days_to_end, "expiry_reminder": ""}
    if st == ContractStatus.draft:
        return {"effective_status": "draft", "days_to_end": days_to_end, "expiry_reminder": ""}

    if ed and ed < today:
        return {"effective_status": "expired", "days_to_end": days_to_end, "expiry_reminder": "已过期"}
    if sd and sd > today:
        d0 = (sd - today).days
        return {
            "effective_status": "pending_start",
            "days_to_end": days_to_end,
            "expiry_reminder": f"距开始还有 {d0} 天",
        }
    if days_to_end is None:
        return {"effective_status": "active", "days_to_end": None, "expiry_reminder": ""}
    if days_to_end == 0:
        return {"effective_status": "expiring_soon", "days_to_end": 0, "expiry_reminder": "今天到期"}
    if 1 <= days_to_end <= 30:
        return {
            "effective_status": "expiring_soon",
            "days_to_end": days_to_end,
            "expiry_reminder": f"{days_to_end}天后到期",
        }
    if days_to_end > 30:
        return {
            "effective_status": "active",
            "days_to_end": days_to_end,
            "expiry_reminder": f"还剩 {days_to_end} 天",
        }
    return {"effective_status": "active", "days_to_end": days_to_end, "expiry_reminder": ""}


def _assert_contract_item_payload(payload: ContractItemIn) -> tuple[str, str]:
    gm = (payload.game_name or "").strip()
    ch = (payload.channel_name or "").strip()
    if not gm:
        raise HTTPException(status_code=400, detail="游戏名称不能为空")
    if not ch:
        raise HTTPException(status_code=400, detail="明细渠道名称不能为空")
    dr = payload.discount_rate
    if dr < 0 or dr > Decimal("100"):
        raise HTTPException(status_code=400, detail="折扣率需在0~100之间")
    for label, v in (
        ("渠道分成比例", payload.channel_share_percent),
        ("通道费比例", payload.channel_fee_percent),
        ("税点", payload.tax_percent),
        ("私点", payload.private_percent),
    ):
        if v < 0 or v > Decimal("100"):
            raise HTTPException(status_code=400, detail=f"{label}需在0~100之间")
    return gm, ch


def _contract_item_dict(it: ContractItem) -> dict:
    return {
        "id": it.id,
        "contract_id": it.contract_id,
        "game_name": it.game_name,
        "channel_name": it.channel_name or "",
        "discount_label": it.discount_label or "",
        "discount_rate": float(it.discount_rate) if it.discount_rate is not None else 0.0,
        "channel_share_percent": float(it.channel_share_percent) if it.channel_share_percent is not None else 0.0,
        "channel_fee_percent": float(it.channel_fee_percent) if it.channel_fee_percent is not None else 0.0,
        "tax_percent": float(it.tax_percent) if it.tax_percent is not None else 0.0,
        "private_percent": float(it.private_percent) if it.private_percent is not None else 0.0,
        "item_remark": it.item_remark or "",
        "rd_share_note": it.rd_share_note or "",
        "is_active": bool(it.is_active),
        "created_at": str(it.created_at) if it.created_at else "",
        "updated_at": str(it.updated_at) if it.updated_at else "",
    }


def _contract_header_dict(h: ContractHeader, include_items: bool) -> dict:
    meta = _contract_effective_meta(h)
    d: dict = {
        "id": h.id,
        "contract_no": h.contract_no,
        "contract_name": h.contract_name,
        "channel_name": h.channel_name,
        "platform_party_name": h.platform_party_name or "",
        "platform_party_address": h.platform_party_address or "",
        "developer_party_name": h.developer_party_name or "",
        "developer_party_address": h.developer_party_address or "",
        "start_date": h.start_date.isoformat() if h.start_date else None,
        "end_date": h.end_date.isoformat() if h.end_date else None,
        "stored_status": _contract_status_value(h.status),
        "status": meta["effective_status"],
        "effective_status": meta["effective_status"],
        "days_to_end": meta["days_to_end"],
        "expiry_reminder": meta["expiry_reminder"],
        "remark": h.remark or "",
        "created_at": str(h.created_at) if h.created_at else "",
        "updated_at": str(h.updated_at) if h.updated_at else "",
    }
    if include_items:
        items = sorted(h.items, key=lambda x: x.id)
        d["items"] = [_contract_item_dict(x) for x in items]
    return d


@app.get("/contracts")
def list_contracts(
    channel: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech, Role.ops])),
):
    today = dt.date.today()
    stmt = select(ContractHeader).order_by(ContractHeader.id.desc())
    if channel and channel.strip():
        stmt = stmt.where(ContractHeader.channel_name.contains(channel.strip()))
    ef = (status or "").strip()
    if ef == "draft":
        stmt = stmt.where(ContractHeader.status == ContractStatus.draft)
    elif ef == "terminated":
        stmt = stmt.where(ContractHeader.status == ContractStatus.terminated)
    elif ef == "archived":
        stmt = stmt.where(ContractHeader.status == ContractStatus.archived)
    elif ef == "active":
        stmt = stmt.where(
            ContractHeader.status == ContractStatus.active,
            ContractHeader.start_date <= today,
            ContractHeader.end_date > today + dt.timedelta(days=30),
        )
    elif ef == "expiring_soon":
        stmt = stmt.where(
            ContractHeader.status == ContractStatus.active,
            ContractHeader.start_date <= today,
            ContractHeader.end_date >= today,
            ContractHeader.end_date <= today + dt.timedelta(days=30),
        )
    elif ef == "expired":
        stmt = stmt.where(
            ContractHeader.status == ContractStatus.active,
            ContractHeader.end_date < today,
        )
    elif ef == "pending_start":
        stmt = stmt.where(
            ContractHeader.status == ContractStatus.active,
            ContractHeader.start_date > today,
        )
    rows = db.scalars(stmt).all()
    return [_contract_header_dict(r, False) for r in rows]


@app.post("/contracts")
def create_contract(
    payload: ContractHeaderIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    cn = (payload.contract_no or "").strip()
    if not cn:
        raise HTTPException(status_code=400, detail="合同编号不能为空")
    if db.scalar(select(ContractHeader).where(ContractHeader.contract_no == cn)):
        raise HTTPException(status_code=400, detail="合同编号已存在")
    if payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
    row = ContractHeader(
        contract_no=cn,
        contract_name=(payload.contract_name or "").strip() or cn,
        channel_name=(payload.channel_name or "").strip(),
        platform_party_name=(payload.platform_party_name or "广州熊动科技有限公司").strip() or "广州熊动科技有限公司",
        platform_party_address=(payload.platform_party_address or "").strip(),
        developer_party_name=(payload.developer_party_name or "").strip(),
        developer_party_address=(payload.developer_party_address or "").strip(),
        start_date=payload.start_date,
        end_date=payload.end_date,
        status=payload.status,
        remark=(payload.remark or "").strip(),
    )
    if not row.channel_name:
        raise HTTPException(status_code=400, detail="渠道名称不能为空")
    db.add(row)
    db.flush()
    write_system_audit(db, ctx["user"], "create_contract", "contract_header", str(row.id), f"新增合同: {row.contract_no}")
    db.commit()
    db.refresh(row)
    return _contract_header_dict(row, False)


@app.get("/contracts/{contract_id}")
def get_contract(
    contract_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech, Role.ops])),
):
    row = db.get(ContractHeader, contract_id)
    if not row:
        raise HTTPException(status_code=404, detail="合同不存在")
    return _contract_header_dict(row, True)


@app.put("/contracts/{contract_id}")
def update_contract(
    contract_id: int,
    payload: ContractHeaderIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    row = db.get(ContractHeader, contract_id)
    if not row:
        raise HTTPException(status_code=404, detail="合同不存在")
    cn = (payload.contract_no or "").strip()
    if not cn:
        raise HTTPException(status_code=400, detail="合同编号不能为空")
    dup = db.scalar(select(ContractHeader).where(ContractHeader.contract_no == cn, ContractHeader.id != contract_id))
    if dup:
        raise HTTPException(status_code=400, detail="合同编号已存在")
    if payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
    ch = (payload.channel_name or "").strip()
    if not ch:
        raise HTTPException(status_code=400, detail="渠道名称不能为空")
    row.contract_no = cn
    row.contract_name = (payload.contract_name or "").strip() or cn
    row.channel_name = ch
    row.platform_party_name = (payload.platform_party_name or "广州熊动科技有限公司").strip() or "广州熊动科技有限公司"
    row.platform_party_address = (payload.platform_party_address or "").strip()
    row.developer_party_name = (payload.developer_party_name or "").strip()
    row.developer_party_address = (payload.developer_party_address or "").strip()
    row.start_date = payload.start_date
    row.end_date = payload.end_date
    row.status = payload.status
    row.remark = (payload.remark or "").strip()
    row.updated_at = dt.datetime.now()
    write_system_audit(db, ctx["user"], "update_contract", "contract_header", str(row.id), f"更新合同: {row.contract_no}")
    db.commit()
    db.refresh(row)
    return _contract_header_dict(row, True)


@app.post("/contracts/{contract_id}/lifecycle")
def contract_lifecycle(
    contract_id: int,
    payload: ContractLifecycleIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    row = db.get(ContractHeader, contract_id)
    if not row:
        raise HTTPException(status_code=404, detail="合同不存在")
    act = (payload.action or "").strip().lower()
    if act == "activate":
        if row.status != ContractStatus.draft:
            raise HTTPException(status_code=400, detail="仅草稿可设为生效")
        row.status = ContractStatus.active
    elif act == "terminate":
        if row.status != ContractStatus.active:
            raise HTTPException(status_code=400, detail="仅生效中的合同可终止")
        row.status = ContractStatus.terminated
    elif act == "archive":
        if row.status != ContractStatus.terminated:
            raise HTTPException(status_code=400, detail="仅已终止的合同可归档")
        row.status = ContractStatus.archived
    elif act == "restore_active":
        if row.status not in (ContractStatus.terminated, ContractStatus.archived):
            raise HTTPException(status_code=400, detail="仅已终止或已归档的合同可恢复为生效")
        row.status = ContractStatus.active
    elif act == "restore_draft":
        if row.status not in (
            ContractStatus.active,
            ContractStatus.terminated,
            ContractStatus.archived,
        ):
            raise HTTPException(status_code=400, detail="当前状态无法恢复为草稿")
        row.status = ContractStatus.draft
    else:
        raise HTTPException(status_code=400, detail="未知操作：activate | terminate | archive | restore_active | restore_draft")
    row.updated_at = dt.datetime.now()
    write_system_audit(
        db,
        ctx["user"],
        "contract_lifecycle",
        "contract_header",
        str(row.id),
        f"合同状态流转: {act} -> {row.status.value}",
    )
    db.commit()
    db.refresh(row)
    return _contract_header_dict(row, True)


@app.post("/contracts/{contract_id}/items")
def create_contract_item(
    contract_id: int,
    payload: ContractItemIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    header = db.get(ContractHeader, contract_id)
    if not header:
        raise HTTPException(status_code=404, detail="合同不存在")
    gm, ch = _assert_contract_item_payload(payload)
    row = ContractItem(
        contract_id=contract_id,
        game_name=gm,
        channel_name=ch,
        discount_label=(payload.discount_label or "").strip(),
        discount_rate=payload.discount_rate,
        channel_share_percent=payload.channel_share_percent,
        channel_fee_percent=payload.channel_fee_percent,
        tax_percent=payload.tax_percent,
        private_percent=payload.private_percent,
        item_remark=(payload.item_remark or "").strip()[:500],
        rd_share_note=(payload.rd_share_note or "").strip(),
        is_active=bool(payload.is_active),
    )
    db.add(row)
    db.flush()
    write_system_audit(db, ctx["user"], "create_contract_item", "contract_item", str(row.id), f"合同{contract_id} 新增明细: {gm}")
    db.commit()
    db.refresh(row)
    return _contract_item_dict(row)


@app.put("/contract-items/{item_id}")
def update_contract_item(
    item_id: int,
    payload: ContractItemIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    row = db.get(ContractItem, item_id)
    if not row:
        raise HTTPException(status_code=404, detail="合同明细不存在")
    gm, ch = _assert_contract_item_payload(payload)
    row.game_name = gm
    row.channel_name = ch
    row.discount_label = (payload.discount_label or "").strip()
    row.discount_rate = payload.discount_rate
    row.channel_share_percent = payload.channel_share_percent
    row.channel_fee_percent = payload.channel_fee_percent
    row.tax_percent = payload.tax_percent
    row.private_percent = payload.private_percent
    row.item_remark = (payload.item_remark or "").strip()[:500]
    row.rd_share_note = (payload.rd_share_note or "").strip()
    row.is_active = bool(payload.is_active)
    row.updated_at = dt.datetime.now()
    write_system_audit(db, ctx["user"], "update_contract_item", "contract_item", str(row.id), f"更新明细: {gm}")
    db.commit()
    db.refresh(row)
    return _contract_item_dict(row)


@app.delete("/contract-items/{item_id}")
def delete_contract_item(
    item_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    row = db.get(ContractItem, item_id)
    if not row:
        raise HTTPException(status_code=404, detail="合同明细不存在")
    cid = row.contract_id
    db.delete(row)
    write_system_audit(db, ctx["user"], "delete_contract_item", "contract_item", str(item_id), f"合同{cid} 删除明细")
    db.commit()
    return {"id": item_id, "deleted": True}


def _pdf_extract_text(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail="服务端未安装 PDF 解析依赖 pypdf，无法解析。请在 backend 执行 pip install -r requirements.txt",
        ) from e
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 PDF：{e}") from e
    parts: list[str] = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
    return "\n".join(parts)


def _re_field(text: str, patterns: list[str]) -> str:
    for pat in patterns:
        m = re.search(pat, text, re.MULTILINE | re.DOTALL)
        if m:
            v = m.group(1).strip()
            v = re.sub(r"\s+", " ", v)
            if v:
                return v
    return ""


def _parse_contract_dates(text: str) -> tuple[Optional[str], Optional[str]]:
    m = re.search(
        r"自\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日\s*起?\s*至\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日止?",
        text,
    )
    if m:
        a = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        b = f"{m.group(4)}-{int(m.group(5)):02d}-{int(m.group(6)):02d}"
        return a, b
    dates: list[str] = []
    for g in re.finditer(r"(20\d{2}|19\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text):
        dates.append(f"{g.group(1)}-{int(g.group(2)):02d}-{int(g.group(3)):02d}")
    for g in re.finditer(r"(20\d{2}|19\d{2})[./-](\d{1,2})[./-](\d{1,2})", text):
        dates.append(f"{g.group(1)}-{int(g.group(2)):02d}-{int(g.group(3)):02d}")
    seen: set[str] = set()
    uniq: list[str] = []
    for d in dates:
        if d not in seen:
            seen.add(d)
            uniq.append(d)
    if len(uniq) >= 2:
        return uniq[0], uniq[1]
    if len(uniq) == 1:
        return uniq[0], None
    return None, None


def _clamp_pct(d: Decimal) -> Decimal:
    if d < 0:
        return Decimal("0")
    if d > 100:
        return Decimal("100")
    return d


def _try_parse_item_line(line: str) -> Optional[ContractDraftItemOut]:
    s = line.strip()
    if len(s) < 6 or s.startswith(("注：", "说明：", "附件：", "——", "--", "序号")):
        return None
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if len(nums) < 4:
        return None
    m = re.match(r"^([^\d:：]{2,28}?)\s*[:：]?\s*\d", s)
    if not m:
        m = re.match(r"^([^\d|｜\t]{2,28}?)[\s|｜\t]", s)
    if not m:
        return None
    game = re.sub(r"^[（(]?\d+[）).\s]+", "", m.group(1).strip())
    game = game.strip("·- ：:")
    if len(game) < 2:
        return None
    try:
        vals = [Decimal(x) for x in nums[:6]]
    except Exception:
        return None
    if len(vals) >= 6:
        dr, cs, cf, tx, pv = vals[0], vals[1], vals[2], vals[3], vals[4]
    elif len(vals) >= 5:
        dr, cs, cf, tx, pv = Decimal("0"), vals[0], vals[1], vals[2], vals[3]
    else:
        return None
    return ContractDraftItemOut(
        game_name=game,
        discount_label="",
        discount_rate=_clamp_pct(dr),
        channel_share_percent=_clamp_pct(cs),
        channel_fee_percent=_clamp_pct(cf),
        tax_percent=_clamp_pct(tx),
        private_percent=_clamp_pct(pv),
    )


def _parse_contract_items_heuristic(text: str) -> list[ContractDraftItemOut]:
    items: list[ContractDraftItemOut] = []
    for line in text.splitlines():
        it = _try_parse_item_line(line)
        if it and it.game_name:
            if items and items[-1].game_name == it.game_name:
                continue
            items.append(it)
    return items[:50]


def parse_contract_draft_from_pdf_bytes(data: bytes) -> ContractDraftParseOut:
    text = _pdf_extract_text(data)
    text = text.strip()
    base_remark = "由 PDF 自动识别生成，请人工核对后再保存。"
    if not text:
        return ContractDraftParseOut(
            platform_party_name="广州熊动科技有限公司",
            remark=base_remark + "（未提取到文本，可能为扫描件或加密 PDF。）",
            status="draft",
        )
    t = text
    contract_no = _re_field(
        t,
        [
            r"合同编号[：:\s]*([A-Za-z0-9\-_\.【】\[\]《》\u4e00-\u9fff]{2,50})",
            r"协议编号[：:\s]*([A-Za-z0-9\-_\.【】\[\]《》\u4e00-\u9fff]{2,50})",
        ],
    )
    contract_name = _re_field(
        t,
        [
            r"《\s*([^》\n]{2,80}?)》",
            r"合同名称[：:\s]*([^\n]{2,80})",
            r"协议名称[：:\s]*([^\n]{2,80})",
        ],
    )
    party_a = _re_field(
        t,
        [
            r"甲方[（(][^）)]*[）)]?\s*[：:]([^\n]{2,120})",
            r"甲方\s*[：:]\s*([^\n]{2,120})",
        ],
    )
    party_b = _re_field(
        t,
        [
            r"乙方[（(][^）)]*[）)]?\s*[：:]([^\n]{2,120})",
            r"乙方\s*[：:]\s*([^\n]{2,120})",
        ],
    )
    addr_a = _re_field(t, [r"甲方地址\s*[：:]\s*([^\n]+)", r"住所地\s*[：:]\s*([^\n]+)"])
    addr_b = _re_field(t, [r"乙方地址\s*[：:]\s*([^\n]+)"])
    channel = _re_field(t, [r"渠道(?:名称)?\s*[：:]\s*([^\n]{2,60})", r"合作方\s*[：:]\s*([^\n]{2,60})"])
    sd, ed = _parse_contract_dates(t)
    channel_name = (channel or party_b or "").strip()
    platform_party_name = (party_a or "广州熊动科技有限公司").strip() or "广州熊动科技有限公司"
    developer_party_name = (party_b or "").strip()
    items = _parse_contract_items_heuristic(t)
    return ContractDraftParseOut(
        contract_no=contract_no,
        contract_name=contract_name,
        channel_name=channel_name,
        platform_party_name=platform_party_name,
        platform_party_address=addr_a.strip(),
        developer_party_address=addr_b.strip(),
        developer_party_name=developer_party_name,
        start_date=sd,
        end_date=ed,
        status="draft",
        remark=base_remark,
        items=items,
    )


@app.post("/contracts/import-draft/parse", response_model=ContractDraftParseOut)
async def parse_contract_import_draft(
    file: UploadFile = File(...),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    name = (file.filename or "").lower()
    if not name.endswith(".pdf"):
        raise HTTPException(status_code=400, detail="请上传 PDF 文件")
    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过 10MB")
    if not raw:
        raise HTTPException(status_code=400, detail="空文件")
    return parse_contract_draft_from_pdf_bytes(raw)


_CONTRACT_EXCEL_SYNONYMS: dict[str, list[str]] = {
    "contract_no": ["合同编号", "协议编号", "编号"],
    "contract_name": ["合同名称", "合同名"],
    "contract_type": ["合同类型", "类型"],
    "sign_party": ["合同签约方", "签约方", "对方", "乙方", "客户名称", "合作方"],
    "channel_short": ["渠道简称", "渠道"],
    "start_date": ["签约日期", "开始日期", "生效日期", "起始日期"],
    "end_date": ["终止日期", "结束日期", "到期日", "截止日期"],
    "account_type": ["账款类型"],
    "attachment": ["合同附件", "附件"],
    "share_ratio": ["分成比例", "分成", "渠道分成", "分成%", "分成比例%"],
}


def _contract_excel_strip_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _contract_excel_find_column(df: pd.DataFrame, key: str) -> Optional[str]:
    syns = _CONTRACT_EXCEL_SYNONYMS.get(key, [])
    cols = [str(c).strip() for c in df.columns]
    colset = {c: c for c in cols}
    for s in syns:
        if s in colset:
            return s
    for s in syns:
        for c in cols:
            if s == c or (len(s) >= 2 and s in c):
                return c
    return None


def _contract_excel_cell_str(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    if isinstance(val, dt.datetime):
        return val.date().isoformat()
    if isinstance(val, dt.date):
        return val.isoformat()
    return str(val).strip()


def _contract_excel_parse_date(val: object) -> Optional[dt.date]:
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    try:
        if pd.api.types.is_scalar(val) and pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    # pd.Timestamp 是 datetime 子类；NaT 需先于 .date() 拦截，否则可能得到不可与 date 比较的 NaT
    if isinstance(val, pd.Timestamp):
        if pd.isna(val):
            return None
        return val.date()
    if isinstance(val, dt.datetime):
        try:
            if pd.isna(val):
                return None
        except TypeError:
            pass
        try:
            return val.date()
        except (ValueError, OSError):
            return None
    if isinstance(val, dt.date):
        return val
    s = str(val).strip()
    if not s or s.lower() in {"nat", "nan", "none"}:
        return None
    ts = pd.to_datetime(s, errors="coerce")
    if pd.isna(ts):
        return None
    try:
        out = ts.date()
    except (ValueError, OSError, OverflowError):
        return None
    try:
        if isinstance(out, dt.date) and pd.isna(out):
            return None
    except (TypeError, ValueError):
        pass
    return out if isinstance(out, dt.date) else None


def _contract_excel_read_frame(filename: str, raw: bytes) -> pd.DataFrame:
    low = (filename or "").lower()
    bio = io.BytesIO(raw)
    if low.endswith(".csv"):
        return _contract_excel_strip_cols(pd.read_csv(bio))
    if low.endswith(".xlsx"):
        return _contract_excel_strip_cols(pd.read_excel(bio, engine="openpyxl"))
    raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv")


def _contract_excel_build_remark(
    *,
    contract_type: str,
    account_type: str,
    attachment: str,
    share_ratio: str,
    start_note: str,
) -> str:
    parts: list[str] = []
    if contract_type:
        parts.append(f"合同类型(Excel): {contract_type}")
    if account_type:
        parts.append(f"账款类型: {account_type}")
    if attachment:
        parts.append(f"合同附件: {attachment}")
    if share_ratio:
        parts.append(f"分成比例(Excel): {share_ratio}")
    if start_note:
        parts.append(start_note)
    parts.append("来源: Excel 台账批量导入（草稿）")
    text = "\n".join(parts)
    return text[:1000]


def _contract_excel_preview_from_file(
    filename: str, raw: bytes, db: Session
) -> ContractExcelPreviewOut:
    df = _contract_excel_read_frame(filename, raw)
    if df.empty or len(df.columns) == 0:
        raise HTTPException(status_code=400, detail="表格为空或无法识别表头")

    def col(key: str) -> Optional[str]:
        return _contract_excel_find_column(df, key)

    c_name_col = col("contract_name")
    cno_col = col("contract_no")
    sign_col = col("sign_party")
    ch_col = col("channel_short")
    sd_col = col("start_date")
    ed_col = col("end_date")
    type_col = col("contract_type")
    acc_col = col("account_type")
    att_col = col("attachment")
    share_col = col("share_ratio")

    preview_rows: list[ContractExcelPreviewRow] = []
    triplet_counts: dict[tuple[str, str, str], int] = {}

    for i in range(len(df)):
        excel_row = i + 2
        row = df.iloc[i]
        def get(col_name: Optional[str]) -> str:
            if not col_name or col_name not in df.columns:
                return ""
            return _contract_excel_cell_str(row.get(col_name))

        contract_name = get(c_name_col).strip()
        contract_no_excel = get(cno_col).strip()
        sign_party = get(sign_col).strip()
        channel_short = get(ch_col).strip()
        channel_name = channel_short or sign_party
        developer_party_name = sign_party
        type_text = get(type_col).strip()
        account_type = get(acc_col).strip()
        attachment = get(att_col).strip()
        share_raw = get(share_col).strip()

        start_d = _contract_excel_parse_date(row.get(sd_col)) if sd_col else None
        end_d = _contract_excel_parse_date(row.get(ed_col)) if ed_col else None

        no_data = not contract_name and not channel_name and not sign_party and end_d is None
        if no_data:
            continue

        issues: list[str] = []
        status = "ready"
        if not contract_name:
            issues.append("合同名称不能为空")
            status = "skip"
        party_ok = bool((channel_short or sign_party).strip())
        if not party_ok:
            issues.append("对方/签约方不能为空（请填签约方或渠道简称）")
            status = "skip"
        if end_d is None:
            issues.append("终止日期缺失或无法解析")
            status = "skip"

        start_note = ""
        if sd_col and row.get(sd_col) is not None and str(row.get(sd_col)).strip():
            raw_sd = _contract_excel_cell_str(row.get(sd_col))
            if start_d is None:
                start_note = f"签约日期(原始): {raw_sd}（未能解析为日期，开始日期待人工核对）"
            else:
                start_note = f"签约日期(Excel列): {raw_sd}"
        if start_d is None:
            start_d = dt.date.today()

        if status != "skip" and end_d is not None and end_d < start_d:
            issues.append("结束日期早于开始日期")
            status = "skip"

        remark = _contract_excel_build_remark(
            contract_type=type_text,
            account_type=account_type,
            attachment=attachment,
            share_ratio=share_raw,
            start_note=start_note,
        )

        imp_id = uuid.uuid4().hex[:8].upper()
        proposed_no = (contract_no_excel or "").strip()
        if not proposed_no:
            proposed_no = f"IMP-{dt.date.today().strftime('%Y%m%d')}-{excel_row:04d}-{imp_id}"
        elif len(proposed_no) > 75:
            proposed_no = proposed_no[:75]

        dup_key = (
            contract_name.strip(),
            channel_name.strip(),
            end_d.isoformat() if end_d else "",
        )
        if status != "skip" and dup_key[2]:
            triplet_counts[dup_key] = triplet_counts.get(dup_key, 0) + 1

        preview_rows.append(
            ContractExcelPreviewRow(
                excel_row=excel_row,
                contract_no=proposed_no,
                contract_name=contract_name,
                channel_name=channel_name.strip(),
                developer_party_name=developer_party_name,
                platform_party_name="广州熊动科技有限公司",
                start_date=start_d.isoformat() if start_d else None,
                end_date=end_d.isoformat() if end_d else None,
                attachment_preview=(attachment[:200] + "…") if len(attachment) > 200 else attachment,
                remark=remark,
                status=status,
                issues=issues,
                duplicate_hint=False,
            )
        )

    # 文件内重复
    for r in preview_rows:
        if r.status == "skip" or not r.end_date:
            continue
        key = (r.contract_name.strip(), r.channel_name.strip(), r.end_date)
        if triplet_counts.get(key, 0) > 1:
            r.duplicate_hint = True
            r.status = "warn"
            if "疑似重复：本文件内存在相同合同名称+渠道+终止日" not in r.issues:
                r.issues.append("疑似重复：本文件内存在相同合同名称+渠道+终止日")

    # 库内重复
    if preview_rows:
        ready_triples = [
            (r.contract_name.strip(), r.channel_name.strip(), r.end_date)
            for r in preview_rows
            if r.status != "skip" and r.end_date
        ]
        if ready_triples:
            for r in preview_rows:
                if r.status == "skip" or not r.end_date:
                    continue
                try:
                    ed = dt.date.fromisoformat(r.end_date)
                except ValueError:
                    continue
                exists = db.scalar(
                    select(func.count())
                    .select_from(ContractHeader)
                    .where(
                        ContractHeader.contract_name == r.contract_name.strip(),
                        ContractHeader.channel_name == r.channel_name.strip(),
                        ContractHeader.end_date == ed,
                    )
                )
                if exists and exists > 0:
                    r.duplicate_hint = True
                    r.status = "warn"
                    msg = "疑似重复：系统中已存在相同合同名称+渠道+终止日"
                    if msg not in r.issues:
                        r.issues.append(msg)

    return ContractExcelPreviewOut(rows=preview_rows, file_label=filename or "")


@app.post("/contracts/import-excel/preview", response_model=ContractExcelPreviewOut)
async def contract_excel_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    name = (file.filename or "").lower()
    if not name.endswith((".xlsx", ".csv")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .csv")
    raw = await file.read()
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过 20MB")
    if not raw:
        raise HTTPException(status_code=400, detail="空文件")
    return _contract_excel_preview_from_file(file.filename or "upload", raw, db)


@app.post("/contracts/import-excel/commit", response_model=ContractExcelCommitOut)
def contract_excel_commit(
    payload: ContractExcelCommitIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    created = 0
    skipped = 0
    created_ids: list[int] = []
    skip_reasons: list[str] = []
    for it in payload.items:
        cn = (it.contract_no or "").strip()
        cname = (it.contract_name or "").strip()
        ch = (it.channel_name or "").strip()
        if not cname or not ch:
            skipped += 1
            skip_reasons.append(f"{cn or '(无编号)'}: 名称或渠道为空")
            continue
        if not cn:
            skipped += 1
            skip_reasons.append(f"{cname}: 合同编号为空")
            continue
        try:
            sd = dt.date.fromisoformat((it.start_date or "").strip())
            ed = dt.date.fromisoformat((it.end_date or "").strip())
        except ValueError:
            skipped += 1
            skip_reasons.append(f"{cn}: 日期格式无效")
            continue
        if ed < sd:
            skipped += 1
            skip_reasons.append(f"{cn}: 结束日期早于开始日期")
            continue
        if db.scalar(select(ContractHeader).where(ContractHeader.contract_no == cn)):
            skipped += 1
            skip_reasons.append(f"{cn}: 合同编号已存在")
            continue
        row = ContractHeader(
            contract_no=cn,
            contract_name=cname or cn,
            channel_name=ch,
            platform_party_name=(it.platform_party_name or "广州熊动科技有限公司").strip() or "广州熊动科技有限公司",
            platform_party_address=(it.platform_party_address or "").strip(),
            developer_party_name=(it.developer_party_name or "").strip(),
            developer_party_address=(it.developer_party_address or "").strip(),
            start_date=sd,
            end_date=ed,
            status=ContractStatus.draft,
            remark=(it.remark or "").strip()[:1000],
        )
        db.add(row)
        db.flush()
        created_ids.append(row.id)
        write_system_audit(
            db,
            ctx["user"],
            "create_contract",
            "contract_header",
            str(row.id),
            f"Excel导入草稿: {row.contract_no}",
        )
        created += 1
    db.commit()
    return ContractExcelCommitOut(created=created, skipped=skipped, created_ids=created_ids, skip_reasons=skip_reasons)


@app.post("/games")
def create_game(payload: GameIn, db: Session = Depends(get_db), ctx: dict = Depends(require_role([Role.admin, Role.biz, Role.ops]))):
    if payload.rd_share_percent < 0 or payload.rd_share_percent > 100:
        raise HTTPException(status_code=400, detail="研发分成需在 0~100 之间")
    gc = (payload.game_code or "").strip() or None
    game = Game(
        name=payload.name,
        rd_company=payload.rd_company,
        rd_share_percent=payload.rd_share_percent,
        game_code=gc,
    )
    db.add(game)
    db.flush()
    write_system_audit(db, ctx["user"], "create_game", "game", str(game.id), f"新增游戏: {payload.name}")
    db.commit()
    return {"id": game.id, "name": game.name}


@app.post("/games/bulk-create")
def bulk_create_games(
    payload: GameBulkCreateIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.biz, Role.ops])),
):
    success_count = 0
    failed_names: list[str] = []
    cleaned_names: list[str] = []
    seen = set()
    for raw in payload.names:
        name = (raw or "").strip()
        if not name:
            continue
        if name in seen:
            continue
        seen.add(name)
        cleaned_names.append(name)
    exists = {x.name for x in db.scalars(select(Game).where(Game.name.in_(cleaned_names))).all()} if cleaned_names else set()
    for name in cleaned_names:
        if name in exists:
            failed_names.append(name)
            continue
        db.add(Game(name=name, rd_company="待补充", rd_share_percent=Decimal("0")))
        success_count += 1
    write_system_audit(
        db,
        ctx["user"],
        "bulk_create_games",
        "game",
        "",
        f"批量新增游戏: 成功{success_count}, 跳过{len(failed_names)}",
    )
    db.commit()
    return {"success_count": success_count, "failed_count": len(failed_names), "failed_names": failed_names}


@app.get("/games")
def list_games(
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.finance, Role.tech, Role.biz, Role.ops])),
):
    return db.scalars(select(Game).order_by(Game.id.desc())).all()


@app.put("/games/{game_id}")
def update_game(game_id: int, payload: GameIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.biz]))):
    row = db.get(Game, game_id)
    if not row:
        raise HTTPException(status_code=404, detail="游戏不存在")
    if payload.rd_share_percent < 0 or payload.rd_share_percent > 100:
        raise HTTPException(status_code=400, detail="研发分成需在 0~100 之间")
    row.name = payload.name
    row.rd_company = payload.rd_company
    row.rd_share_percent = payload.rd_share_percent
    row.game_code = (payload.game_code or "").strip() or None
    write_system_audit(db, _["user"], "update_game", "game", str(row.id), f"编辑游戏: {payload.name}")
    db.commit()
    return {"id": row.id, "name": row.name}


@app.delete("/games/{game_id}")
def delete_game(game_id: int, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin]))):
    row = db.get(Game, game_id)
    if not row:
        raise HTTPException(status_code=404, detail="游戏不存在")
    write_system_audit(db, _["user"], "delete_game", "game", str(row.id), f"删除游戏: {row.name}")
    db.delete(row)
    db.commit()
    return {"id": game_id, "deleted": True}


@app.get("/projects")
def list_projects(db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops]))):
    return db.scalars(select(Project).order_by(Project.id.desc())).all()


@app.post("/projects")
def create_project(payload: ProjectIn, db: Session = Depends(get_db), ctx: dict = Depends(require_role([Role.admin, Role.biz, Role.ops]))):
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="项目名称不能为空")
    dup = db.scalar(select(Project).where(Project.name == name))
    if dup:
        raise HTTPException(status_code=400, detail="项目名称已存在")
    row = Project(name=name, status=payload.status, remark=(payload.remark or "").strip())
    db.add(row)
    db.flush()
    write_system_audit(db, ctx["user"], "create_project", "project", str(row.id), f"新增项目: {name}")
    db.commit()
    return row


@app.put("/projects/{project_id}")
def update_project(
    project_id: int,
    payload: ProjectIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.biz])),
):
    row = db.get(Project, project_id)
    if not row:
        raise HTTPException(status_code=404, detail="项目不存在")
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="项目名称不能为空")
    dup = db.scalar(select(Project).where(Project.name == name, Project.id != project_id))
    if dup:
        raise HTTPException(status_code=400, detail="项目名称已存在")
    row.name = name
    row.status = payload.status
    row.remark = (payload.remark or "").strip()
    write_system_audit(db, ctx["user"], "update_project", "project", str(row.id), f"编辑项目: {name}")
    db.commit()
    return row


@app.patch("/projects/{project_id}/status")
def patch_project_status(
    project_id: int,
    payload: ProjectStatusPatch,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.biz])),
):
    row = db.get(Project, project_id)
    if not row:
        raise HTTPException(status_code=404, detail="项目不存在")
    row.status = payload.status
    write_system_audit(db, ctx["user"], "update_project_status", "project", str(row.id), f"项目状态: {payload.status.value}")
    db.commit()
    return row


@app.get("/game-variants")
def list_game_variants(
    project_id: Optional[int] = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    stmt = select(GameVariant).order_by(GameVariant.id.desc())
    if project_id is not None:
        stmt = stmt.where(GameVariant.project_id == project_id)
    return db.scalars(stmt).all()


@app.post("/game-variants")
def create_game_variant(payload: GameVariantIn, db: Session = Depends(get_db), ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech]))):
    if not db.get(Project, payload.project_id):
        raise HTTPException(status_code=400, detail="所属项目不存在")
    raw = (payload.raw_game_name or "").strip()
    variant = (payload.variant_name or "").strip()
    if not raw or not variant:
        raise HTTPException(status_code=400, detail="版本名称与原始游戏名不能为空")
    dup = db.scalar(select(GameVariant).where(GameVariant.raw_game_name == raw))
    if dup:
        raise HTTPException(status_code=400, detail="原始游戏名已存在")
    rd_share, publish_share = normalize_variant_shares(payload)
    row = GameVariant(
        project_id=payload.project_id,
        variant_name=variant,
        raw_game_name=raw,
        discount_type=payload.discount_type,
        version_type=payload.version_type,
        server_type=payload.server_type,
        status=payload.status,
        remark=(payload.remark or "").strip(),
        rd_company=(payload.rd_company or "").strip() or None,
        publish_company=(payload.publish_company or "广州熊动科技有限公司").strip() or "广州熊动科技有限公司",
        rd_share_percent=rd_share,
        publish_share_percent=publish_share,
        settlement_remark=(payload.settlement_remark or "").strip() or None,
    )
    db.add(row)
    db.flush()
    write_system_audit(db, ctx["user"], "create_game_variant", "game_variant", str(row.id), f"新增版本: {variant} / {raw}")
    db.commit()
    return row


@app.put("/game-variants/{variant_id}")
def update_game_variant(
    variant_id: int,
    payload: GameVariantIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    row = db.get(GameVariant, variant_id)
    if not row:
        raise HTTPException(status_code=404, detail="版本不存在")
    if not db.get(Project, payload.project_id):
        raise HTTPException(status_code=400, detail="所属项目不存在")
    raw = (payload.raw_game_name or "").strip()
    variant = (payload.variant_name or "").strip()
    if not raw or not variant:
        raise HTTPException(status_code=400, detail="版本名称与原始游戏名不能为空")
    dup = db.scalar(select(GameVariant).where(GameVariant.raw_game_name == raw, GameVariant.id != variant_id))
    if dup:
        raise HTTPException(status_code=400, detail="原始游戏名已存在")
    rd_share, publish_share = normalize_variant_shares(payload)
    row.project_id = payload.project_id
    row.variant_name = variant
    row.raw_game_name = raw
    row.discount_type = payload.discount_type
    row.version_type = payload.version_type
    row.server_type = payload.server_type
    row.status = payload.status
    row.remark = (payload.remark or "").strip()
    row.rd_company = (payload.rd_company or "").strip() or None
    row.publish_company = (payload.publish_company or "广州熊动科技有限公司").strip() or "广州熊动科技有限公司"
    row.rd_share_percent = rd_share
    row.publish_share_percent = publish_share
    row.settlement_remark = (payload.settlement_remark or "").strip() or None
    write_system_audit(db, ctx["user"], "update_game_variant", "game_variant", str(row.id), f"编辑版本: {variant}")
    db.commit()
    return row


@app.patch("/game-variants/{variant_id}/status")
def patch_variant_status(
    variant_id: int,
    payload: VariantStatusPatch,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    row = db.get(GameVariant, variant_id)
    if not row:
        raise HTTPException(status_code=404, detail="版本不存在")
    row.status = payload.status
    write_system_audit(db, ctx["user"], "update_game_variant_status", "game_variant", str(row.id), f"版本状态: {payload.status.value}")
    db.commit()
    return row


@app.post("/channel-game-map")
def create_map(payload: MapIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech]))):
    exists = db.scalar(
        select(ChannelGameMap).where(ChannelGameMap.channel_id == payload.channel_id, ChannelGameMap.game_id == payload.game_id)
    )
    if exists:
        raise HTTPException(status_code=400, detail="关系已存在")
    item = ChannelGameMap(**payload.model_dump())
    db.add(item)
    db.flush()
    write_system_audit(db, _["user"], "create_channel_game_map", "channel_game_map", str(item.id), "新增渠道游戏映射")
    db.commit()
    return {"id": item.id}


@app.post("/channel-game-map/bulk-create")
def bulk_create_map(
    payload: MapBulkCreateIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    success_count = 0
    failed_items = []
    seen = set()
    for raw in payload.items:
        channel_name = (raw.channel_name or "").strip()
        game_name = (raw.game_name or "").strip()
        key = f"{channel_name}::{game_name}"
        if not channel_name or not game_name:
            failed_items.append({"channel_name": channel_name, "game_name": game_name, "reason": "格式错误"})
            continue
        if key in seen:
            failed_items.append({"channel_name": channel_name, "game_name": game_name, "reason": "重复输入"})
            continue
        seen.add(key)
        channel = db.scalar(select(Channel).where(Channel.name == channel_name))
        if not channel:
            failed_items.append({"channel_name": channel_name, "game_name": game_name, "reason": "渠道不存在"})
            continue
        game = db.scalar(select(Game).where(Game.name == game_name))
        if not game:
            failed_items.append({"channel_name": channel_name, "game_name": game_name, "reason": "游戏不存在"})
            continue
        exists = db.scalar(select(ChannelGameMap).where(ChannelGameMap.channel_id == channel.id, ChannelGameMap.game_id == game.id))
        if exists:
            failed_items.append({"channel_name": channel_name, "game_name": game_name, "reason": "映射已存在"})
            continue
        db.add(ChannelGameMap(channel_id=channel.id, game_id=game.id))
        success_count += 1
    write_system_audit(
        db,
        ctx["user"],
        "bulk_create_channel_game_map",
        "channel_game_map",
        "",
        f"批量新增映射: 成功{success_count}, 跳过{len(failed_items)}",
    )
    db.commit()
    return {"success_count": success_count, "failed_count": len(failed_items), "failed_items": failed_items}


@app.get("/channel-game-map")
def list_map(db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech]))):
    rows = db.scalars(select(ChannelGameMap).order_by(ChannelGameMap.id.desc())).all()
    return [
        {
            "id": x.id,
            "channel": x.channel.name,
            "game": x.game.name,
            "revenue_share_ratio": x.revenue_share_ratio,
            "rd_settlement_ratio": x.rd_settlement_ratio,
        }
        for x in rows
    ]


@app.put("/channel-game-map/{map_id}")
def update_map(map_id: int, payload: MapIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech]))):
    row = db.get(ChannelGameMap, map_id)
    if not row:
        raise HTTPException(status_code=404, detail="映射不存在")
    row.channel_id = payload.channel_id
    row.game_id = payload.game_id
    row.revenue_share_ratio = payload.revenue_share_ratio
    row.rd_settlement_ratio = payload.rd_settlement_ratio
    write_system_audit(db, _["user"], "update_channel_game_map", "channel_game_map", str(row.id), "编辑渠道游戏映射")
    db.commit()
    return {"id": row.id}


@app.delete("/channel-game-map/{map_id}")
def delete_map(map_id: int, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin]))):
    row = db.get(ChannelGameMap, map_id)
    if not row:
        raise HTTPException(status_code=404, detail="映射不存在")
    write_system_audit(db, _["user"], "delete_channel_game_map", "channel_game_map", str(row.id), "删除渠道游戏映射")
    db.delete(row)
    db.commit()
    return {"id": map_id, "deleted": True}


@app.post("/recon/import")
async def import_statement(
    period: str = Query(..., description="账期，例如 2026-03"),
    import_type: str = Query(default="template"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.tech])),
):
    content = await file.read()
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(status_code=400, detail={"code": 400, "message": "仅支持 CSV / XLSX 文件"})
    suffix = ".csv" if filename.endswith(".csv") else ".xlsx"
    with tempfile.NamedTemporaryFile(prefix="recon_import_", suffix=suffix, dir="/tmp", delete=False) as f:
        f.write(content)
        tmp = f.name
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(tmp)
        else:
            wb = load_workbook(tmp, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(x).strip() if x is not None else "" for x in next(rows_iter, [])]
            rows_data = []
            for row in rows_iter:
                if row is None:
                    continue
                values = list(row)
                if all(v is None or str(v).strip() == "" for v in values):
                    continue
                rows_data.append({headers[i] if i < len(headers) else f"col_{i}": values[i] for i in range(len(values))})
            df = pd.DataFrame(rows_data)
            wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail={"code": 400, "message": "仅支持 CSV / XLSX 文件"}) from e
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    required = {"channel_name", "game_name", "gross_amount"}
    if not required.issubset(set(df.columns)):
        raise HTTPException(status_code=400, detail="缺少字段: channel_name, game_name, gross_amount")
    task = ReconTask(period=period, status=ReconStatus.pending)
    db.add(task)
    db.flush()
    issue_count = 0
    matched_variant_count = 0
    unmatched_variant_count = 0
    total_count = int(len(df.index))
    amount_sum = Decimal("0")
    projects = {x.id: x for x in db.scalars(select(Project)).all()}
    variants = {x.raw_game_name: x for x in db.scalars(select(GameVariant)).all()}
    for _, row in df.iterrows():
        channel_name = str(row["channel_name"]).strip()
        game_name = str(row["game_name"]).strip()
        gross_amount = Decimal(str(row["gross_amount"]))
        amount_sum += gross_amount
        matched_variant = variants.get(game_name)
        matched_project = projects.get(matched_variant.project_id) if matched_variant else None
        if matched_variant:
            matched_variant_count += 1
        else:
            unmatched_variant_count += 1
        db.add(
            RawStatement(
                recon_task_id=task.id,
                channel_name=channel_name,
                game_name=game_name,
                period=period,
                gross_amount=gross_amount,
                project_id=matched_project.id if matched_project else None,
                project_name=matched_project.name if matched_project else None,
                variant_id=matched_variant.id if matched_variant else None,
                variant_name=matched_variant.variant_name if matched_variant else None,
                rd_company=matched_variant.rd_company if matched_variant else None,
                publish_company=matched_variant.publish_company if matched_variant else None,
                rd_share_percent=matched_variant.rd_share_percent if matched_variant else None,
                publish_share_percent=matched_variant.publish_share_percent if matched_variant else None,
                variant_match_status="已匹配版本" if matched_variant else "未匹配版本",
            )
        )
        channel = db.scalar(select(Channel).where(Channel.name == channel_name))
        game = db.scalar(select(Game).where(Game.name == game_name))
        if channel is None or game is None:
            db.add(
                ReconIssue(
                    recon_task_id=task.id,
                    issue_type="master_data",
                    detail=f"渠道或游戏不存在: {channel_name}/{game_name}",
                )
            )
            issue_count += 1
            continue
        link = db.scalar(
            select(ChannelGameMap).where(ChannelGameMap.channel_id == channel.id, ChannelGameMap.game_id == game.id)
        )
        if link is None:
            db.add(
                ReconIssue(
                    recon_task_id=task.id,
                    issue_type="mapping",
                    detail=f"渠道游戏映射不存在: {channel_name}/{game_name}",
                )
            )
            issue_count += 1
    task.status = ReconStatus.issue if issue_count > 0 else ReconStatus.pending
    valid_count = max(total_count - issue_count, 0)
    summary_text = f"总行数:{total_count}, 正常:{valid_count}, 异常:{issue_count}, 流水合计:{amount_sum}"
    db.add(
        ImportHistory(
            import_type=import_type,
            period=period,
            file_name=file.filename or "",
            task_id=task.id,
            total_count=total_count,
            valid_count=valid_count,
            invalid_count=issue_count,
            amount_sum=amount_sum,
            status="异常待处理" if issue_count > 0 else "待确认",
            matched_variant_count=matched_variant_count,
            unmatched_variant_count=unmatched_variant_count,
            lifecycle_status="active",
            summary=summary_text,
            created_by=ctx.get("user", "system"),
        )
    )
    write_system_audit(db, ctx["user"], "import_data", "recon_task", str(task.id), f"导入数据: {file.filename or 'unknown'}")
    db.commit()
    return {
        "recon_task_id": task.id,
        "issue_count": issue_count,
        "summary": {
            "total_count": total_count,
            "valid_count": valid_count,
            "invalid_count": issue_count,
            "amount_sum": amount_sum,
            "matched_variant_count": matched_variant_count,
            "unmatched_variant_count": unmatched_variant_count,
        },
    }


@app.post("/recon/{task_id}/confirm")
def confirm_recon(task_id: int, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    task = db.get(ReconTask, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    history = db.scalar(select(ImportHistory).where(ImportHistory.task_id == task_id).order_by(ImportHistory.id.desc()).limit(1))
    if history and (history.lifecycle_status or "active") == "discarded":
        raise HTTPException(status_code=400, detail="该批次已作废，不能确认入账")
    unresolved = db.scalar(select(func.count(ReconIssue.id)).where(ReconIssue.recon_task_id == task_id, ReconIssue.resolved.is_(False)))
    if unresolved > 0:
        raise HTTPException(status_code=400, detail="仍有异常未处理")
    for h in db.scalars(select(ImportHistory).where(ImportHistory.period == task.period)).all():
        if h.task_id == task_id:
            continue
        if (h.lifecycle_status or "active") != "active":
            continue
        other = db.get(ReconTask, h.task_id)
        if other and other.status == ReconStatus.confirmed:
            raise HTTPException(status_code=400, detail=BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS)
    task.status = ReconStatus.confirmed
    write_system_audit(db, _["user"], "confirm_recon_period", "recon_task", str(task.id), f"确认账期: {task.period}")
    db.commit()
    return {"task_id": task_id, "status": task.status}


@app.post("/recon/{task_id}/revert-confirm")
def revert_confirm_recon(
    task_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance])),
):
    task = db.get(ReconTask, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    if task.status != ReconStatus.confirmed:
        raise HTTPException(status_code=400, detail="仅已确认任务允许撤销")
    history = db.scalar(select(ImportHistory).where(ImportHistory.task_id == task_id).order_by(ImportHistory.id.desc()).limit(1))
    if history and (history.lifecycle_status or "active") == "discarded":
        raise HTTPException(status_code=400, detail="已作废批次不可撤销入账")

    period = task.period
    bill_cnt = int(
        db.scalar(select(func.count(Bill.id)).where(Bill.period == period, Bill.lifecycle_status == "active")) or 0
    )
    if bill_cnt > 0:
        raise HTTPException(status_code=400, detail="该账期已生成账单，禁止撤销入账")

    stmt_cnt = int(db.scalar(select(func.count(ChannelSettlementStatement.id)).where(ChannelSettlementStatement.period == period)) or 0)
    if stmt_cnt > 0:
        raise HTTPException(status_code=400, detail="该账期已生成结算单，禁止撤销入账")

    inv_cnt = int(
        db.scalar(
            select(func.count(Invoice.id))
            .select_from(Invoice)
            .join(Bill, Bill.id == Invoice.bill_id)
            .where(Bill.period == period)
        )
        or 0
    )
    if inv_cnt > 0:
        raise HTTPException(status_code=400, detail="该账期已有发票，禁止撤销入账")

    rec_cnt = int(
        db.scalar(
            select(func.count(Receipt.id))
            .select_from(Receipt)
            .join(Bill, Bill.id == Receipt.bill_id)
            .where(Bill.period == period)
        )
        or 0
    )
    if rec_cnt > 0:
        raise HTTPException(status_code=400, detail="该账期已有回款，禁止撤销入账")

    task.status = ReconStatus.pending
    if history:
        history.status = "待确认"
    write_system_audit(db, ctx["user"], "revert_confirm_recon_period", "recon_task", str(task.id), f"撤销确认账期: {task.period}")
    db.commit()
    return {"task_id": task_id, "status": task.status}


@app.post("/recon/issues/{issue_id}/resolve")
def resolve_issue(
    issue_id: int,
    payload: ResolveIssueIn,
    db: Session = Depends(get_db),
    operator_ctx: dict = Depends(require_role([Role.admin, Role.finance, Role.ops])),
):
    issue = db.get(ReconIssue, issue_id)
    if not issue:
        raise HTTPException(status_code=404, detail="异常不存在")
    from_status = "已处理" if issue.resolved else "未处理"
    issue.resolved = payload.status in {"已处理", "resolved", "true", "1"}
    to_status = "已处理" if issue.resolved else "未处理"
    meta = db.scalar(select(ReconIssueMeta).where(ReconIssueMeta.issue_id == issue.id))
    if meta:
        meta.remark = payload.remark
        meta.updated_at = dt.datetime.now()
    else:
        db.add(ReconIssueMeta(issue_id=issue.id, remark=payload.remark, updated_at=dt.datetime.now()))
    db.add(
        ReconIssueTimeline(
            issue_id=issue.id,
            action="resolve",
            from_status=from_status,
            to_status=to_status,
            remark=payload.remark,
            operator=operator_ctx.get("user", "system"),
            created_at=dt.datetime.now(),
        )
    )
    db.commit()
    return {"issue_id": issue.id, "resolved": issue.resolved, "status": "已处理" if issue.resolved else "未处理", "remark": payload.remark}


@app.post("/recon/issues/bulk-resolve")
def bulk_resolve_issues(
    payload: BulkResolveIssuesIn,
    db: Session = Depends(get_db),
    operator_ctx: dict = Depends(require_role([Role.admin, Role.finance, Role.ops])),
):
    success_count = 0
    failed_ids = []
    for issue_id in payload.issue_ids:
        issue = db.get(ReconIssue, issue_id)
        if not issue:
            failed_ids.append(issue_id)
            continue
        from_status = "已处理" if issue.resolved else "未处理"
        issue.resolved = True
        to_status = "已处理"
        meta = db.scalar(select(ReconIssueMeta).where(ReconIssueMeta.issue_id == issue.id))
        if meta:
            meta.remark = payload.remark
            meta.updated_at = dt.datetime.now()
        else:
            db.add(ReconIssueMeta(issue_id=issue.id, remark=payload.remark, updated_at=dt.datetime.now()))
        db.add(
            ReconIssueTimeline(
                issue_id=issue.id,
                action="bulk_resolve",
                from_status=from_status,
                to_status=to_status,
                remark=payload.remark,
                operator=operator_ctx.get("user", "system"),
                created_at=dt.datetime.now(),
            )
        )
        success_count += 1
    db.commit()
    return {"success_count": success_count, "failed_count": len(failed_ids), "failed_ids": failed_ids}


@app.get("/recon/tasks")
def list_recon(db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops]))):
    tasks = db.scalars(select(ReconTask).order_by(ReconTask.id.desc())).all()
    return [{"id": x.id, "period": x.period, "status": x.status} for x in tasks]


@app.get("/recon/issues")
def list_recon_issues(task_id: int = Query(...), db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance, Role.ops]))):
    rows = db.scalars(select(ReconIssue).where(ReconIssue.recon_task_id == task_id).order_by(ReconIssue.id.desc())).all()
    result = []
    for x in rows:
        meta = db.scalar(select(ReconIssueMeta).where(ReconIssueMeta.issue_id == x.id))
        latest_timeline = db.scalar(select(ReconIssueTimeline).where(ReconIssueTimeline.issue_id == x.id).order_by(ReconIssueTimeline.id.desc()).limit(1))
        result.append(
            {
                "issue_id": x.id,
                "task_id": x.recon_task_id,
                "issue_type": x.issue_type,
                "message": x.detail,
                "status": "已处理" if x.resolved else "未处理",
                "row_no": None,
                "raw_data": None,
                "created_at": "",
                "remark": meta.remark if meta else "",
                "updated_at": str(meta.updated_at) if meta else "",
                "latest_operator": latest_timeline.operator if latest_timeline else "",
                "latest_updated_at": str(latest_timeline.created_at) if latest_timeline else "",
            }
        )
    return result


@app.get("/recon/issues/{issue_id}/timeline")
def get_issue_timeline(
    issue_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.ops, Role.biz])),
):
    issue = db.get(ReconIssue, issue_id)
    if not issue:
        raise HTTPException(status_code=404, detail="异常不存在")
    rows = db.scalars(select(ReconIssueTimeline).where(ReconIssueTimeline.issue_id == issue_id).order_by(ReconIssueTimeline.id.asc())).all()
    return [
        {
            "id": x.id,
            "issue_id": x.issue_id,
            "action": x.action,
            "from_status": x.from_status,
            "to_status": x.to_status,
            "remark": x.remark,
            "operator": x.operator,
            "created_at": str(x.created_at),
        }
        for x in rows
    ]


@app.post("/billing/rules")
def create_rule(payload: RuleIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    rule = BillingRule(**payload.model_dump())
    db.add(rule)
    db.flush()
    write_system_audit(db, _["user"], "create_billing_rule", "billing_rule", str(rule.id), f"新增规则: {rule.name}")
    db.commit()
    return {"id": rule.id}


@app.get("/billing/rules")
def list_rules(
    channel: Optional[str] = None,
    game: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz])),
):
    rows = db.scalars(select(BillingRule).order_by(BillingRule.id.desc())).all()
    if channel:
        rows = [x for x in rows if channel in x.name]
    if game:
        rows = [x for x in rows if game in x.name]
    return [
        {
            "id": x.id,
            "name": x.name,
            "bill_type": x.bill_type.value if isinstance(x.bill_type, BillType) else str(x.bill_type),
            "default_ratio": float(x.default_ratio) if x.default_ratio is not None else None,
            "active": x.active,
        }
        for x in rows
    ]


@app.put("/billing/rules/{rule_id}")
def update_rule(rule_id: int, payload: RuleIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    row = db.get(BillingRule, rule_id)
    if not row:
        raise HTTPException(status_code=404, detail="规则不存在")
    row.name = payload.name
    row.bill_type = payload.bill_type
    row.default_ratio = payload.default_ratio
    write_system_audit(db, _["user"], "update_billing_rule", "billing_rule", str(row.id), f"编辑规则: {row.name}")
    db.commit()
    return {"id": row.id}


@app.post("/billing/rules/bulk-import")
def bulk_import_rules(payload: RuleBulkIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    created_count = 0
    updated_count = 0
    failed_count = 0
    error_details = []
    channel_set = {x.name for x in db.scalars(select(Channel)).all()}
    game_set = {x.name for x in db.scalars(select(Game)).all()}
    valid_discount = {"无", "0.1折", "0.05折"}
    valid_status = {"启用", "停用", "enabled", "disabled", "true", "false", "1", "0"}
    for idx, row in enumerate(payload.rows):
        try:
            err_fields = []
            err_msgs = []
            if not row.game:
                err_fields.append("game")
                err_msgs.append("游戏为空")
            if not row.channel:
                err_fields.append("channel")
                err_msgs.append("渠道为空")
            if row.game and row.game not in game_set:
                err_fields.append("game")
                err_msgs.append("游戏不存在")
            if row.channel and row.channel not in channel_set:
                err_fields.append("channel")
                err_msgs.append("渠道不存在")
            if row.discount_type not in valid_discount:
                err_fields.append("discount_type")
                err_msgs.append("折扣类型非法")
            if row.status not in valid_status:
                err_fields.append("status")
                err_msgs.append("状态非法")
            for val, name, field_name in [
                (row.channel_fee, "通道费格式非法", "channel_fee"),
                (row.tax_rate, "税点格式非法", "tax_rate"),
                (row.rd_share, "研发分成格式非法", "rd_share"),
                (row.private_rate, "私点格式非法", "private_rate"),
            ]:
                if val is None:
                    err_fields.append(field_name)
                    err_msgs.append(name)
            if err_msgs:
                failed_count += 1
                error_details.append(
                    {
                        "row_no": row.row_no or idx + 2,
                        "raw_data": row.model_dump(),
                        "error_fields": err_fields,
                        "error_message": ";".join(err_msgs),
                    }
                )
                continue
            name = f"{row.channel}-{row.game}-rule"
            existing = db.scalar(select(BillingRule).where(BillingRule.name == name))
            if existing:
                existing.default_ratio = row.rd_share
                existing.active = row.status in {"启用", "enabled", "true", "1"}
                updated_count += 1
            else:
                db.add(
                    BillingRule(
                        name=name,
                        bill_type=BillType.channel,
                        default_ratio=row.rd_share,
                        active=row.status in {"启用", "enabled", "true", "1"},
                    )
                )
                created_count += 1
        except Exception:
            failed_count += 1
    write_system_audit(
        db,
        _["user"],
        "bulk_import_billing_rules",
        "billing_rule",
        "",
        f"规则批量导入: 新增{created_count}, 更新{updated_count}, 失败{failed_count}",
    )
    db.commit()
    return {
        "created_count": created_count,
        "updated_count": updated_count,
        "failed_count": failed_count,
        "error_details": error_details,
    }


@app.post("/billing/rules/bulk-validate")
def bulk_validate_rules(payload: RuleBulkIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    channel_set = {x.name for x in db.scalars(select(Channel)).all()}
    game_set = {x.name for x in db.scalars(select(Game)).all()}
    valid_discount = {"无", "0.1折", "0.05折"}
    valid_status = {"启用", "停用", "enabled", "disabled", "true", "false", "1", "0"}
    error_details = []
    for idx, row in enumerate(payload.rows):
        err_fields = []
        err_msgs = []
        if not row.game:
            err_fields.append("game")
            err_msgs.append("游戏为空")
        if not row.channel:
            err_fields.append("channel")
            err_msgs.append("渠道为空")
        if row.game and row.game not in game_set:
            err_fields.append("game")
            err_msgs.append("游戏不存在")
        if row.channel and row.channel not in channel_set:
            err_fields.append("channel")
            err_msgs.append("渠道不存在")
        if row.discount_type not in valid_discount:
            err_fields.append("discount_type")
            err_msgs.append("折扣类型非法")
        if row.status not in valid_status:
            err_fields.append("status")
            err_msgs.append("状态非法")
        for val, name, field_name in [
            (row.channel_fee, "通道费格式非法", "channel_fee"),
            (row.tax_rate, "税点格式非法", "tax_rate"),
            (row.rd_share, "研发分成格式非法", "rd_share"),
            (row.private_rate, "私点格式非法", "private_rate"),
        ]:
            if val is None:
                err_fields.append(field_name)
                err_msgs.append(name)
        if err_msgs:
            error_details.append(
                {
                    "row_no": row.row_no or idx + 2,
                    "raw_data": row.model_dump(),
                    "error_fields": err_fields,
                    "error_message": ";".join(err_msgs),
                }
            )
    return {
        "total_count": len(payload.rows),
        "failed_count": len(error_details),
        "valid_count": len(payload.rows) - len(error_details),
        "error_details": error_details,
    }


def _money2(value: Decimal) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"))


def _ensure_period_format(period: str):
    raw = (period or "").strip()
    try:
        dt.datetime.strptime(raw, "%Y-%m")
    except ValueError as e:
        raise HTTPException(status_code=400, detail="period 格式非法，应为 YYYY-MM") from e


def _normalize_period_yyyymm(period: str) -> str:
    """接受 2026-3、2026/03、2026-03 等，统一为与库内一致的 YYYY-MM。"""
    raw = (period or "").strip().replace("/", "-")
    m = re.match(r"^(\d{4})-(\d{1,2})$", raw)
    if not m:
        raise HTTPException(status_code=400, detail="账期格式应为 YYYY-MM，例如 2026-03")
    y, mo = m.group(1), int(m.group(2))
    if mo < 1 or mo > 12:
        raise HTTPException(status_code=400, detail="账期月份应在 1–12 之间")
    return f"{y}-{mo:02d}"


def _resolve_parties_for_channel(db: Session, channel: Channel) -> tuple[str, str]:
    """甲方（平台）/ 乙方：优先有效合同中快照，否则默认平台名与渠道名。"""
    platform = "广州熊动科技有限公司"
    partner = (channel.name or "").strip() or "——"
    row = db.scalar(
        select(ContractHeader)
        .where(ContractHeader.channel_name == channel.name, ContractHeader.status == ContractStatus.active)
        .order_by(ContractHeader.end_date.desc())
        .limit(1)
    )
    if row:
        if (row.platform_party_name or "").strip():
            platform = row.platform_party_name.strip()
        if (row.developer_party_name or "").strip():
            partner = row.developer_party_name.strip()
    return platform, partner


def _build_settlement_snapshot(db: Session, period: str, channel_id: int) -> tuple[Channel, dict[int, dict], Decimal, Decimal, Decimal, Decimal, Decimal, str, str]:
    """基于该账期唯一有效已确认导入批次，按渠道聚合渠道+游戏行并计算结算列。"""
    period = _normalize_period_yyyymm(period)
    channel = db.get(Channel, channel_id)
    if not channel:
        raise HTTPException(status_code=404, detail="渠道不存在")
    task_ids = _active_confirmed_recon_task_ids_for_billing(db, period)
    if len(task_ids) > 1:
        raise HTTPException(status_code=400, detail=BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS)
    if len(task_ids) == 0:
        any_c = db.scalars(select(ReconTask).where(ReconTask.period == period, ReconTask.status == ReconStatus.confirmed)).all()
        if not any_c:
            raise HTTPException(status_code=400, detail="该账期无已确认核对任务")
        raise HTTPException(
            status_code=400,
            detail="该账期无已确认且有效的导入批次（可能均已作废），请先在导入数据中心处理后再生成月结单。",
        )
    rows = db.scalars(
        select(RawStatement).where(
            RawStatement.recon_task_id.in_(task_ids),
            RawStatement.period == period,
            RawStatement.channel_name == channel.name,
        )
    ).all()
    if not rows:
        raise HTTPException(status_code=400, detail="该渠道该账期无可用流水")
    maps = db.scalars(select(ChannelGameMap).where(ChannelGameMap.channel_id == channel_id)).all()
    map_by_game_name = {m.game.name: m for m in maps}
    if not map_by_game_name:
        raise HTTPException(status_code=400, detail="该渠道缺少渠道游戏映射，无法生成对账单")
    unmatched_games: set[str] = set()
    items_by_game: dict[int, dict] = {}
    for row in rows:
        link = map_by_game_name.get(row.game_name)
        if not link:
            unmatched_games.add(row.game_name)
            continue
        game_id = link.game_id
        item = items_by_game.get(game_id)
        rev = Decimal(str(link.revenue_share_ratio))
        if not item:
            item = {
                "game_id": game_id,
                "raw_game_name_snapshot": row.game_name or "",
                "game_name_snapshot": link.game.name,
                "gross_amount": Decimal("0"),
                "discount_amount": Decimal("0"),
                "test_fee_amount": Decimal("0"),
                "coupon_amount": Decimal("0"),
                "settlement_base_amount": Decimal("0"),
                "share_ratio": rev,
                "channel_fee_rate": rev,
                "channel_fee_amount": Decimal("0"),
                "settlement_amount": Decimal("0"),
            }
            items_by_game[game_id] = item
        item["gross_amount"] += Decimal(str(row.gross_amount or 0))
    if unmatched_games:
        sample = "、".join(sorted(list(unmatched_games))[:5])
        raise HTTPException(status_code=400, detail=f"存在未映射游戏，无法生成：{sample}")
    total_gross_amount = Decimal("0")
    total_discount_amount = Decimal("0")
    total_settlement_base_amount = Decimal("0")
    total_channel_fee_amount = Decimal("0")
    total_settlement_amount = Decimal("0")
    for _, item in items_by_game.items():
        item["gross_amount"] = _money2(item["gross_amount"])
        item["test_fee_amount"] = Decimal("0.00")
        item["coupon_amount"] = Decimal("0.00")
        item["discount_amount"] = _money2(item["test_fee_amount"] + item["coupon_amount"])
        item["settlement_base_amount"] = _money2(item["gross_amount"] - item["discount_amount"])
        item["channel_fee_amount"] = _money2(item["settlement_base_amount"] * item["channel_fee_rate"])
        item["settlement_amount"] = _money2(item["settlement_base_amount"] - item["channel_fee_amount"])
        total_gross_amount += item["gross_amount"]
        total_discount_amount += item["discount_amount"]
        total_settlement_base_amount += item["settlement_base_amount"]
        total_channel_fee_amount += item["channel_fee_amount"]
        total_settlement_amount += item["settlement_amount"]
    party_a, party_b = _resolve_parties_for_channel(db, channel)
    return (
        channel,
        items_by_game,
        _money2(total_gross_amount),
        _money2(total_discount_amount),
        _money2(total_settlement_base_amount),
        _money2(total_channel_fee_amount),
        _money2(total_settlement_amount),
        party_a,
        party_b,
    )


def _settlement_channel_ids_for_period(db: Session, period: str, task_ids: list[int]) -> list[int]:
    rows = db.scalars(
        select(RawStatement).where(RawStatement.recon_task_id.in_(task_ids), RawStatement.period == period)
    ).all()
    maps = db.scalars(select(ChannelGameMap)).all()
    map_key = {(m.channel.name, m.game.name): m for m in maps}
    ids: set[int] = set()
    for r in rows:
        link = map_key.get((r.channel_name, r.game_name))
        if link:
            ids.add(link.channel_id)
    return sorted(ids)


def _settlement_summary_payload(
    db: Session,
    period: str,
    channel_id: Optional[int],
    keyword: Optional[str],
) -> dict:
    """按账期 + 与列表一致的渠道筛选，统计月结单与对账进度（基于导入可映射渠道范围）。"""
    period_n = _normalize_period_yyyymm(period)
    task_ids = _active_confirmed_recon_task_ids_for_billing(db, period_n)
    warn: Optional[str] = None
    if len(task_ids) > 1:
        warn = BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS
        eligible_ids: list[int] = []
    elif len(task_ids) == 0:
        eligible_ids = []
    else:
        eligible_ids = _settlement_channel_ids_for_period(db, period_n, task_ids)
    ch_map: dict[int, str] = {}
    if eligible_ids:
        ch_rows = db.scalars(select(Channel).where(Channel.id.in_(eligible_ids))).all()
        ch_map = {c.id: c.name or "" for c in ch_rows}
    if channel_id is not None:
        eligible_ids = [cid for cid in eligible_ids if cid == channel_id]
    if keyword and str(keyword).strip():
        key = str(keyword).strip().lower()
        eligible_ids = [cid for cid in eligible_ids if key in (ch_map.get(cid, "") or "").lower()]
    statements: list[ChannelSettlementStatement] = []
    if eligible_ids:
        statements = list(
            db.scalars(
                select(ChannelSettlementStatement).where(
                    ChannelSettlementStatement.period == period_n,
                    ChannelSettlementStatement.channel_id.in_(eligible_ids),
                    ChannelSettlementStatement.status == "generated",
                )
            ).all()
        )
    stmt_by_cid = {s.channel_id: s for s in statements}
    eligible_total = len(eligible_ids)
    generated_count = len(statements)
    exported_count = 0
    reconciled_done = 0
    pending_recon_count = 0
    for cid in eligible_ids:
        st = stmt_by_cid.get(cid)
        if not st:
            pending_recon_count += 1
            continue
        rs = getattr(st, "reconciliation_status", None) or "pending"
        if rs == "exported":
            exported_count += 1
            reconciled_done += 1
        elif rs == "confirmed":
            reconciled_done += 1
        else:
            pending_recon_count += 1
    total_settlement = sum((Decimal(str(s.total_settlement_amount or 0)) for s in statements), start=Decimal("0"))
    return {
        "period": period_n,
        "eligible_channel_count": eligible_total,
        "generated_statement_count": generated_count,
        "pending_reconciliation_count": pending_recon_count,
        "reconciled_channel_count": reconciled_done,
        "exported_channel_count": exported_count,
        "total_settlement_amount": total_settlement,
        "warning": warn,
    }


@app.get("/settlement-statements/summary")
def settlement_statements_summary(
    period: str = Query(..., description="账期 YYYY-MM"),
    channel_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.ops])),
):
    try:
        payload = _settlement_summary_payload(db, period, channel_id, keyword)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="账期格式无效或无法汇总")
    return payload


@app.get("/settlement-statements")
def list_settlement_statements(
    period: Optional[str] = None,
    channel_id: Optional[int] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.ops])),
):
    stmt = select(ChannelSettlementStatement).order_by(ChannelSettlementStatement.id.desc())
    if period:
        p_filter = period.strip()
        try:
            p_filter = _normalize_period_yyyymm(p_filter)
        except HTTPException:
            pass
        stmt = stmt.where(ChannelSettlementStatement.period == p_filter)
    if channel_id:
        stmt = stmt.where(ChannelSettlementStatement.channel_id == channel_id)
    if status:
        stmt = stmt.where(ChannelSettlementStatement.status == status)
    rows = db.scalars(stmt).all()
    channel_name_map: dict[int, str] = {}
    if rows:
        ids = list({x.channel_id for x in rows})
        channels = db.scalars(select(Channel).where(Channel.id.in_(ids))).all()
        channel_name_map = {c.id: c.name for c in channels}
    items = [
        {
            "id": row.id,
            "period": row.period,
            "channel_id": row.channel_id,
            "channel_name": channel_name_map.get(row.channel_id, ""),
            "total_gross_amount": row.total_gross_amount,
            "total_discount_amount": row.total_discount_amount,
            "total_settlement_base_amount": row.total_settlement_base_amount,
            "total_channel_fee_amount": row.total_channel_fee_amount,
            "total_settlement_amount": row.total_settlement_amount,
            "status": row.status,
            "reconciliation_status": getattr(row, "reconciliation_status", None) or "pending",
            "updated_at": row.updated_at,
            "created_at": row.created_at,
        }
        for row in rows
    ]
    if keyword:
        key = keyword.strip().lower()
        items = [x for x in items if key in (x["channel_name"] or "").lower()]
    return {"items": items, "total": len(items)}


@app.post("/settlement-statements/generate")
def generate_settlement_statement(
    payload: SettlementStatementGenerateIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance])),
):
    period = _normalize_period_yyyymm(payload.period)
    existing = db.scalar(
        select(ChannelSettlementStatement).where(
            ChannelSettlementStatement.period == period,
            ChannelSettlementStatement.channel_id == payload.channel_id,
        )
    )
    if existing and not payload.overwrite:
        raise HTTPException(status_code=400, detail="已存在，请确认是否覆盖")
    (
        channel,
        items_by_game,
        total_gross_amount,
        total_discount_amount,
        total_settlement_base_amount,
        total_channel_fee_amount,
        total_settlement_amount,
        party_platform_name,
        party_channel_name,
    ) = _build_settlement_snapshot(db, period, payload.channel_id)
    now = dt.datetime.now()
    if existing:
        old_items = db.scalars(select(ChannelSettlementStatementItem).where(ChannelSettlementStatementItem.statement_id == existing.id)).all()
        for old in old_items:
            db.delete(old)
        statement = existing
    else:
        statement = ChannelSettlementStatement(
            period=period,
            channel_id=payload.channel_id,
            created_by=ctx["user"],
            status="generated",
            reconciliation_status="pending",
            created_at=now,
            updated_at=now,
        )
        db.add(statement)
        db.flush()
    statement.reconciliation_status = "pending"
    statement.total_gross_amount = total_gross_amount
    statement.total_discount_amount = total_discount_amount
    statement.total_settlement_base_amount = total_settlement_base_amount
    statement.total_channel_fee_amount = total_channel_fee_amount
    statement.total_settlement_amount = total_settlement_amount
    statement.party_platform_name = party_platform_name
    statement.party_channel_name = party_channel_name
    statement.status = "generated"
    statement.updated_at = now
    sort_order = 1
    for item in sorted(items_by_game.values(), key=lambda x: x["game_name_snapshot"]):
        db.add(
            ChannelSettlementStatementItem(
                statement_id=statement.id,
                game_id=item["game_id"],
                raw_game_name_snapshot=item["raw_game_name_snapshot"],
                game_name_snapshot=item["game_name_snapshot"],
                gross_amount=item["gross_amount"],
                discount_amount=item["discount_amount"],
                test_fee_amount=item["test_fee_amount"],
                coupon_amount=item["coupon_amount"],
                settlement_base_amount=item["settlement_base_amount"],
                share_ratio=item["share_ratio"],
                channel_fee_rate=item["channel_fee_rate"],
                channel_fee_amount=item["channel_fee_amount"],
                settlement_amount=item["settlement_amount"],
                sort_order=sort_order,
                created_at=now,
                updated_at=now,
            )
        )
        sort_order += 1
    action = "regenerate_settlement_statement" if existing else "generate_settlement_statement"
    write_system_audit(
        db,
        ctx["user"],
        action,
        "settlement_statement",
        str(statement.id),
        f"生成渠道月结单：{period} / {channel.name}",
    )
    db.commit()
    return {"id": statement.id, "period": statement.period, "channel_id": statement.channel_id, "status": statement.status}


@app.post("/settlement-statements/generate-all-for-period")
def generate_all_channel_settlements_for_period(
    payload: SettlementGenerateAllForPeriodIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance])),
):
    """按账期为所有在导入数据中出现且可映射的渠道生成/覆盖月结单（一渠道一单、多游戏明细）。"""
    period = _normalize_period_yyyymm(payload.period)
    task_ids = _active_confirmed_recon_task_ids_for_billing(db, period)
    if len(task_ids) > 1:
        raise HTTPException(status_code=400, detail=BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS)
    if len(task_ids) == 0:
        any_c = db.scalars(select(ReconTask).where(ReconTask.period == period, ReconTask.status == ReconStatus.confirmed)).all()
        if not any_c:
            raise HTTPException(status_code=400, detail="该账期无已确认核对任务")
        raise HTTPException(
            status_code=400,
            detail="该账期无已确认且有效的导入批次，请先在导入数据中心处理后再生成月结单。",
        )
    channel_ids = _settlement_channel_ids_for_period(db, period, task_ids)
    if not channel_ids:
        raise HTTPException(status_code=400, detail="该账期无可用渠道游戏映射数据，无法生成月结单")
    ok: list[dict] = []
    errors: list[dict] = []
    for cid in channel_ids:
        inner_existing = db.scalar(
            select(ChannelSettlementStatement).where(
                ChannelSettlementStatement.period == period,
                ChannelSettlementStatement.channel_id == cid,
            )
        )
        if inner_existing and not payload.overwrite:
            errors.append({"channel_id": cid, "detail": "已存在月结单，请勾选覆盖重生成或先删除后重试"})
            continue
        try:
            (
                ch,
                items_by_game,
                total_gross_amount,
                total_discount_amount,
                total_settlement_base_amount,
                total_channel_fee_amount,
                total_settlement_amount,
                party_platform_name,
                party_channel_name,
            ) = _build_settlement_snapshot(db, period, cid)
            now = dt.datetime.now()
            if inner_existing:
                old_items = db.scalars(
                    select(ChannelSettlementStatementItem).where(ChannelSettlementStatementItem.statement_id == inner_existing.id)
                ).all()
                for old in old_items:
                    db.delete(old)
                st = inner_existing
            else:
                st = ChannelSettlementStatement(
                    period=period,
                    channel_id=cid,
                    created_by=ctx["user"],
                    status="generated",
                    reconciliation_status="pending",
                    created_at=now,
                    updated_at=now,
                )
                db.add(st)
                db.flush()
            st.reconciliation_status = "pending"
            st.total_gross_amount = total_gross_amount
            st.total_discount_amount = total_discount_amount
            st.total_settlement_base_amount = total_settlement_base_amount
            st.total_channel_fee_amount = total_channel_fee_amount
            st.total_settlement_amount = total_settlement_amount
            st.party_platform_name = party_platform_name
            st.party_channel_name = party_channel_name
            st.status = "generated"
            st.updated_at = now
            sort_order = 1
            for item in sorted(items_by_game.values(), key=lambda x: x["game_name_snapshot"]):
                db.add(
                    ChannelSettlementStatementItem(
                        statement_id=st.id,
                        game_id=item["game_id"],
                        raw_game_name_snapshot=item["raw_game_name_snapshot"],
                        game_name_snapshot=item["game_name_snapshot"],
                        gross_amount=item["gross_amount"],
                        discount_amount=item["discount_amount"],
                        test_fee_amount=item["test_fee_amount"],
                        coupon_amount=item["coupon_amount"],
                        settlement_base_amount=item["settlement_base_amount"],
                        share_ratio=item["share_ratio"],
                        channel_fee_rate=item["channel_fee_rate"],
                        channel_fee_amount=item["channel_fee_amount"],
                        settlement_amount=item["settlement_amount"],
                        sort_order=sort_order,
                        created_at=now,
                        updated_at=now,
                    )
                )
                sort_order += 1
            write_system_audit(
                db,
                ctx["user"],
                "generate_settlement_batch",
                "settlement_statement",
                str(st.id),
                f"批量生成渠道月结单：{period} / {ch.name}",
            )
            ok.append({"id": st.id, "channel_id": cid, "channel_name": ch.name})
        except HTTPException as e:
            errors.append({"channel_id": cid, "detail": e.detail})
    db.commit()
    return {"period": period, "generated": ok, "errors": errors}


@app.get("/settlement-statements/{statement_id}")
def get_settlement_statement(
    statement_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.ops])),
):
    statement = db.get(ChannelSettlementStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="对账单不存在")
    channel = db.get(Channel, statement.channel_id)
    items = db.scalars(
        select(ChannelSettlementStatementItem)
        .where(ChannelSettlementStatementItem.statement_id == statement_id)
        .order_by(ChannelSettlementStatementItem.sort_order.asc(), ChannelSettlementStatementItem.id.asc())
    ).all()
    party_a = getattr(statement, "party_platform_name", None) or "广州熊动科技有限公司"
    party_b = getattr(statement, "party_channel_name", None) or ""
    return {
        "id": statement.id,
        "period": statement.period,
        "channel_id": statement.channel_id,
        "channel_name": channel.name if channel else "",
        "status": statement.status,
        "total_gross_amount": statement.total_gross_amount,
        "total_discount_amount": statement.total_discount_amount,
        "total_settlement_base_amount": statement.total_settlement_base_amount,
        "total_channel_fee_amount": statement.total_channel_fee_amount,
        "total_settlement_amount": statement.total_settlement_amount,
        "note": statement.note,
        "party_platform_name": party_a,
        "party_channel_name": party_b,
        "reconciliation_status": getattr(statement, "reconciliation_status", None) or "pending",
        "created_by": statement.created_by,
        "created_at": statement.created_at,
        "updated_at": statement.updated_at,
        "items": [
            {
                "id": x.id,
                "game_id": x.game_id,
                "raw_game_name_snapshot": x.raw_game_name_snapshot,
                "game_name_snapshot": x.game_name_snapshot,
                "gross_amount": x.gross_amount,
                "discount_amount": x.discount_amount,
                "test_fee_amount": getattr(x, "test_fee_amount", Decimal("0")),
                "coupon_amount": getattr(x, "coupon_amount", Decimal("0")),
                "settlement_base_amount": x.settlement_base_amount,
                "share_ratio": getattr(x, "share_ratio", x.channel_fee_rate),
                "channel_fee_rate": x.channel_fee_rate,
                "channel_fee_amount": x.channel_fee_amount,
                "settlement_amount": x.settlement_amount,
                "sort_order": x.sort_order,
            }
            for x in items
        ],
    }


@app.patch("/settlement-statements/{statement_id}/reconciliation-status")
def patch_settlement_reconciliation_status(
    statement_id: int,
    payload: SettlementReconciliationStatusIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance])),
):
    statement = db.get(ChannelSettlementStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="对账单不存在")
    statement.reconciliation_status = payload.reconciliation_status
    statement.updated_at = dt.datetime.now()
    write_system_audit(
        db,
        ctx["user"],
        "update_settlement_reconciliation_status",
        "settlement_statement",
        str(statement.id),
        f"更新对账状态：{statement.period} → {payload.reconciliation_status}",
    )
    db.commit()
    return {"id": statement.id, "reconciliation_status": statement.reconciliation_status}


@app.get("/settlement-statements/{statement_id}/export")
def export_settlement_statement(
    statement_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance])),
):
    statement = db.get(ChannelSettlementStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="对账单不存在")
    channel = db.get(Channel, statement.channel_id)
    items = db.scalars(
        select(ChannelSettlementStatementItem)
        .where(ChannelSettlementStatementItem.statement_id == statement_id)
        .order_by(ChannelSettlementStatementItem.sort_order.asc(), ChannelSettlementStatementItem.id.asc())
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "月结单"
    ch_label = channel.name if channel else ""
    yymm = statement.period.replace("-", "")
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"{ch_label} & 熊动结算对账单{yymm}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    note_row = 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
    c_note = ws.cell(
        row=note_row,
        column=1,
        value="说明：合作总收入、参与分成金额、结算金额等为系统按导入流水与映射规则汇总，测试费/代金券暂无导入字段时为 0。",
    )
    c_note.alignment = Alignment(wrap_text=True, vertical="top")
    headers = [
        "结算月份",
        "游戏名称",
        "合作总收入",
        "测试费",
        "代金券",
        "参与分成金额",
        "分成比例",
        "通道费",
        "结算金额",
    ]
    start_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)
    row_idx = start_row + 1
    for item in items:
        tf = float(getattr(item, "test_fee_amount", 0) or 0)
        cp = float(getattr(item, "coupon_amount", 0) or 0)
        sh = float(getattr(item, "share_ratio", item.channel_fee_rate) or 0)
        ws.cell(row=row_idx, column=1, value=statement.period)
        ws.cell(row=row_idx, column=2, value=item.game_name_snapshot)
        ws.cell(row=row_idx, column=3, value=float(item.gross_amount))
        ws.cell(row=row_idx, column=4, value=tf)
        ws.cell(row=row_idx, column=5, value=cp)
        ws.cell(row=row_idx, column=6, value=float(item.settlement_base_amount))
        c_ratio = ws.cell(row=row_idx, column=7, value=sh)
        c_ratio.number_format = "0.00%"
        ws.cell(row=row_idx, column=8, value=float(item.channel_fee_amount))
        ws.cell(row=row_idx, column=9, value=float(item.settlement_amount))
        for col in (3, 4, 5, 6, 8, 9):
            ws.cell(row=row_idx, column=col).number_format = "#,##0.00"
        row_idx += 1
    ws.cell(row=row_idx, column=1, value="合计")
    sum_row = row_idx
    ws.cell(row=sum_row, column=3, value=float(statement.total_gross_amount))
    ws.cell(row=sum_row, column=4, value=sum(float(getattr(x, "test_fee_amount", 0) or 0) for x in items))
    ws.cell(row=sum_row, column=5, value=sum(float(getattr(x, "coupon_amount", 0) or 0) for x in items))
    ws.cell(row=sum_row, column=6, value=float(statement.total_settlement_base_amount))
    ws.cell(row=sum_row, column=8, value=float(statement.total_channel_fee_amount))
    ws.cell(row=sum_row, column=9, value=float(statement.total_settlement_amount))
    for col in (3, 4, 5, 6, 8, 9):
        ws.cell(row=sum_row, column=col).number_format = "#,##0.00"
        cell_font = Font(bold=True)
        ws.cell(row=sum_row, column=col).font = cell_font
    ws.cell(row=sum_row, column=1).font = Font(bold=True)
    row_idx = sum_row + 2
    party_a = getattr(statement, "party_platform_name", None) or "广州熊动科技有限公司"
    party_b = getattr(statement, "party_channel_name", None) or ch_label
    ws.cell(row=row_idx, column=1, value="甲方（平台）")
    ws.cell(row=row_idx, column=2, value=party_a)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="乙方")
    ws.cell(row=row_idx, column=2, value=party_b)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="备注")
    ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=9)
    ws.cell(row=row_idx, column=2, value=statement.note or "")
    ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    widths = [12, 26, 14, 10, 10, 16, 12, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    bio = io.BytesIO()
    wb.save(bio)
    statement.reconciliation_status = "exported"
    statement.updated_at = dt.datetime.now()
    write_system_audit(
        db,
        ctx["user"],
        "export_settlement_statement",
        "settlement_statement",
        str(statement.id),
        f"导出渠道对账单：{statement.period} / {(channel.name if channel else '')}",
    )
    db.commit()
    filename = f"settlement_statement_{statement.period}_{statement.id}.xlsx"
    return Response(
        content=bio.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/billing/rules/export")
def export_rules(db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz]))):
    rows = db.scalars(select(BillingRule).order_by(BillingRule.id.desc())).all()
    return [
        {
            "id": x.id,
            "name": x.name,
            "bill_type": x.bill_type,
            "default_ratio": x.default_ratio,
            "active": x.active,
        }
        for x in rows
    ]


BILL_SUSPICIOUS_CHANNEL_BUCKET = "未匹配渠道（原始流水渠道名需修正）"

# 账单生成：同一账期仅允许一条「lifecycle=active + 任务已确认」的导入批次，避免多批 raw 静默合并导致金额偏大
BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS = (
    "该账期存在多条已确认且有效的导入批次，请先撤销或作废旧批次后再生成账单。"
)


def _active_confirmed_recon_task_ids_for_billing(db: Session, period: str) -> list[int]:
    """收集账期内「导入批次仍有效」且「对应核对任务已确认」的 task_id（去重、顺序稳定）。"""
    out: list[int] = []
    seen: set[int] = set()
    for h in db.scalars(select(ImportHistory).where(ImportHistory.period == period)).all():
        if (h.lifecycle_status or "active") != "active":
            continue
        t = db.get(ReconTask, h.task_id)
        if not t or t.period != period or t.status != ReconStatus.confirmed:
            continue
        if h.task_id not in seen:
            seen.add(h.task_id)
            out.append(h.task_id)
    return out


def _is_placeholder_channel_label(name: Optional[str]) -> bool:
    """占位或明显异常的渠道名字符串：不作为对外账单渠道名展示。"""
    s = (name or "").strip()
    if not s:
        return True
    sl = s.lower()
    bad_tokens = {
        "待补充",
        "未填写",
        "未知",
        "无",
        "待定",
        "null",
        "none",
        "-",
        "—",
        "n/a",
        "na",
        "tbd",
        "(待补充)",
        "【待补充】",
    }
    if s in bad_tokens or sl in {x.lower() for x in bad_tokens}:
        return True
    # 较短纯数字多為行号/临时填充，避免当正式渠道名进账单
    if s.isdigit() and len(s) <= 3:
        return True
    return False


def _bill_channel_target_name(raw_statement_channel: str, canonical_channel_name: str) -> str:
    raw = (raw_statement_channel or "").strip()
    canon = (canonical_channel_name or "").strip()
    if _is_placeholder_channel_label(raw) or _is_placeholder_channel_label(canon):
        return BILL_SUSPICIOUS_CHANNEL_BUCKET
    return raw


def _compute_billing_split_from_raw(db: Session, task_ids: list[int]) -> dict:
    """与账单生成相同的 raw→渠道/研发拆分；保留金额=总流水−渠道拆分−研发拆分（含未映射与分成剩余）。"""
    rows = db.scalars(select(RawStatement).where(RawStatement.recon_task_id.in_(task_ids))).all()
    map_rows = db.scalars(select(ChannelGameMap)).all()
    map_key = {(m.channel.name, m.game.name): m for m in map_rows}
    total_gross = Decimal("0")
    channel_sum: dict[str, Decimal] = {}
    rd_sum: dict[str, Decimal] = {}
    unmapped_gross = Decimal("0")
    unmapped_rows = 0
    for r in rows:
        g = Decimal(str(r.gross_amount or 0))
        total_gross += g
        link = map_key.get((r.channel_name, r.game_name))
        if not link:
            unmapped_gross += g
            unmapped_rows += 1
            continue
        channel_amount = g * link.revenue_share_ratio
        rd_amount = g * link.rd_settlement_ratio
        bill_ch = _bill_channel_target_name(r.channel_name, link.channel.name)
        channel_sum[bill_ch] = channel_sum.get(bill_ch, Decimal("0")) + channel_amount
        rd_sum[link.game.rd_company] = rd_sum.get(link.game.rd_company, Decimal("0")) + rd_amount
    ch_tot = sum(channel_sum.values(), Decimal("0"))
    rd_tot = sum(rd_sum.values(), Decimal("0"))
    retention = total_gross - ch_tot - rd_tot
    balance_delta = total_gross - ch_tot - rd_tot - retention
    return {
        "channel_sum": channel_sum,
        "rd_sum": rd_sum,
        "total_import_gross": total_gross,
        "channel_split_total": ch_tot,
        "rd_split_total": rd_tot,
        "publisher_retention_total": retention,
        "balance_delta": balance_delta,
        "unmapped_gross": unmapped_gross,
        "unmapped_row_count": unmapped_rows,
        "mapped_row_count": len(rows) - unmapped_rows,
        "raw_row_count": len(rows),
    }


CHANNEL_BRIDGE_UNMAPPED_LABEL = "（未映射渠道游戏，未进入账单拆分）"


def _channel_rows_for_settlement_reconciliation(db: Session, task_ids: list[int], period: str, split: dict) -> list[dict]:
    """按与账单生成相同的渠道桶（bill_ch）汇总：原始流水、渠道拆分、归属研发拆分、保留；并对比有效渠道账单行。"""
    rows = db.scalars(select(RawStatement).where(RawStatement.recon_task_id.in_(task_ids))).all()
    map_rows = db.scalars(select(ChannelGameMap)).all()
    map_key = {(m.channel.name, m.game.name): m for m in map_rows}
    gross_by_ch: dict[str, Decimal] = {}
    ch_split_by_ch: dict[str, Decimal] = {}
    rd_split_by_ch: dict[str, Decimal] = {}
    unmapped_gross = Decimal("0")
    for r in rows:
        g = Decimal(str(r.gross_amount or 0))
        link = map_key.get((r.channel_name, r.game_name))
        if not link:
            unmapped_gross += g
            continue
        bill_ch = _bill_channel_target_name(r.channel_name, link.channel.name)
        gross_by_ch[bill_ch] = gross_by_ch.get(bill_ch, Decimal("0")) + g
        ch_a = g * link.revenue_share_ratio
        rd_a = g * link.rd_settlement_ratio
        ch_split_by_ch[bill_ch] = ch_split_by_ch.get(bill_ch, Decimal("0")) + ch_a
        rd_split_by_ch[bill_ch] = rd_split_by_ch.get(bill_ch, Decimal("0")) + rd_a

    active_bills = db.scalars(select(Bill).where(Bill.period == period, Bill.lifecycle_status == "active")).all()
    bill_ch_by_target: dict[str, Decimal] = {}
    for b in active_bills:
        if b.bill_type == BillType.channel:
            bill_ch_by_target[b.target_name] = bill_ch_by_target.get(b.target_name, Decimal("0")) + Decimal(str(b.amount))

    def _near(a: Decimal, b: Decimal) -> bool:
        return abs(a - b) <= Decimal("0.009")

    out: list[dict] = []
    for ch in sorted(gross_by_ch.keys()):
        gtot = gross_by_ch[ch]
        ctot = ch_split_by_ch.get(ch, Decimal("0"))
        rtot = rd_split_by_ch.get(ch, Decimal("0"))
        ret = gtot - ctot - rtot
        bal = gtot - ctot - rtot - ret
        bpart = bill_ch_by_target.get(ch, Decimal("0"))
        out.append(
            {
                "channel": ch,
                "import_gross_total": str(gtot),
                "channel_split_total": str(ctot),
                "rd_split_total": str(rtot),
                "retention_total": str(ret),
                "balance_delta": str(bal),
                "is_balanced": abs(bal) <= Decimal("0.0001"),
                "active_bills_channel_total": str(bpart),
                "bills_match_channel_split": _near(bpart, ctot),
            }
        )

    if split.get("unmapped_row_count", 0) > 0 or unmapped_gross > 0:
        out.append(
            {
                "channel": CHANNEL_BRIDGE_UNMAPPED_LABEL,
                "import_gross_total": str(unmapped_gross),
                "channel_split_total": "0",
                "rd_split_total": "0",
                "retention_total": str(unmapped_gross),
                "balance_delta": "0",
                "is_balanced": True,
                "active_bills_channel_total": "0",
                "bills_match_channel_split": True,
            }
        )
    return out


@app.get("/settlement-statements/period-reconciliation")
def get_settlement_period_reconciliation(
    period: str = Query(...),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    period = _normalize_period_yyyymm(period)
    task_ids = _active_confirmed_recon_task_ids_for_billing(db, period)
    if len(task_ids) > 1:
        raise HTTPException(status_code=400, detail=BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS)
    if len(task_ids) == 0:
        any_confirmed = db.scalars(select(ReconTask).where(ReconTask.period == period, ReconTask.status == ReconStatus.confirmed)).all()
        if not any_confirmed:
            raise HTTPException(status_code=400, detail="该账期无已确认核对任务")
        raise HTTPException(
            status_code=400,
            detail="该账期无已确认且有效的导入批次（可能均已作废），请先在导入数据中心恢复有效批次或撤销入账后再试。",
        )
    split = _compute_billing_split_from_raw(db, task_ids)
    if split["raw_row_count"] == 0:
        raise HTTPException(status_code=400, detail="无原始数据")
    channel_rows = _channel_rows_for_settlement_reconciliation(db, task_ids, period, split)
    active_bills = db.scalars(select(Bill).where(Bill.period == period, Bill.lifecycle_status == "active")).all()
    bills_ch = sum((Decimal(str(b.amount)) for b in active_bills if b.bill_type == BillType.channel), Decimal("0"))
    bills_rd = sum((Decimal(str(b.amount)) for b in active_bills if b.bill_type == BillType.rd), Decimal("0"))
    raw_ch: Decimal = split["channel_split_total"]
    raw_rd: Decimal = split["rd_split_total"]

    def _money_eq(a: Decimal, b: Decimal) -> bool:
        return abs(a - b) <= Decimal("0.009")

    return {
        "period": period,
        "recon_task_id": task_ids[0],
        "summary": {
            "total_import_gross": str(split["total_import_gross"]),
            "channel_split_total": str(raw_ch),
            "rd_split_total": str(raw_rd),
            "publisher_retention_total": str(split["publisher_retention_total"]),
            "balance_delta": str(split["balance_delta"]),
            "unmapped_gross": str(split["unmapped_gross"]),
            "raw_row_count": split["raw_row_count"],
            "mapped_row_count": split["mapped_row_count"],
            "unmapped_row_count": split["unmapped_row_count"],
            "active_bills_channel_total": str(bills_ch),
            "active_bills_rd_total": str(bills_rd),
            "bills_match_raw_split": _money_eq(bills_ch, raw_ch) and _money_eq(bills_rd, raw_rd),
        },
        "channels": channel_rows,
        "intro_note": (
            "本页为渠道结算对账与导入流水核对入口：数据来自该账期唯一「有效且已确认」导入批次，渠道/研发拆分校验与账单生成逻辑一致。"
            "账单列表与流程仍以「账单管理」为准；原始流水批次核对仍以「导入数据中心」为准。"
        ),
    }


@app.get("/billing/period-bridge")
def get_billing_period_bridge(
    period: str = Query(...),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    task_ids = _active_confirmed_recon_task_ids_for_billing(db, period)
    if len(task_ids) > 1:
        raise HTTPException(status_code=400, detail=BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS)
    if len(task_ids) == 0:
        any_confirmed = db.scalars(select(ReconTask).where(ReconTask.period == period, ReconTask.status == ReconStatus.confirmed)).all()
        if not any_confirmed:
            raise HTTPException(status_code=400, detail="该账期无已确认核对任务")
        raise HTTPException(
            status_code=400,
            detail="该账期无已确认且有效的导入批次（可能均已作废），请先在导入数据中心恢复有效批次或撤销入账后再生成账单。",
        )
    split = _compute_billing_split_from_raw(db, task_ids)
    if split["raw_row_count"] == 0:
        raise HTTPException(status_code=400, detail="无原始数据")
    active_bills = db.scalars(select(Bill).where(Bill.period == period, Bill.lifecycle_status == "active")).all()
    bills_ch = sum((Decimal(str(b.amount)) for b in active_bills if b.bill_type == BillType.channel), Decimal("0"))
    bills_rd = sum((Decimal(str(b.amount)) for b in active_bills if b.bill_type == BillType.rd), Decimal("0"))
    raw_ch: Decimal = split["channel_split_total"]
    raw_rd: Decimal = split["rd_split_total"]

    def _money_eq(a: Decimal, b: Decimal) -> bool:
        return abs(a - b) <= Decimal("0.009")

    return {
        "period": period,
        "recon_task_id": task_ids[0],
        "total_import_gross": str(split["total_import_gross"]),
        "channel_split_total": str(raw_ch),
        "rd_split_total": str(raw_rd),
        "publisher_retention_total": str(split["publisher_retention_total"]),
        "balance_delta": str(split["balance_delta"]),
        "unmapped_gross": str(split["unmapped_gross"]),
        "unmapped_row_count": split["unmapped_row_count"],
        "mapped_row_count": split["mapped_row_count"],
        "raw_row_count": split["raw_row_count"],
        "active_bills_channel_total": str(bills_ch),
        "active_bills_rd_total": str(bills_rd),
        "bills_match_raw_split": _money_eq(bills_ch, raw_ch) and _money_eq(bills_rd, raw_rd),
        "note": (
            "原始流水合计为唯一有效已确认批次下 RawStatement.gross 汇总；渠道/研发拆分为与「生成账单」相同映射与公式；"
            "发行或公司保留=原始流水−渠道拆分合计−研发拆分合计（含未映射流水及单条流水分成剩余 1−渠道分成−研发分成）。"
            "差额恒为 0；若有效账单合计与拆分不一致，说明账单未与当前数据同步，请覆盖重生成。"
        ),
    }


@app.post("/billing/generate")
def generate_bills(
    period: str = Query(...),
    overwrite: bool = Query(default=False),
    force_new_version: bool = Query(default=False),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance])),
):
    # 兼容历史参数：旧前端可能仍传 force_new_version，语义上等同于重生成
    overwrite = bool(overwrite or force_new_version)
    period_bills = db.scalars(select(Bill).where(Bill.period == period)).all()
    active_period_bills = [b for b in period_bills if (b.lifecycle_status or "active") == "active"]
    # 与列表默认口径一致：仅当仍有「未作废」账单时禁止重复生成；全部为 discarded 时可重新生成
    if active_period_bills and not overwrite:
        raise HTTPException(status_code=400, detail="该账期账单已存在")
    task_ids = _active_confirmed_recon_task_ids_for_billing(db, period)
    if len(task_ids) > 1:
        raise HTTPException(status_code=400, detail=BILLING_PERIOD_MULTI_ACTIVE_CONFIRMED_IMPORTS)
    if len(task_ids) == 1:
        recon_tasks = [db.get(ReconTask, task_ids[0])]
        if not recon_tasks[0]:
            raise HTTPException(status_code=400, detail="该账期无已确认核对任务")
        task_ids = [recon_tasks[0].id]
    else:
        any_confirmed = db.scalars(select(ReconTask).where(ReconTask.period == period, ReconTask.status == ReconStatus.confirmed)).all()
        if not any_confirmed:
            raise HTTPException(status_code=400, detail="该账期无已确认核对任务")
        raise HTTPException(
            status_code=400,
            detail="该账期无已确认且有效的导入批次（可能均已作废），请先在导入数据中心恢复有效批次或撤销入账后再生成账单。",
        )
    split = _compute_billing_split_from_raw(db, task_ids)
    if split["raw_row_count"] == 0:
        raise HTTPException(status_code=400, detail="无原始数据")
    channel_sum: dict[str, Decimal] = split["channel_sum"]
    rd_sum: dict[str, Decimal] = split["rd_sum"]
    existing_key_map: dict[tuple[BillType, str], Bill] = {}
    for bill in active_period_bills:
        key = (bill.bill_type, bill.target_name)
        prev = existing_key_map.get(key)
        if not prev or bill.id > prev.id:
            existing_key_map[key] = bill
    created = 0
    updated = 0
    for target, amount in channel_sum.items():
        key = (BillType.channel, target)
        existing = existing_key_map.get(key)
        if existing:
            # 覆盖模式仅同步「待发送」草稿金额，避免改写已进入对账/开票/回款流程的账单
            if overwrite and existing.status == BillStatus.draft:
                existing.amount = amount
                updated += 1
            continue
        db.add(Bill(bill_type=BillType.channel, period=period, target_name=target, amount=amount, version=1))
        created += 1
    for target, amount in rd_sum.items():
        key = (BillType.rd, target)
        existing = existing_key_map.get(key)
        if existing:
            if overwrite and existing.status == BillStatus.draft:
                existing.amount = amount
                updated += 1
            continue
        db.add(Bill(bill_type=BillType.rd, period=period, target_name=target, amount=amount, version=1))
        created += 1
    audit_action = "regenerate_bills" if overwrite else "generate_bills"
    write_system_audit(db, _["user"], audit_action, "bill", period, f"生成账单: 新增{created}, 更新{updated}")
    db.commit()
    return {"created_bills": created, "updated_bills": updated, "overwrite": overwrite}


@app.get("/billing/bills")
def list_bills(
    period: Optional[str] = None,
    bill_type: Optional[BillType] = None,
    lifecycle_status: Optional[str] = "active",
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    stmt = select(Bill).order_by(Bill.id.desc())
    if lifecycle_status in {"active", "discarded"}:
        stmt = stmt.where(Bill.lifecycle_status == lifecycle_status)
    if period:
        stmt = stmt.where(Bill.period == period)
    if bill_type:
        stmt = stmt.where(Bill.bill_type == bill_type)
    rows = db.scalars(stmt).all()
    result = []
    for b in rows:
        invoices = db.scalars(select(Invoice).where(Invoice.bill_id == b.id).order_by(Invoice.id.desc())).all()
        receipts = db.scalars(select(Receipt).where(Receipt.bill_id == b.id).order_by(Receipt.id.desc())).all()
        invoiced_total = sum((Decimal(str(x.total_amount)) for x in invoices), Decimal("0"))
        received_total = sum((Decimal(str(x.amount)) for x in receipts), Decimal("0"))
        outstanding_amount = max(Decimal("0"), Decimal(str(b.amount)) - received_total)
        invoice_status = "已开票" if invoices else "待开票"
        receipt_status = "待回款"
        if received_total >= Decimal(str(b.amount)):
            receipt_status = "已回款"
        elif received_total > 0:
            receipt_status = "部分回款"
        flow_status = calc_bill_flow_status(Decimal(str(b.amount)), b.status, bool(invoices), received_total)
        result.append(
            {
                "id": b.id,
                "bill_type": b.bill_type,
                "period": b.period,
                "target_name": b.target_name,
                "amount": b.amount,
                "status": b.status,
                "version": b.version,
                "collection_status": b.collection_status,
                "lifecycle_status": b.lifecycle_status,
                "invoice_status": invoice_status,
                "receipt_status": receipt_status,
                "flow_status": flow_status,
                "invoiced_total": invoiced_total,
                "received_total": received_total,
                "outstanding_amount": outstanding_amount,
                "latest_receipt_date": str(receipts[0].received_at) if receipts else "",
            }
        )
    return result


@app.get("/billing/{bill_id}")
def get_bill_detail(
    bill_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    b = db.get(Bill, bill_id)
    if not b:
        raise HTTPException(status_code=404, detail="账单不存在")
    invoices = db.scalars(select(Invoice).where(Invoice.bill_id == b.id).order_by(Invoice.id.desc())).all()
    receipts = db.scalars(select(Receipt).where(Receipt.bill_id == b.id).order_by(Receipt.id.desc())).all()
    invoiced_total = sum((Decimal(str(x.total_amount)) for x in invoices), Decimal("0"))
    received_total = sum((Decimal(str(x.amount)) for x in receipts), Decimal("0"))
    outstanding_amount = max(Decimal("0"), Decimal(str(b.amount)) - received_total)
    latest_invoice = invoices[0] if invoices else None
    latest_receipt = receipts[0] if receipts else None
    invoice_status = "已开票" if invoices else "待开票"
    receipt_status = "待回款"
    if received_total >= Decimal(str(b.amount)):
        receipt_status = "已回款"
    elif received_total > 0:
        receipt_status = "部分回款"
    flow_status = calc_bill_flow_status(Decimal(str(b.amount)), b.status, bool(invoices), received_total)
    return {
        "id": b.id,
        "bill_type": b.bill_type,
        "period": b.period,
        "target_name": b.target_name,
        "amount": b.amount,
        "status": b.status,
        "version": b.version,
        "collection_status": b.collection_status,
        "lifecycle_status": b.lifecycle_status,
        "invoice_status": invoice_status,
        "receipt_status": receipt_status,
        "flow_status": flow_status,
        "invoiced_total": invoiced_total,
        "received_total": received_total,
        "outstanding_amount": outstanding_amount,
        "invoice_info": {
            "has_invoice": bool(invoices),
            "invoice_no": latest_invoice.invoice_no if latest_invoice else "",
            "invoice_amount": latest_invoice.total_amount if latest_invoice else Decimal("0"),
            "issue_date": str(latest_invoice.issue_date) if latest_invoice else "",
        },
        "receipt_info": {
            "received_total": received_total,
            "outstanding_amount": outstanding_amount,
            "latest_receipt_date": str(latest_receipt.received_at) if latest_receipt else "",
            "receipt_status": receipt_status,
        },
    }


@app.post("/billing/{bill_id}/discard")
def discard_bill(
    bill_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance])),
):
    bill = db.get(Bill, bill_id)
    if not bill:
        raise HTTPException(status_code=404, detail="账单不存在")
    if (bill.lifecycle_status or "active") == "discarded":
        return {"bill_id": bill.id, "lifecycle_status": "discarded", "already_discarded": True}

    invoices = db.scalars(select(Invoice).where(Invoice.bill_id == bill.id)).all()
    if invoices:
        raise HTTPException(status_code=400, detail="存在关联发票，禁止作废")
    receipts = db.scalars(select(Receipt).where(Receipt.bill_id == bill.id)).all()
    if receipts:
        raise HTTPException(status_code=400, detail="存在关联回款，禁止作废")
    received_total = sum((Decimal(str(x.amount)) for x in receipts), Decimal("0"))
    if received_total > 0:
        raise HTTPException(status_code=400, detail="已发生回款，禁止作废")

    if bill.status != BillStatus.draft:
        raise HTTPException(status_code=400, detail="仅草稿账单允许作废")
    delivery_count = int(db.scalar(select(func.count(BillDeliveryLog.id)).where(BillDeliveryLog.bill_id == bill.id)) or 0)
    if delivery_count > 0:
        raise HTTPException(status_code=400, detail="存在发送/交付记录，禁止作废")

    outstanding_amount = max(Decimal("0"), Decimal(str(bill.amount)) - received_total)
    if outstanding_amount != Decimal(str(bill.amount)):
        raise HTTPException(status_code=400, detail="存在已回款金额，禁止作废")

    bill.lifecycle_status = "discarded"
    write_system_audit(db, ctx["user"], "discard_bill", "bill", str(bill.id), f"作废账单: {bill.period}/{bill.bill_type}/{bill.target_name}")
    db.commit()
    return {"bill_id": bill.id, "lifecycle_status": bill.lifecycle_status}


@app.post("/billing/cleanup-duplicates")
def cleanup_duplicate_bills(
    payload: CleanupDuplicateBillsIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin])),
):
    rows = db.scalars(select(Bill).order_by(Bill.created_at.asc(), Bill.id.asc())).all()
    grouped: dict[tuple[str, BillType, str], list[Bill]] = {}
    for row in rows:
        key = (row.period, row.bill_type, row.target_name)
        grouped.setdefault(key, []).append(row)
    duplicate_groups = {k: v for k, v in grouped.items() if len(v) > 1}
    deleted_ids: list[int] = []
    kept_groups: list[dict] = []
    skipped: list[dict] = []

    def can_delete_safely(bill: Bill) -> tuple[bool, str]:
        # 仅清理：草稿 + 未回款 + 无下游依赖
        if bill.status != BillStatus.draft:
            return False, f"账单状态非草稿({bill.status.value})"
        if bill.collection_status != CollectionStatus.pending:
            return False, f"回款状态非待回款({bill.collection_status.value})"
        invoice_count = int(db.scalar(select(func.count(Invoice.id)).where(Invoice.bill_id == bill.id)) or 0)
        if invoice_count > 0:
            return False, "存在关联发票"
        receipt_count = int(db.scalar(select(func.count(Receipt.id)).where(Receipt.bill_id == bill.id)) or 0)
        if receipt_count > 0:
            return False, "存在关联回款"
        delivery_count = int(db.scalar(select(func.count(BillDeliveryLog.id)).where(BillDeliveryLog.bill_id == bill.id)) or 0)
        if delivery_count > 0:
            return False, "存在发送/交付记录"
        audit_count = int(
            db.scalar(
                select(func.count(SystemAuditLog.id)).where(
                    SystemAuditLog.target_type == "bill",
                    SystemAuditLog.target_id == str(bill.id),
                )
            )
            or 0
        )
        if audit_count > 0:
            return False, "存在账单审计依赖"
        return True, ""

    for (period, bill_type, target_name), bills in duplicate_groups.items():
        sorted_bills = sorted(bills, key=lambda x: (x.created_at or dt.datetime.min, x.id))
        keep = sorted_bills[0]
        kept_groups.append(
            {
                "period": period,
                "bill_type": bill_type.value if isinstance(bill_type, BillType) else str(bill_type),
                "target_name": target_name,
                "keep_id": keep.id,
                "group_size": len(sorted_bills),
            }
        )
        for candidate in sorted_bills[1:]:
            ok, reason = can_delete_safely(candidate)
            if not ok:
                skipped.append(
                    {
                        "bill_id": candidate.id,
                        "period": period,
                        "bill_type": bill_type.value if isinstance(bill_type, BillType) else str(bill_type),
                        "target_name": target_name,
                        "reason": reason,
                    }
                )
                continue
            deleted_ids.append(candidate.id)
            if not payload.dry_run:
                db.delete(candidate)

    write_system_audit(
        db,
        ctx["user"],
        "cleanup_duplicate_bills",
        "bill",
        "",
        f"重复账单清理 dry_run={payload.dry_run}: 删除候选{len(deleted_ids)}, 分组{len(kept_groups)}, 跳过{len(skipped)}",
    )
    if not payload.dry_run:
        db.commit()
    else:
        db.rollback()
    return {
        "dry_run": payload.dry_run,
        "duplicate_group_count": len(kept_groups),
        "kept_group_count": len(kept_groups),
        "deleted_count": len(deleted_ids),
        "deleted_ids": deleted_ids,
        "kept_groups": kept_groups,
        "skipped_count": len(skipped),
        "skipped": skipped,
    }


@app.post("/billing/{bill_id}/send")
def send_bill(bill_id: int, payload: BillStatusIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz]))):
    bill = db.get(Bill, bill_id)
    if not bill:
        raise HTTPException(status_code=404, detail="账单不存在")
    if (bill.lifecycle_status or "active") == "discarded":
        raise HTTPException(status_code=400, detail="已作废账单不可发送/确认")
    bill.status = payload.status
    if payload.status in (BillStatus.sent, BillStatus.acknowledged, BillStatus.disputed):
        db.add(BillDeliveryLog(bill_id=bill_id, note=payload.note))
    write_system_audit(db, _["user"], "send_bill", "bill", str(bill_id), f"账单状态更新为: {payload.status}")
    db.commit()
    return {"bill_id": bill_id, "status": bill.status}


@app.post("/invoices")
def create_invoice(payload: InvoiceIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    bill = db.get(Bill, payload.bill_id)
    if not bill:
        raise HTTPException(status_code=404, detail="账单不存在")
    if (bill.lifecycle_status or "active") == "discarded":
        raise HTTPException(status_code=400, detail="已作废账单不可开票")
    invoice = Invoice(
        invoice_no=payload.invoice_no,
        bill_id=payload.bill_id,
        issue_date=payload.issue_date,
        amount_without_tax=payload.amount_without_tax,
        tax_amount=payload.tax_amount,
        total_amount=payload.total_amount,
        status=payload.status,
    )
    db.add(invoice)
    db.flush()
    db.add(InvoiceMeta(invoice_id=invoice.id, remark=payload.remark))
    write_system_audit(db, _["user"], "create_invoice", "invoice", str(invoice.id), f"新增发票: {payload.invoice_no}")
    db.commit()
    return {"id": invoice.id, "invoice_no": invoice.invoice_no}


@app.get("/invoices")
def list_invoices(
    status: Optional[str] = None,
    period: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz])),
):
    rows = db.scalars(select(Invoice).order_by(Invoice.id.desc())).all()
    result = []
    for x in rows:
        bill = db.get(Bill, x.bill_id)
        meta = db.scalar(select(InvoiceMeta).where(InvoiceMeta.invoice_id == x.id))
        item = {
            "id": x.id,
            "invoice_no": x.invoice_no,
            "bill_id": x.bill_id,
            "issue_date": x.issue_date,
            "total_amount": x.total_amount,
            "status": x.status,
            "target_name": bill.target_name if bill else "",
            "period": bill.period if bill else "",
            "created_at": str(meta.created_at if meta else x.issue_date),
            "remark": meta.remark if meta else "",
        }
        result.append(item)
    if status:
        result = [x for x in result if str(x["status"]) == status]
    if period:
        result = [x for x in result if period in str(x["period"])]
    if keyword:
        result = [x for x in result if keyword in f'{x["invoice_no"]}{x["bill_id"]}{x["target_name"]}']
    return result


@app.put("/invoices/{invoice_id}")
def update_invoice(invoice_id: int, payload: InvoiceIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    row = db.get(Invoice, invoice_id)
    if not row:
        raise HTTPException(status_code=404, detail="发票不存在")
    row.invoice_no = payload.invoice_no
    row.bill_id = payload.bill_id
    row.issue_date = payload.issue_date
    row.amount_without_tax = payload.amount_without_tax
    row.tax_amount = payload.tax_amount
    row.total_amount = payload.total_amount
    row.status = payload.status
    meta = db.scalar(select(InvoiceMeta).where(InvoiceMeta.invoice_id == row.id))
    if meta:
        meta.remark = payload.remark
    else:
        db.add(InvoiceMeta(invoice_id=row.id, remark=payload.remark))
    write_system_audit(db, _["user"], "update_invoice", "invoice", str(row.id), f"编辑发票: {payload.invoice_no}")
    db.commit()
    return {"id": row.id, "status": row.status}


@app.post("/receipts")
def register_receipt(payload: ReceiptIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    bill = db.get(Bill, payload.bill_id)
    if not bill:
        raise HTTPException(status_code=404, detail="账单不存在")
    if (bill.lifecycle_status or "active") == "discarded":
        raise HTTPException(status_code=400, detail="已作废账单不可回款")
    receipt = Receipt(
        bill_id=payload.bill_id,
        received_at=payload.received_at,
        amount=payload.amount,
        bank_ref=payload.bank_ref,
        account_name=payload.account_name,
    )
    db.add(receipt)
    db.flush()
    db.add(ReceiptMeta(receipt_id=receipt.id, remark=payload.remark))
    received_total = db.scalar(select(func.coalesce(func.sum(Receipt.amount), 0)).where(Receipt.bill_id == payload.bill_id))
    if received_total >= bill.amount:
        bill.collection_status = CollectionStatus.paid
    elif received_total > 0:
        bill.collection_status = CollectionStatus.partial
    else:
        bill.collection_status = CollectionStatus.pending
    write_system_audit(db, _["user"], "create_receipt", "receipt", str(receipt.id), f"新增回款: bill={payload.bill_id}")
    db.commit()
    return {"receipt_id": receipt.id, "collection_status": bill.collection_status}


@app.get("/receipts")
def list_receipts(
    status: Optional[str] = None,
    period: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz])),
):
    rows = db.scalars(select(Receipt).order_by(Receipt.id.desc())).all()
    result = []
    for x in rows:
        bill = db.get(Bill, x.bill_id)
        meta = db.scalar(select(ReceiptMeta).where(ReceiptMeta.receipt_id == x.id))
        bill_total = Decimal(str(bill.amount)) if bill else Decimal("0")
        received_total = db.scalar(select(func.coalesce(func.sum(Receipt.amount), 0)).where(Receipt.bill_id == x.bill_id)) if bill else Decimal("0")
        item = {
            "id": x.id,
            "bill_id": x.bill_id,
            "received_at": x.received_at,
            "amount": x.amount,
            "bank_ref": x.bank_ref,
            "account_name": x.account_name,
            "target_name": bill.target_name if bill else "",
            "period": bill.period if bill else "",
            "status": bill.collection_status if bill else "",
            "outstanding_amount": max(Decimal("0"), bill_total - Decimal(str(received_total))),
            "remark": meta.remark if meta else "",
            "created_at": str(meta.created_at if meta else x.received_at),
        }
        result.append(item)
    if status:
        result = [x for x in result if str(x["status"]) == status]
    if period:
        result = [x for x in result if period in str(x["period"])]
    if keyword:
        result = [x for x in result if keyword in f'{x["bill_id"]}{x["target_name"]}{x["bank_ref"]}']
    return result


@app.put("/receipts/{receipt_id}")
def update_receipt(receipt_id: int, payload: ReceiptIn, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance]))):
    row = db.get(Receipt, receipt_id)
    if not row:
        raise HTTPException(status_code=404, detail="回款不存在")
    row.bill_id = payload.bill_id
    row.received_at = payload.received_at
    row.amount = payload.amount
    row.bank_ref = payload.bank_ref
    row.account_name = payload.account_name
    meta = db.scalar(select(ReceiptMeta).where(ReceiptMeta.receipt_id == row.id))
    if meta:
        meta.remark = payload.remark
    else:
        db.add(ReceiptMeta(receipt_id=row.id, remark=payload.remark))
    if payload.status:
        bill = db.get(Bill, row.bill_id)
        if bill:
            bill.collection_status = payload.status
    write_system_audit(db, _["user"], "update_receipt", "receipt", str(row.id), f"编辑回款: bill={row.bill_id}")
    db.commit()
    return {"id": row.id}


def _filter_import_history_rows(
    db: Session,
    rows: list[ImportHistory],
    period: Optional[str],
    import_type: Optional[str],
    status: Optional[str],
    keyword: Optional[str],
    task_status: Optional[str] = None,
    lifecycle_status: Optional[str] = "active",
) -> list[ImportHistory]:
    out = list(rows)
    if lifecycle_status in {"active", "discarded"}:
        out = [x for x in out if (x.lifecycle_status or "active") == lifecycle_status]
    if period:
        out = [x for x in out if period in x.period]
    if import_type:
        out = [x for x in out if x.import_type == import_type]
    if status:
        out = [x for x in out if status in x.status]
    if keyword:
        kw = keyword.strip()
        if kw:
            out = [x for x in out if kw in f"{x.file_name}{x.summary}{x.created_by}{x.period}{x.task_id}"]
    if task_status:
        out = [x for x in out if _recon_task_status_str(db, x.task_id) == task_status]
    return out


def _recon_task_status_str(db: Session, task_id: int) -> str:
    task = db.get(ReconTask, task_id)
    if not task:
        return ""
    st = task.status
    return st.value if isinstance(st, ReconStatus) else str(st)


@app.get("/imports/history")
def list_import_history(
    period: Optional[str] = None,
    import_type: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    task_status: Optional[str] = None,
    lifecycle_status: Optional[str] = "active",
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    rows = db.scalars(select(ImportHistory).order_by(ImportHistory.id.desc())).all()
    filtered = _filter_import_history_rows(db, rows, period, import_type, status, keyword, task_status, lifecycle_status)
    total = len(filtered)
    amt = Decimal("0")
    matched_v = 0
    unmatched_v = 0
    for x in filtered:
        amt += Decimal(str(x.amount_sum or 0))
        matched_v += int(x.matched_variant_count or 0)
        unmatched_v += int(x.unmatched_variant_count or 0)
    summary = {
        "batch_count": total,
        "total_import_rows": sum(int(x.total_count or 0) for x in filtered),
        "valid_rows": sum(int(x.valid_count or 0) for x in filtered),
        "invalid_rows": sum(int(x.invalid_count or 0) for x in filtered),
        "amount_sum": str(amt),
        "matched_variant_rows": matched_v,
        "unmatched_variant_rows": unmatched_v,
    }
    start = (max(page, 1) - 1) * max(page_size, 1)
    end = start + max(page_size, 1)
    page_rows = filtered[start:end]
    items = []
    for x in page_rows:
        payload = ImportHistoryOut.model_validate(x).model_dump()
        payload["task_status"] = _recon_task_status_str(db, x.task_id)
        items.append(payload)
    return {"items": items, "total": total, "page": page, "page_size": page_size, "summary": summary}


@app.get("/imports/history/{history_id}", response_model=ImportHistoryOut)
def get_import_history(history_id: int, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops]))):
    row = db.get(ImportHistory, history_id)
    if not row:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    unresolved = db.scalar(select(func.count(ReconIssue.id)).where(ReconIssue.recon_task_id == row.task_id, ReconIssue.resolved.is_(False)))
    resolved = db.scalar(select(func.count(ReconIssue.id)).where(ReconIssue.recon_task_id == row.task_id, ReconIssue.resolved.is_(True)))
    payload = ImportHistoryOut.model_validate(row)
    payload.unresolved_issue_count = int(unresolved or 0)
    payload.resolved_issue_count = int(resolved or 0)
    payload.task_status = _recon_task_status_str(db, row.task_id)
    return payload


@app.get("/imports/history/{history_id}/issues")
def get_import_history_issues(history_id: int, db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops]))):
    history = db.get(ImportHistory, history_id)
    if not history:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    rows = db.scalars(select(ReconIssue).where(ReconIssue.recon_task_id == history.task_id).order_by(ReconIssue.id.desc())).all()
    result = []
    for x in rows:
        meta = db.scalar(select(ReconIssueMeta).where(ReconIssueMeta.issue_id == x.id))
        latest_timeline = db.scalar(select(ReconIssueTimeline).where(ReconIssueTimeline.issue_id == x.id).order_by(ReconIssueTimeline.id.desc()).limit(1))
        result.append(
            {
                "issue_id": x.id,
                "task_id": x.recon_task_id,
                "issue_type": x.issue_type,
                "message": x.detail,
                "status": "已处理" if x.resolved else "未处理",
                "row_no": None,
                "raw_data": None,
                "created_at": "",
                "remark": meta.remark if meta else "",
                "updated_at": str(meta.updated_at) if meta else "",
                "latest_operator": latest_timeline.operator if latest_timeline else "",
                "latest_updated_at": str(latest_timeline.created_at) if latest_timeline else "",
            }
        )
    return result


@app.get("/imports/history/{history_id}/unmatched-variants")
def get_import_history_unmatched_variants(
    history_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    history = db.get(ImportHistory, history_id)
    if not history:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    stmt = (
        select(
            RawStatement.game_name,
            func.count(RawStatement.id).label("cnt"),
            func.max(RawStatement.period).label("latest_period"),
        )
        .where(
            RawStatement.recon_task_id == history.task_id,
            RawStatement.variant_match_status == "未匹配版本",
        )
        .group_by(RawStatement.game_name)
        .order_by(RawStatement.game_name.asc())
    )
    rows = db.execute(stmt).mappings().all()
    return [
        {
            "game_name": row["game_name"],
            "count": int(row["cnt"]),
            "period": row["latest_period"] or history.period,
        }
        for row in rows
    ]


@app.get("/imports/history/{history_id}/batch-stats")
def get_import_history_batch_stats(
    history_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    history = db.get(ImportHistory, history_id)
    if not history:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    task_id = history.task_id
    channel_names_db = set(db.scalars(select(Channel.name)).all())
    game_names_db = set(db.scalars(select(Game.name)).all())
    raw_rows = db.scalars(select(RawStatement).where(RawStatement.recon_task_id == task_id)).all()

    ch_agg: dict[str, dict] = {}
    game_agg: dict[str, dict] = {}
    for r in raw_rows:
        ch = r.channel_name or ""
        gm = r.game_name or ""
        ga = Decimal(str(r.gross_amount or 0))
        if ch not in ch_agg:
            ch_agg[ch] = {"row_count": 0, "gross_amount": Decimal("0")}
        ch_agg[ch]["row_count"] += 1
        ch_agg[ch]["gross_amount"] += ga
        if gm not in game_agg:
            game_agg[gm] = {"row_count": 0, "gross_amount": Decimal("0")}
        game_agg[gm]["row_count"] += 1
        game_agg[gm]["gross_amount"] += ga

    by_channel = [
        {"channel_name": k, "row_count": v["row_count"], "gross_amount": str(v["gross_amount"])}
        for k, v in sorted(ch_agg.items(), key=lambda x: x[0])
    ]
    by_game = [
        {"game_name": k, "row_count": v["row_count"], "gross_amount": str(v["gross_amount"])}
        for k, v in sorted(game_agg.items(), key=lambda x: x[0])
    ]

    unmatched_channels = sorted({r.channel_name for r in raw_rows if r.channel_name and r.channel_name not in channel_names_db})
    unmatched_games = sorted({r.game_name for r in raw_rows if r.game_name and r.game_name not in game_names_db})
    mapping_issues = db.scalars(
        select(ReconIssue).where(ReconIssue.recon_task_id == task_id, ReconIssue.issue_type == "mapping").order_by(ReconIssue.id.asc())
    ).all()
    unmapped_pairs = [x.detail for x in mapping_issues]
    variant_stmt = (
        select(
            RawStatement.game_name,
            func.count(RawStatement.id).label("cnt"),
        )
        .where(RawStatement.recon_task_id == task_id, RawStatement.variant_match_status == "未匹配版本")
        .group_by(RawStatement.game_name)
        .order_by(RawStatement.game_name.asc())
    )
    variant_rows = db.execute(variant_stmt).mappings().all()
    variant_unmatched = [{"game_name": row["game_name"], "count": int(row["cnt"])} for row in variant_rows]

    return {
        "by_channel": by_channel,
        "by_game": by_game,
        "unique_channel_count": len(ch_agg),
        "unique_game_count": len(game_agg),
        "exceptions": {
            "unmatched_channels": unmatched_channels,
            "unmatched_games": unmatched_games,
            "unmapped_pairs": unmapped_pairs,
            "variant_unmatched": variant_unmatched,
        },
    }


@app.get("/imports/history/{history_id}/raw-rows")
def get_import_history_raw_rows(
    history_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    history = db.get(ImportHistory, history_id)
    if not history:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    rows = db.scalars(select(RawStatement).where(RawStatement.recon_task_id == history.task_id).order_by(RawStatement.id.asc())).all()
    return [
        {
            "id": r.id,
            "channel_name": r.channel_name,
            "game_name": r.game_name,
            "period": r.period,
            "gross_amount": str(r.gross_amount),
            "channel_id": r.channel_id,
            "game_id": r.game_id,
            "mapping_id": r.mapping_id,
            "channel_share_percent": str(r.channel_share_percent) if r.channel_share_percent is not None else "",
            "rd_share_percent": str(r.rd_share_percent) if r.rd_share_percent is not None else "",
            "publish_share_percent": str(r.publish_share_percent) if r.publish_share_percent is not None else "",
            "match_status": r.match_status,
            "variant_match_status": r.variant_match_status,
            "project_name": r.project_name or "",
            "variant_name": r.variant_name or "",
        }
        for r in rows
    ]


@app.post("/imports/history/{history_id}/recompute")
def recompute_import_history(
    history_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    history = db.get(ImportHistory, history_id)
    if not history:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    if (history.lifecycle_status or "active") == "discarded":
        raise HTTPException(status_code=400, detail="已作废批次不支持重新计算")
    task = db.get(ReconTask, history.task_id)
    if not task:
        raise HTTPException(status_code=404, detail="导入任务不存在")

    rows = db.scalars(select(RawStatement).where(RawStatement.recon_task_id == task.id).order_by(RawStatement.id.asc())).all()
    channels = {x.name: x for x in db.scalars(select(Channel)).all()}
    games = {x.name: x for x in db.scalars(select(Game)).all()}
    variants = {x.raw_game_name: x for x in db.scalars(select(GameVariant)).all()}
    projects = {x.id: x for x in db.scalars(select(Project)).all()}
    maps = {(x.channel_id, x.game_id): x for x in db.scalars(select(ChannelGameMap)).all()}

    old_issues = db.scalars(select(ReconIssue).where(ReconIssue.recon_task_id == task.id)).all()
    for issue in old_issues:
        meta = db.scalar(select(ReconIssueMeta).where(ReconIssueMeta.issue_id == issue.id))
        if meta:
            db.delete(meta)
        timelines = db.scalars(select(ReconIssueTimeline).where(ReconIssueTimeline.issue_id == issue.id)).all()
        for tl in timelines:
            db.delete(tl)
        db.delete(issue)

    total = len(rows)
    matched = 0
    unmatched = 0
    matched_variant_count = 0
    unmatched_variant_count = 0
    issue_count = 0
    amount_sum = Decimal("0")
    new_issues: list[ReconIssue] = []

    for row in rows:
        amount_sum += Decimal(str(row.gross_amount or 0))
        channel_name = (row.channel_name or "").strip()
        game_name = (row.game_name or "").strip()
        channel = channels.get(channel_name)
        game = games.get(game_name)
        mapping = maps.get((channel.id, game.id)) if channel and game else None

        row.channel_id = channel.id if channel else None
        row.game_id = game.id if game else None
        row.mapping_id = mapping.id if mapping else None

        rd_share = Decimal(str(game.rd_share_percent if game and game.rd_share_percent is not None else 0))
        channel_share = (
            Decimal(str(mapping.revenue_share_ratio or 0)) * Decimal("100")
            if mapping and mapping.revenue_share_ratio is not None
            else None
        )
        row.rd_share_percent = rd_share if game else None
        row.channel_share_percent = channel_share
        row.publish_share_percent = (
            max(Decimal("0"), Decimal("100") - rd_share - channel_share) if channel_share is not None and game else None
        )

        row_issue = False
        if not channel:
            row_issue = True
            new_issues.append(ReconIssue(recon_task_id=task.id, issue_type="unmatched_channel", detail=f"未匹配渠道: {channel_name}"))
        if not game:
            row_issue = True
            new_issues.append(ReconIssue(recon_task_id=task.id, issue_type="unmatched_game", detail=f"未匹配游戏: {game_name}"))
        if channel and game and not mapping:
            row_issue = True
            new_issues.append(
                ReconIssue(recon_task_id=task.id, issue_type="unmapped_pair", detail=f"未映射组合: {channel_name}/{game_name}")
            )

        variant = variants.get(game_name)
        matched_project = projects.get(variant.project_id) if variant else None
        if variant:
            matched_variant_count += 1
            row.variant_id = variant.id
            row.variant_name = variant.variant_name
            row.project_id = matched_project.id if matched_project else None
            row.project_name = matched_project.name if matched_project else None
            row.rd_company = variant.rd_company
            row.variant_match_status = "已匹配版本"
        else:
            unmatched_variant_count += 1
            row.variant_id = None
            row.variant_name = None
            row.project_id = None
            row.project_name = None
            row.rd_company = None
            row.variant_match_status = "未匹配版本"
            row_issue = True
            new_issues.append(ReconIssue(recon_task_id=task.id, issue_type="variant_unmatched", detail=f"版本未匹配: {game_name}"))

        if row_issue:
            row.match_status = "未匹配"
            unmatched += 1
        else:
            row.match_status = "已匹配"
            matched += 1

    for issue in new_issues:
        db.add(issue)
    issue_count = len(new_issues)

    task.status = ReconStatus.issue if issue_count > 0 else ReconStatus.pending
    history.total_count = total
    history.valid_count = matched
    history.invalid_count = unmatched
    history.matched_variant_count = matched_variant_count
    history.unmatched_variant_count = unmatched_variant_count
    history.amount_sum = amount_sum
    history.status = "异常待处理" if issue_count > 0 else "待确认"
    history.summary = f"总行数:{total}, 正常:{matched}, 异常:{unmatched}, 流水合计:{amount_sum}"

    write_system_audit(
        db,
        ctx["user"],
        "recompute_import_history",
        "import_history",
        str(history.id),
        f"重算批次: total={total}, matched={matched}, unmatched={unmatched}, issues={issue_count}",
    )
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"重算失败: {str(e)}")

    return {"total": total, "matched": matched, "unmatched": unmatched, "issues": issue_count}


@app.post("/imports/history/{history_id}/discard")
def discard_import_history(
    history_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    history = db.get(ImportHistory, history_id)
    if not history:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    if (history.lifecycle_status or "active") == "discarded":
        return {"id": history.id, "lifecycle_status": "discarded", "already_discarded": True}
    task = db.get(ReconTask, history.task_id)
    if task and task.status == ReconStatus.confirmed:
        raise HTTPException(status_code=400, detail="已确认入账的批次不可作废")
    history.lifecycle_status = "discarded"
    history.status = "已作废"
    write_system_audit(
        db,
        ctx["user"],
        "discard_import_history",
        "import_history",
        str(history.id),
        f"作废导入批次: task_id={history.task_id}",
    )
    db.commit()
    return {"id": history.id, "lifecycle_status": "discarded"}


@app.post("/imports/history/{history_id}/rematch-variants")
def rematch_import_history_variants(
    history_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance, Role.biz, Role.ops])),
):
    history = db.get(ImportHistory, history_id)
    if not history:
        raise HTTPException(status_code=404, detail="导入历史不存在")
    variants = {x.raw_game_name: x for x in db.scalars(select(GameVariant)).all()}
    projects = {x.id: x for x in db.scalars(select(Project)).all()}
    rows = db.scalars(select(RawStatement).where(RawStatement.recon_task_id == history.task_id)).all()
    rematched_count = 0
    matched_count = 0
    unmatched_count = 0
    for row in rows:
        matched_variant = variants.get(row.game_name)
        matched_project = projects.get(matched_variant.project_id) if matched_variant else None
        prev_unmatched = row.variant_match_status == "未匹配版本"
        if matched_variant:
            row.project_id = matched_project.id if matched_project else None
            row.project_name = matched_project.name if matched_project else None
            row.variant_id = matched_variant.id
            row.variant_name = matched_variant.variant_name
            row.rd_company = matched_variant.rd_company
            row.publish_company = matched_variant.publish_company
            row.rd_share_percent = matched_variant.rd_share_percent
            row.publish_share_percent = matched_variant.publish_share_percent
            row.variant_match_status = "已匹配版本"
            matched_count += 1
            if prev_unmatched:
                rematched_count += 1
        else:
            row.project_id = None
            row.project_name = None
            row.variant_id = None
            row.variant_name = None
            row.rd_company = None
            row.publish_company = None
            row.rd_share_percent = None
            row.publish_share_percent = None
            row.variant_match_status = "未匹配版本"
            unmatched_count += 1
    history.matched_variant_count = matched_count
    history.unmatched_variant_count = unmatched_count
    write_system_audit(
        db,
        ctx["user"],
        "rematch_variants_for_import_history",
        "import_history",
        str(history.id),
        f"重新匹配版本: rematched={rematched_count}, remaining_unmatched={unmatched_count}",
    )
    db.commit()
    return {"rematched_count": rematched_count, "remaining_unmatched_count": unmatched_count}


@app.get("/dashboard/finance")
def finance_dashboard(db: Session = Depends(get_db), _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.ops_manager, Role.tech]))):
    total_receivable = db.scalar(select(func.coalesce(func.sum(Bill.amount), 0)))
    total_received = db.scalar(select(func.coalesce(func.sum(Receipt.amount), 0)))
    total_invoiced = db.scalar(select(func.coalesce(func.sum(Invoice.total_amount), 0)))
    pending_count = db.scalar(select(func.count(Bill.id)).where(Bill.collection_status == CollectionStatus.pending))
    partial_count = db.scalar(select(func.count(Bill.id)).where(Bill.collection_status == CollectionStatus.partial))
    paid_count = db.scalar(select(func.count(Bill.id)).where(Bill.collection_status == CollectionStatus.paid))
    outstanding = Decimal(str(total_receivable)) - Decimal(str(total_received))
    current_period = dt.date.today().strftime("%Y-%m")
    overdue_amount = db.scalar(
        select(func.coalesce(func.sum(Bill.amount), 0)).where(Bill.collection_status != CollectionStatus.paid, Bill.period < current_period)
    )
    recent_period_rows = db.scalars(select(Bill).order_by(Bill.period.desc(), Bill.id.desc()).limit(300)).all()
    period_map: dict[str, dict[str, Decimal]] = {}
    for b in recent_period_rows:
        if b.period not in period_map:
            period_map[b.period] = {"receivable": Decimal("0"), "received": Decimal("0"), "count": Decimal("0")}
        period_map[b.period]["receivable"] += Decimal(str(b.amount))
        period_map[b.period]["count"] += Decimal("1")
        rec_sum = db.scalar(select(func.coalesce(func.sum(Receipt.amount), 0)).where(Receipt.bill_id == b.id))
        period_map[b.period]["received"] += Decimal(str(rec_sum))
    recent_period_summary = []
    for period_key in sorted(period_map.keys(), reverse=True)[:6]:
        item = period_map[period_key]
        recent_period_summary.append(
            {
                "period": period_key,
                "bill_count": int(item["count"]),
                "receivable": item["receivable"],
                "received": item["received"],
                "outstanding": max(Decimal("0"), item["receivable"] - item["received"]),
            }
        )
    pending_bills = db.scalars(select(Bill).where(Bill.collection_status != CollectionStatus.paid).order_by(Bill.id.desc()).limit(20)).all()
    pending_list = []
    flow_status_breakdown = {"草稿": 0, "已生成": 0, "已发送": 0, "已开票": 0, "部分回款": 0, "已回款": 0}
    all_bills = db.scalars(select(Bill)).all()
    for b in all_bills:
        inv_count = db.scalar(select(func.count(Invoice.id)).where(Invoice.bill_id == b.id))
        rec_sum = db.scalar(select(func.coalesce(func.sum(Receipt.amount), 0)).where(Receipt.bill_id == b.id))
        flow_status = calc_bill_flow_status(Decimal(str(b.amount)), b.status, bool(inv_count), Decimal(str(rec_sum)))
        flow_status_breakdown[flow_status] = flow_status_breakdown.get(flow_status, 0) + 1
    for b in pending_bills:
        rec_sum = db.scalar(select(func.coalesce(func.sum(Receipt.amount), 0)).where(Receipt.bill_id == b.id))
        rec_total = Decimal(str(rec_sum))
        pending_list.append(
            {
                "bill_id": b.id,
                "period": b.period,
                "target_name": b.target_name,
                "bill_type": b.bill_type,
                "amount": b.amount,
                "received_total": rec_total,
                "outstanding_amount": max(Decimal("0"), Decimal(str(b.amount)) - rec_total),
                "receipt_status": "部分回款" if rec_total > 0 else "待回款",
                "flow_status": calc_bill_flow_status(Decimal(str(b.amount)), b.status, False, rec_total),
            }
        )
    return {
        "total_receivable": total_receivable,
        "total_invoiced": total_invoiced,
        "total_received": total_received,
        "outstanding": outstanding,
        "overdue_amount": overdue_amount,
        "status_breakdown": {"待回款": pending_count, "部分回款": partial_count, "已回款": paid_count},
        "flow_status_breakdown": flow_status_breakdown,
        "recent_period_summary": recent_period_summary,
        "pending_bills": pending_list,
    }


def _parse_dashboard_range(range_text: str) -> int:
    mapping = {"7d": 7, "30d": 30}
    if range_text not in mapping:
        raise HTTPException(status_code=400, detail="range 仅支持 7d 或 30d")
    return mapping[range_text]


def _safe_ratio_change(current: Decimal, previous: Decimal) -> float:
    if previous == 0:
        return 0.0 if current == 0 else 100.0
    return float(((current - previous) / previous) * Decimal("100"))


def _exception_status_text(status: ExceptionHandleStatus) -> str:
    if status == ExceptionHandleStatus.ignored:
        return "已忽略"
    if status == ExceptionHandleStatus.resolved:
        return "已解决"
    return "待处理"


def _collect_exception_data(
    db: Session,
    days: int,
    status_filter: str = "all",
    type_filter: str = "all",
):
    valid_types = {
        "share",
        "channel",
        "game",
        "import",
        "overdue",
        "unmatched_channel",
        "unmatched_game",
        "unmapped_pair",
        "variant_unmatched",
        "import_failed",
    }
    if type_filter not in {"all", *valid_types}:
        raise HTTPException(status_code=400, detail="type 参数非法")
    if status_filter not in {"all", "pending", "ignored", "resolved"}:
        raise HTTPException(status_code=400, detail="status 参数非法")

    cutoff_dt = dt.datetime.combine(dt.date.today() - dt.timedelta(days=days - 1), dt.time.min)
    channels = set(db.scalars(select(Channel.name)).all())
    active_variants = {
        x for x in db.scalars(select(GameVariant.raw_game_name).where(GameVariant.status == VariantStatus.active)).all() if x
    }

    status_rows = db.scalars(select(ExceptionHandleRecord)).all()
    status_map = {(x.exception_type, x.exception_id): x for x in status_rows}

    def status_for(exception_type: str, exception_id: str) -> ExceptionHandleStatus:
        row = status_map.get((exception_type, exception_id))
        return row.status if row else ExceptionHandleStatus.pending

    def allow(exception_type: str, exception_id: str) -> bool:
        if type_filter != "all" and exception_type != type_filter:
            return False
        if status_filter == "all":
            return True
        return status_for(exception_type, exception_id).value == status_filter

    def allow_with_status(status_value: ExceptionHandleStatus, exception_type: str) -> bool:
        if type_filter != "all" and exception_type != type_filter:
            return False
        if status_filter == "all":
            return True
        return status_value.value == status_filter

    items: dict[str, list[dict]] = {
        "share": [],
        "channel": [],
        "game": [],
        "import": [],
        "overdue": [],
        "unmatched_channel": [],
        "unmatched_game": [],
        "unmapped_pair": [],
        "variant_unmatched": [],
        "import_failed": [],
    }

    map_rows = db.scalars(select(ChannelGameMap)).all()
    for row in map_rows:
        total_ratio = Decimal(str(row.revenue_share_ratio or 0)) + Decimal(str(row.rd_settlement_ratio or 0))
        if not (total_ratio > Decimal("1.0001") or total_ratio < Decimal("0.9999")):
            continue
        exception_id = f"share-{row.id}"
        if not allow("share", exception_id):
            continue
        channel_share = Decimal(str(row.revenue_share_ratio or 0))
        rd_share = Decimal(str(row.rd_settlement_ratio or 0))
        tax_rate = Decimal("0")
        private_share = Decimal("0")
        publisher_share = Decimal("1") - channel_share - rd_share - tax_rate - private_share
        items["share"].append(
            {
                "id": exception_id,
                "type": "share",
                "status": status_for("share", exception_id).value,
                "detected_at": dt.datetime.now().isoformat(),
                "updated_at": (status_map.get(("share", exception_id)).updated_at.isoformat() if status_map.get(("share", exception_id)) else None),
                "source_module": "channel_game_map",
                "channel_name": row.channel.name if row.channel else "",
                "game_name": row.game.name if row.game else "",
                "channel_share": channel_share,
                "tax_rate": tax_rate,
                "rd_share": rd_share,
                "private_share": private_share,
                "publisher_share": publisher_share,
                "total_ratio": total_ratio,
                "status_text": _exception_status_text(status_for("share", exception_id)),
            }
        )

    raw_rows = db.scalars(select(RawStatement).where(RawStatement.created_at >= cutoff_dt)).all()
    for row in raw_rows:
        if row.channel_name not in channels:
            exception_id = f"channel-{row.id}"
            if allow("channel", exception_id):
                history = db.scalar(select(ImportHistory).where(ImportHistory.task_id == row.recon_task_id))
                items["channel"].append(
                    {
                        "id": exception_id,
                        "type": "channel",
                        "status": status_for("channel", exception_id).value,
                        "detected_at": row.created_at.isoformat() if row.created_at else dt.datetime.now().isoformat(),
                        "updated_at": (status_map.get(("channel", exception_id)).updated_at.isoformat() if status_map.get(("channel", exception_id)) else None),
                        "source_module": "import",
                        "import_history_id": history.id if history else None,
                        "batch_name": history.file_name if history else f"task-{row.recon_task_id}",
                        "raw_channel_name": row.channel_name,
                        "match_status": "未匹配渠道",
                        "status_text": _exception_status_text(status_for("channel", exception_id)),
                    }
                )
        if row.game_name not in active_variants:
            exception_id = f"game-{row.id}"
            if allow("game", exception_id):
                history = db.scalar(select(ImportHistory).where(ImportHistory.task_id == row.recon_task_id))
                items["game"].append(
                    {
                        "id": exception_id,
                        "type": "game",
                        "status": status_for("game", exception_id).value,
                        "detected_at": row.created_at.isoformat() if row.created_at else dt.datetime.now().isoformat(),
                        "updated_at": (status_map.get(("game", exception_id)).updated_at.isoformat() if status_map.get(("game", exception_id)) else None),
                        "source_module": "import",
                        "import_history_id": history.id if history else None,
                        "batch_name": history.file_name if history else f"task-{row.recon_task_id}",
                        "raw_game_name": row.game_name,
                        "match_status": "未匹配游戏",
                        "status_text": _exception_status_text(status_for("game", exception_id)),
                    }
                )

    import_rows = db.scalars(
        select(ImportHistory).where(
            ImportHistory.created_at >= cutoff_dt,
            (ImportHistory.invalid_count > 0) | (ImportHistory.status.like("%失败%")),
        )
    ).all()
    for row in import_rows:
        exception_id = f"import-{row.id}"
        if not allow("import", exception_id):
            continue
        fail_reason = f"异常行{row.invalid_count}条，状态={row.status}"
        items["import"].append(
            {
                "id": exception_id,
                "type": "import",
                "status": status_for("import", exception_id).value,
                "detected_at": row.created_at.isoformat() if row.created_at else dt.datetime.now().isoformat(),
                "updated_at": (status_map.get(("import", exception_id)).updated_at.isoformat() if status_map.get(("import", exception_id)) else None),
                "source_module": "import_history",
                "import_history_id": row.id,
                "batch_name": row.file_name or f"history-{row.id}",
                "fail_reason": fail_reason,
                "invalid_count": int(row.invalid_count or 0),
                "status_text": _exception_status_text(status_for("import", exception_id)),
            }
        )

    # 导入批次异常（ReconIssue）统一接入异常中心
    issue_rows = (
        db.execute(
            select(ReconIssue, ImportHistory)
            .join(ImportHistory, ImportHistory.task_id == ReconIssue.recon_task_id)
            .where(ImportHistory.created_at >= cutoff_dt, ImportHistory.lifecycle_status == "active")
            .order_by(ReconIssue.id.desc())
        )
        .all()
    )
    issue_type_map = {
        "unmatched_channel": "unmatched_channel",
        "unmatched_game": "unmatched_game",
        "unmapped_pair": "unmapped_pair",
        "variant_unmatched": "variant_unmatched",
        "mapping": "unmapped_pair",
        "import_failed": "import_failed",
    }
    for issue, history in issue_rows:
        normalized_type = issue_type_map.get(issue.issue_type, "import_failed")
        exception_id = f"recon-{issue.id}"
        row_status = status_map.get((normalized_type, exception_id))
        effective_status = row_status.status if row_status else (ExceptionHandleStatus.resolved if issue.resolved else ExceptionHandleStatus.pending)
        if not allow_with_status(effective_status, normalized_type):
            continue

        payload = {
            "id": exception_id,
            "type": normalized_type,
            "issue_type": issue.issue_type,
            "status": effective_status.value,
            "detected_at": history.created_at.isoformat() if history.created_at else dt.datetime.now().isoformat(),
            "updated_at": row_status.updated_at.isoformat() if row_status and row_status.updated_at else None,
            "source_module": "import_issue",
            "import_history_id": history.id,
            "task_id": issue.recon_task_id,
            "period": history.period,
            "batch_name": history.file_name or f"history-{history.id}",
            "detail": issue.detail,
            "status_text": _exception_status_text(effective_status),
        }
        if normalized_type == "unmatched_channel":
            payload["raw_channel_name"] = issue.detail.split(":", 1)[-1].strip() if ":" in issue.detail else issue.detail
        if normalized_type == "unmatched_game":
            payload["raw_game_name"] = issue.detail.split(":", 1)[-1].strip() if ":" in issue.detail else issue.detail
        items[normalized_type].append(payload)

    current_period = dt.date.today().strftime("%Y-%m")
    overdue_rows = db.scalars(
        select(Bill).where(
            Bill.collection_status != CollectionStatus.paid,
            Bill.period < current_period,
        )
    ).all()
    for row in overdue_rows:
        exception_id = f"overdue-{row.id}"
        if not allow("overdue", exception_id):
            continue
        period_date = dt.datetime.strptime(f"{row.period}-01", "%Y-%m-%d").date() if len(row.period) == 7 else dt.date.today()
        overdue_days = max(0, (dt.date.today() - period_date).days)
        items["overdue"].append(
            {
                "id": exception_id,
                "type": "overdue",
                "status": status_for("overdue", exception_id).value,
                "detected_at": row.created_at.isoformat() if row.created_at else dt.datetime.now().isoformat(),
                "updated_at": (status_map.get(("overdue", exception_id)).updated_at.isoformat() if status_map.get(("overdue", exception_id)) else None),
                "source_module": "billing",
                "channel_name": row.target_name if row.bill_type == BillType.channel else "",
                "game_name": "",
                "period": row.period,
                "bill_amount": row.amount,
                "overdue_days": overdue_days,
                "status_text": _exception_status_text(status_for("overdue", exception_id)),
            }
        )

    summary = {
        "share": len(items["share"]),
        "channel": len(items["channel"]),
        "game": len(items["game"]),
        "import": len(items["import"]),
        "overdue": len(items["overdue"]),
        "unmatched_channel": len(items["unmatched_channel"]),
        "unmatched_game": len(items["unmatched_game"]),
        "unmapped_pair": len(items["unmapped_pair"]),
        "variant_unmatched": len(items["variant_unmatched"]),
        "import_failed": len(items["import_failed"]),
    }
    summary["total"] = (
        summary["share"]
        + summary["channel"]
        + summary["game"]
        + summary["import"]
        + summary["overdue"]
        + summary["unmatched_channel"]
        + summary["unmatched_game"]
        + summary["unmapped_pair"]
        + summary["variant_unmatched"]
        + summary["import_failed"]
    )
    return {"summary": summary, "items": items}


@app.get("/dashboard/overview")
def dashboard_overview(
    range: str = Query(default="7d"),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.ops_manager, Role.tech])),
):
    days = _parse_dashboard_range(range)
    today = dt.date.today()
    month_start = today.replace(day=1)
    prev_month_end = month_start - dt.timedelta(days=1)
    prev_month_start = prev_month_end.replace(day=1)
    current_period = today.strftime("%Y-%m")
    prev_period = prev_month_start.strftime("%Y-%m")

    # 口径：本月总流水 = raw_statements 在当前自然月账期(period=YYYY-MM)的 gross_amount 汇总
    monthly_gross_revenue = Decimal(
        str(db.scalar(select(func.coalesce(func.sum(RawStatement.gross_amount), 0)).where(RawStatement.period == current_period)) or 0)
    )
    previous_month_gross_revenue = Decimal(
        str(db.scalar(select(func.coalesce(func.sum(RawStatement.gross_amount), 0)).where(RawStatement.period == prev_period)) or 0)
    )

    # 口径：本月渠道回款 = Receipt 关联 Bill(period=当前月) 的实际回款金额汇总
    monthly_channel_receipts = Decimal(
        str(
            db.scalar(
                select(func.coalesce(func.sum(Receipt.amount), 0))
                .select_from(Receipt)
                .join(Bill, Bill.id == Receipt.bill_id)
                .where(Bill.period == current_period)
            )
            or 0
        )
    )
    previous_month_channel_receipts = Decimal(
        str(
            db.scalar(
                select(func.coalesce(func.sum(Receipt.amount), 0))
                .select_from(Receipt)
                .join(Bill, Bill.id == Receipt.bill_id)
                .where(Bill.period == prev_period)
            )
            or 0
        )
    )

    # 口径：本月应付研发 = rd 账单(BillType.rd)在当前月账期的 amount 汇总
    monthly_rd_payable = Decimal(
        str(
            db.scalar(
                select(func.coalesce(func.sum(Bill.amount), 0)).where(Bill.bill_type == BillType.rd, Bill.period == current_period)
            )
            or 0
        )
    )
    previous_month_rd_payable = Decimal(
        str(
            db.scalar(
                select(func.coalesce(func.sum(Bill.amount), 0)).where(Bill.bill_type == BillType.rd, Bill.period == prev_period)
            )
            or 0
        )
    )

    # 口径：本月毛利润 = 本月渠道回款 - 本月应付研发
    monthly_gross_profit = monthly_channel_receipts - monthly_rd_payable
    previous_month_gross_profit = previous_month_channel_receipts - previous_month_rd_payable

    # 口径：未结算金额 = 所有未完全回款账单(amount - 已回款)的剩余金额汇总
    receipt_sum = (
        select(Receipt.bill_id.label("bill_id"), func.coalesce(func.sum(Receipt.amount), 0).label("received_total"))
        .group_by(Receipt.bill_id)
        .subquery("receipt_sum")
    )
    outstanding_stmt = (
        select(func.coalesce(func.sum(Bill.amount - func.coalesce(receipt_sum.c.received_total, 0)), 0))
        .select_from(Bill)
        .outerjoin(receipt_sum, receipt_sum.c.bill_id == Bill.id)
        .where(Bill.collection_status != CollectionStatus.paid)
    )
    unsettled_amount = Decimal(str(db.scalar(outstanding_stmt) or 0))
    previous_unsettled_amount = Decimal(
        str(
            db.scalar(
                select(func.coalesce(func.sum(Bill.amount), 0)).where(
                    Bill.collection_status != CollectionStatus.paid,
                    Bill.period == prev_period,
                )
            )
            or 0
        )
    )

    # 异常口径与异常中心同源：复用统一异常聚合函数
    exception_data_current = _collect_exception_data(db, days=days, status_filter="all", type_filter="all")
    exception_bill_count = int(exception_data_current["summary"]["total"])

    previous_exception_bill_count = int(
        db.scalar(
            select(func.count(ImportHistory.id)).where(
                ImportHistory.created_at >= dt.datetime.combine(prev_month_start, dt.time.min),
                ImportHistory.created_at < dt.datetime.combine(month_start, dt.time.min),
                (ImportHistory.invalid_count > 0) | (ImportHistory.status.like("%失败%")),
            )
        )
        or 0
    )

    start_day = today - dt.timedelta(days=days - 1)
    date_keys = [(start_day + dt.timedelta(days=i)).strftime("%Y-%m-%d") for i in range(days)]
    flow_map: dict[str, Decimal] = {d: Decimal("0") for d in date_keys}
    receipt_map: dict[str, Decimal] = {d: Decimal("0") for d in date_keys}
    rd_map: dict[str, Decimal] = {d: Decimal("0") for d in date_keys}

    flow_rows = db.execute(
        select(func.date(RawStatement.created_at).label("d"), func.coalesce(func.sum(RawStatement.gross_amount), 0).label("s"))
        .where(RawStatement.created_at >= cutoff_dt)
        .group_by(func.date(RawStatement.created_at))
    ).all()
    for d, s in flow_rows:
        key = str(d)
        if key in flow_map:
            flow_map[key] = Decimal(str(s or 0))

    receipt_rows = db.execute(
        select(func.date(Receipt.received_at).label("d"), func.coalesce(func.sum(Receipt.amount), 0).label("s"))
        .where(Receipt.received_at >= start_day)
        .group_by(func.date(Receipt.received_at))
    ).all()
    for d, s in receipt_rows:
        key = str(d)
        if key in receipt_map:
            receipt_map[key] = Decimal(str(s or 0))

    rd_rows = db.execute(
        select(func.date(Bill.created_at).label("d"), func.coalesce(func.sum(Bill.amount), 0).label("s"))
        .where(Bill.bill_type == BillType.rd, Bill.created_at >= cutoff_dt)
        .group_by(func.date(Bill.created_at))
    ).all()
    for d, s in rd_rows:
        key = str(d)
        if key in rd_map:
            rd_map[key] = Decimal(str(s or 0))

    trends: list[dict] = []
    for d in date_keys:
        flow_amount = flow_map[d]
        receipt_amount = receipt_map[d]
        profit_amount = receipt_amount - rd_map[d]
        trends.append({"date": d, "type": "流水", "amount": flow_amount})
        trends.append({"date": d, "type": "回款", "amount": receipt_amount})
        trends.append({"date": d, "type": "利润", "amount": profit_amount})

    recent_logs = db.scalars(select(SystemAuditLog).order_by(SystemAuditLog.id.desc()).limit(10)).all()
    recent_activities = [
        {
            "id": x.id,
            "operator": x.operator,
            "action_type": x.action,
            "detail": x.summary,
            "created_at": x.created_at,
        }
        for x in recent_logs
    ]

    return {
        "summary": {
            "monthly_gross_revenue": monthly_gross_revenue,
            "monthly_channel_receipts": monthly_channel_receipts,
            "monthly_rd_payable": monthly_rd_payable,
            "monthly_gross_profit": monthly_gross_profit,
            "unsettled_amount": unsettled_amount,
            "exception_bill_count": exception_bill_count,
        },
        "summary_compare": {
            "monthly_gross_revenue": _safe_ratio_change(monthly_gross_revenue, previous_month_gross_revenue),
            "monthly_channel_receipts": _safe_ratio_change(monthly_channel_receipts, previous_month_channel_receipts),
            "monthly_rd_payable": _safe_ratio_change(monthly_rd_payable, previous_month_rd_payable),
            "monthly_gross_profit": _safe_ratio_change(monthly_gross_profit, previous_month_gross_profit),
            "unsettled_amount": _safe_ratio_change(unsettled_amount, previous_unsettled_amount),
            "exception_bill_count": _safe_ratio_change(Decimal(str(exception_bill_count)), Decimal(str(previous_exception_bill_count))),
        },
        "trends": trends,
        "exceptions": {
            "total": exception_data_current["summary"]["total"],
            "share": exception_data_current["summary"]["share"],
            "channel": exception_data_current["summary"]["channel"],
            "game": exception_data_current["summary"]["game"],
            "import": exception_data_current["summary"]["import"],
            "overdue": exception_data_current["summary"]["overdue"],
        },
        "recent_activities": recent_activities,
    }


@app.get("/exceptions/overview")
def exceptions_overview(
    range: str = Query(default="30d"),
    status: str = Query(default="all"),
    type: str = Query(default="all"),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.ops_manager, Role.tech])),
):
    range_mapping = {"7d": 7, "30d": 30, "90d": 90}
    if range not in range_mapping:
        raise HTTPException(status_code=400, detail="range 仅支持 7d/30d/90d")
    payload = _collect_exception_data(db, days=range_mapping[range], status_filter=status, type_filter=type)
    return payload


@app.post("/exceptions/status")
def update_exception_status(
    payload: ExceptionStatusPatchIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager])),
):
    if payload.type not in {
        "share",
        "channel",
        "game",
        "import",
        "overdue",
        "unmatched_channel",
        "unmatched_game",
        "unmapped_pair",
        "variant_unmatched",
        "import_failed",
    }:
        raise HTTPException(status_code=400, detail="type 参数非法")
    row = db.scalar(
        select(ExceptionHandleRecord).where(
            ExceptionHandleRecord.exception_type == payload.type,
            ExceptionHandleRecord.exception_id == payload.id,
        )
    )
    if row:
        row.status = payload.status
        row.remark = payload.remark
        row.updated_by = ctx["user"]
        row.updated_at = dt.datetime.now()
    else:
        db.add(
            ExceptionHandleRecord(
                exception_type=payload.type,
                exception_id=payload.id,
                status=payload.status,
                remark=payload.remark,
                updated_by=ctx["user"],
                updated_at=dt.datetime.now(),
            )
        )
    write_system_audit(
        db,
        ctx["user"],
        "update_exception_status",
        "exception",
        f"{payload.type}:{payload.id}",
        f"状态更新为 {payload.status.value}",
    )
    db.commit()
    return {"ok": True}


@app.get("/audit/logs")
def list_system_audit_logs(
    operator: Optional[str] = None,
    action: Optional[str] = None,
    target_type: Optional[str] = None,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin])),
):
    rows = db.scalars(select(SystemAuditLog).order_by(SystemAuditLog.id.desc())).all()
    if operator:
        rows = [x for x in rows if operator in x.operator]
    if action:
        rows = [x for x in rows if action in x.action]
    if target_type:
        rows = [x for x in rows if target_type in x.target_type]
    if start_time:
        start_dt = dt.datetime.fromisoformat(start_time)
        rows = [x for x in rows if x.created_at >= start_dt]
    if end_time:
        end_dt = dt.datetime.fromisoformat(end_time)
        rows = [x for x in rows if x.created_at <= end_dt]
    total = len(rows)
    start = (max(page, 1) - 1) * max(page_size, 1)
    end = start + max(page_size, 1)
    return {"items": rows[start:end], "total": total, "page": page, "page_size": page_size}


@app.get("/audit-logs")
def list_audit_logs_compat(
    db: Session = Depends(get_db),
    _: dict = Depends(require_role([Role.admin])),
):
    return db.scalars(select(SystemAuditLog).order_by(SystemAuditLog.id.desc()).limit(200)).all()


# --- 渠道月度结算对账单（导入 → 明细 → 生成账单），API 前缀 monthly-settlement-* 与旧版 /settlement-statements 区分 ---

MONTHLY_SETTLEMENT_READ_ROLES = [
    Role.admin,
    Role.finance_manager,
    Role.finance,
    Role.ops_manager,
    Role.ops,
    Role.tech,
    Role.biz,
]
MONTHLY_SETTLEMENT_WRITE_ROLES = [Role.admin, Role.finance_manager, Role.finance, Role.ops_manager, Role.ops]

_CAPITAL_NUMS = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
_CAPITAL_UNITS = ["", "拾", "佰", "仟"]


def _section_to_chinese(section: int) -> str:
    s = ""
    zero = False
    for i in range(3, -1, -1):
        d = section // (10**i) % 10
        if d == 0:
            zero = True
        else:
            if zero:
                s += _CAPITAL_NUMS[0]
                zero = False
            s += _CAPITAL_NUMS[d] + _CAPITAL_UNITS[i]
    s = s.rstrip(_CAPITAL_NUMS[0])
    return s or _CAPITAL_NUMS[0]


def _integer_part_to_chinese(n: int) -> str:
    if n == 0:
        return _CAPITAL_NUMS[0]
    parts: list[str] = []
    unit_large = ["", "万", "亿"]
    idx = 0
    while n > 0 and idx < len(unit_large):
        section = n % 10000
        if section != 0:
            parts.append(_section_to_chinese(section) + unit_large[idx])
        else:
            if parts and not parts[-1].endswith(_CAPITAL_NUMS[0]):
                parts.append(_CAPITAL_NUMS[0])
        n //= 10000
        idx += 1
    s = "".join(reversed(parts))
    while "零零" in s:
        s = s.replace("零零", "零")
    return s


def _rmb_upper_case(amount: Decimal) -> str:
    amount = _money2(amount)
    if amount < 0:
        return "（负数）" + _rmb_upper_case(-amount)
    if amount == 0:
        return "人民币零元整"
    integer = int(amount)
    frac = _money2(amount - Decimal(integer))
    jiao_fen = int((frac * 100).quantize(Decimal("1")))
    jiao, fen = jiao_fen // 10, jiao_fen % 10
    head = _integer_part_to_chinese(integer)
    if jiao == 0 and fen == 0:
        return f"人民币{head}元整"
    tail = ""
    if jiao != 0:
        tail += _CAPITAL_NUMS[jiao] + "角"
    if fen != 0:
        tail += _CAPITAL_NUMS[fen] + "分"
    if jiao == 0 and fen != 0:
        tail = _CAPITAL_NUMS[0] + tail
    return f"人民币{head}元" + tail


def _normalize_ratio_cell(raw: object) -> Optional[Decimal]:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    if isinstance(raw, str) and not raw.strip():
        return None
    try:
        d = Decimal(str(raw).strip().rstrip("%"))
    except Exception:
        return None
    if d > 1:
        d = d / Decimal("100")
    if d < 0 or d > 1:
        return None
    return d.quantize(Decimal("0.0001"))


def _parse_money_cell(raw: object, default: Decimal = Decimal("0")) -> Decimal:
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return _money2(default)
    s = str(raw).strip().replace(",", "")
    if not s:
        return _money2(default)
    try:
        return _money2(Decimal(s))
    except Exception:
        return _money2(default)


def _pick_contract_for_month(
    db: Session, channel_id: int, game_id: int, settlement_month: str
) -> Optional[ChannelGameContract]:
    ref = dt.datetime.strptime(f"{settlement_month}-01", "%Y-%m-%d").date()
    rows = db.scalars(
        select(ChannelGameContract)
        .where(
            ChannelGameContract.channel_id == channel_id,
            ChannelGameContract.game_id == game_id,
            ChannelGameContract.status == "active",
            ChannelGameContract.effective_start_date <= ref,
        )
        .order_by(ChannelGameContract.effective_start_date.desc(), ChannelGameContract.id.desc())
    ).all()
    for r in rows:
        if r.effective_end_date is None or r.effective_end_date >= ref:
            return r
    return None


def _recalc_settlement_detail_row(
    db: Session,
    row: SettlementDetailRow,
    channel_id: int,
    game_id: Optional[int],
):
    """participation = gross - test - coupon; settlement = participation * (1 - channel_fee) * revenue_share"""
    g = _money2(row.gross_revenue)
    t = _money2(row.test_fee or 0)
    c = _money2(row.coupon_fee or 0)
    row.gross_revenue = g
    row.test_fee = t
    row.coupon_fee = c
    row.participation_amount = _money2(g - t - c)
    rsh = row.revenue_share_ratio
    cf = row.channel_fee_ratio
    if game_id:
        if rsh is None or cf is None:
            contract = _pick_contract_for_month(db, channel_id, game_id, row.settlement_month)
            if contract:
                if rsh is None:
                    rsh = contract.revenue_share_ratio
                if cf is None:
                    cf = contract.channel_fee_ratio
    row.revenue_share_ratio = rsh
    row.channel_fee_ratio = cf
    errs: list[str] = []
    if not game_id:
        errs.append("游戏未匹配")
    if rsh is None:
        errs.append("缺少分成比例（行内与合同均未配置）")
    if cf is None:
        errs.append("缺少渠道费比例（行内与合同均未配置）")
    if row.participation_amount < 0:
        errs.append("参与分成金额为负")
    if errs:
        row.settlement_amount = Decimal("0")
        if row.row_status != "pending_confirm":
            row.row_status = "error"
        if not row.error_message:
            row.error_message = "；".join(errs)
        return
    one_minus_cf = _money2(Decimal("1") - cf)
    row.settlement_amount = _money2(row.participation_amount * one_minus_cf * rsh)
    if row.row_status == "error" and errs == []:
        row.row_status = "normal"
    if row.row_status not in ("pending_confirm", "used_in_statement"):
        row.row_status = "normal"
        row.error_message = ""


def _resolve_game_for_import(db: Session, game_name: str, game_code_hint: Optional[str]) -> tuple[Optional[int], str]:
    name = (game_name or "").strip()
    code = (game_code_hint or "").strip() or None
    if code:
        g = db.scalar(select(Game).where(Game.game_code == code))
        if g:
            return g.id, g.name
    if name:
        g2 = db.scalar(select(Game).where(Game.name == name))
        if g2:
            return g2.id, g2.name
    return None, name


def _df_column_aliases() -> dict[str, list[str]]:
    return {
        "settlement_month": ["settlement_month", "账期", "结算月份", "settlement month"],
        "channel_name": ["channel_name", "渠道名称", "渠道"],
        "game_name": ["game_name", "游戏名称", "游戏"],
        "game_code": ["game_code", "游戏编码", "游戏code"],
        "gross_revenue": ["gross_revenue", "流水", "总收入", "充值流水", "gross"],
        "test_fee": ["test_fee", "测试费", "测试充值"],
        "coupon_fee": ["coupon_fee", "代金券", "优惠券", "券抵扣"],
        "revenue_share_ratio": ["revenue_share_ratio", "分成比例", "流水分成比例"],
        "channel_fee_ratio": ["channel_fee_ratio", "渠道费比例", "渠道费"],
        "remark": ["remark", "备注"],
    }


def _normalize_import_df(df: pd.DataFrame) -> pd.DataFrame:
    aliases = _df_column_aliases()
    col_map: dict[str, str] = {}
    lower_map = {str(c).strip().lower(): str(c).strip() for c in df.columns}
    for canonical, names in aliases.items():
        for n in names:
            k = n.strip().lower()
            if k in lower_map:
                col_map[lower_map[k]] = canonical
                break
    return df.rename(columns=col_map)


def _next_monthly_statement_no(db: Session, settlement_month: str) -> str:
    prefix = "MS" + settlement_month.replace("-", "")
    like_pat = f"{prefix}-%"
    n = int(db.scalar(select(func.count(ChannelMonthlySettlementStatement.id)).where(ChannelMonthlySettlementStatement.statement_no.like(like_pat))) or 0)
    return f"{prefix}-{n + 1:04d}"


@app.get("/settlement-imports/template")
def settlement_import_template(
    _: dict = Depends(require_role(MONTHLY_SETTLEMENT_READ_ROLES)),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"
    headers = [
        "settlement_month",
        "channel_name",
        "game_name",
        "gross_revenue",
        "test_fee",
        "coupon_fee",
        "revenue_share_ratio",
        "channel_fee_ratio",
        "remark",
    ]
    zh = ["账期(YYYY-MM)", "渠道名称", "游戏名称", "流水(必填)", "测试费", "代金券", "分成比例(0~1或%)", "渠道费比例(0~1或%)", "备注"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
        ws.cell(row=2, column=i, value=zh[i - 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="settlement_import_template.xlsx"'},
    )


@app.post("/settlement-imports/upload")
async def settlement_import_upload(
    settlement_month: str = Query(..., description="账期 YYYY-MM"),
    channel_id: int = Query(..., description="渠道 id，与文件内渠道名称一致"),
    batch_name: str = Query("", description="批次名称，可空"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role(MONTHLY_SETTLEMENT_WRITE_ROLES)),
):
    ch = db.get(Channel, channel_id)
    if not ch:
        raise HTTPException(status_code=400, detail="渠道不存在")
    month = _normalize_period_yyyymm(settlement_month)
    raw_name = file.filename or "upload"
    lower = raw_name.lower()
    try:
        content = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取文件失败: {e}") from e
    if lower.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(content))
        import_type = "csv"
    elif lower.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(content))
        import_type = "excel"
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xls/.csv")
    df = _normalize_import_df(df)
    required = {"channel_name", "game_name", "gross_revenue"}
    missing = required - set(df.columns)
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少列: {', '.join(sorted(missing))}")
    batch = SettlementImportBatch(
        batch_name=(batch_name or "").strip() or raw_name,
        source_file_name=raw_name[:500],
        settlement_month=month,
        channel_id=channel_id,
        import_type=import_type,
        total_rows=len(df),
        success_rows=0,
        failed_rows=0,
        import_status="processing",
        created_by=ctx["user"],
    )
    db.add(batch)
    db.flush()

    err_lines: list[str] = []
    for idx, r in df.iterrows():
        row_no = int(idx) + 1 if isinstance(idx, int) else idx + 1
        file_ch = str(r.get("channel_name", "")).strip()
        if file_ch != ch.name:
            detail = SettlementDetailRow(
                batch_id=batch.id,
                settlement_month=month,
                channel_id=channel_id,
                game_id=None,
                raw_game_name=str(r.get("game_name", "")),
                game_name_snapshot="",
                gross_revenue=Decimal("0"),
                row_status="error",
                error_message=f"渠道名称与所选不一致（第{row_no}行）",
            )
            db.add(detail)
            batch.failed_rows += 1
            err_lines.append(f"行{row_no}: 渠道名称不匹配")
            continue
        sm = r.get("settlement_month")
        row_month = _normalize_period_yyyymm(str(sm).strip()) if sm is not None and str(sm).strip() else month
        if row_month != month:
            detail = SettlementDetailRow(
                batch_id=batch.id,
                settlement_month=month,
                channel_id=channel_id,
                game_id=None,
                raw_game_name=str(r.get("game_name", "")),
                game_name_snapshot="",
                gross_revenue=Decimal("0"),
                row_status="error",
                error_message=f"账期与上传所选不一致（第{row_no}行）",
            )
            db.add(detail)
            batch.failed_rows += 1
            err_lines.append(f"行{row_no}: 账期不匹配")
            continue
        game_hint_code = r.get("game_code") if "game_code" in df.columns else None
        gid, snap_name = _resolve_game_for_import(db, str(r.get("game_name", "")), None if game_hint_code is None else str(game_hint_code))
        gross = _parse_money_cell(r.get("gross_revenue"))
        test_f = _parse_money_cell(r.get("test_fee") if "test_fee" in df.columns else None, Decimal("0"))
        coupon_f = _parse_money_cell(r.get("coupon_fee") if "coupon_fee" in df.columns else None, Decimal("0"))
        rsh = _normalize_ratio_cell(r.get("revenue_share_ratio") if "revenue_share_ratio" in df.columns else None)
        cfr = _normalize_ratio_cell(r.get("channel_fee_ratio") if "channel_fee_ratio" in df.columns else None)
        remark = str(r.get("remark") if "remark" in df.columns else "").strip()
        detail = SettlementDetailRow(
            batch_id=batch.id,
            settlement_month=month,
            channel_id=channel_id,
            game_id=gid,
            raw_game_name=str(r.get("game_name", "")),
            game_name_snapshot=snap_name or str(r.get("game_name", "")),
            gross_revenue=gross,
            test_fee=test_f,
            coupon_fee=coupon_f,
            revenue_share_ratio=rsh,
            channel_fee_ratio=cfr,
            remark=remark[:500],
        )
        if not gid:
            detail.row_status = "error"
            detail.error_message = "游戏名称在系统中不存在"
            batch.failed_rows += 1
            err_lines.append(f"行{row_no}: 游戏未匹配")
        db.add(detail)
        db.flush()
        if gid:
            _recalc_settlement_detail_row(db, detail, channel_id, gid)
            if detail.row_status == "error":
                batch.failed_rows += 1
                err_lines.append(f"行{row_no}: {detail.error_message}")
            else:
                batch.success_rows += 1
        batch.total_rows = len(df)

    batch.import_status = "completed" if batch.failed_rows == 0 else ("partial" if batch.success_rows else "failed")
    batch.error_summary = "\n".join(err_lines[:30])[:2000]
    batch.updated_at = dt.datetime.now()
    write_system_audit(
        db,
        ctx["user"],
        "settlement_import_upload",
        "import_batch",
        str(batch.id),
        f"导入结算明细 {month} / {ch.name} 成功{batch.success_rows} 失败{batch.failed_rows}",
    )
    db.commit()
    return {
        "batch_id": batch.id,
        "total_rows": batch.total_rows,
        "success_rows": batch.success_rows,
        "failed_rows": batch.failed_rows,
        "import_status": batch.import_status,
        "error_summary": batch.error_summary,
    }


@app.get("/settlement-imports")
def list_settlement_imports(
    settlement_month: Optional[str] = None,
    channel_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role(MONTHLY_SETTLEMENT_READ_ROLES)),
):
    stmt = select(SettlementImportBatch).order_by(SettlementImportBatch.id.desc())
    if settlement_month:
        stmt = stmt.where(SettlementImportBatch.settlement_month == _normalize_period_yyyymm(settlement_month))
    if channel_id:
        stmt = stmt.where(SettlementImportBatch.channel_id == channel_id)
    rows = db.scalars(stmt).all()
    total = len(rows)
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)
    slice_rows = rows[(page - 1) * page_size : page * page_size]
    _b_cids = [x.channel_id for x in slice_rows]
    ch_names = (
        {c.id: c.name for c in db.scalars(select(Channel).where(Channel.id.in_(_b_cids))).all()} if _b_cids else {}
    )
    return {
        "items": [
            {
                "id": x.id,
                "batch_name": x.batch_name,
                "source_file_name": x.source_file_name,
                "settlement_month": x.settlement_month,
                "channel_id": x.channel_id,
                "channel_name": ch_names.get(x.channel_id, ""),
                "import_type": x.import_type,
                "total_rows": x.total_rows,
                "success_rows": x.success_rows,
                "failed_rows": x.failed_rows,
                "import_status": x.import_status,
                "error_summary": x.error_summary,
                "created_by": x.created_by,
                "created_at": x.created_at,
            }
            for x in slice_rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/settlement-details")
def list_settlement_details(
    batch_id: Optional[int] = None,
    settlement_month: Optional[str] = None,
    channel_id: Optional[int] = None,
    row_status: Optional[str] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role(MONTHLY_SETTLEMENT_READ_ROLES)),
):
    stmt = select(SettlementDetailRow).order_by(SettlementDetailRow.id.desc())
    if batch_id:
        stmt = stmt.where(SettlementDetailRow.batch_id == batch_id)
    if settlement_month:
        stmt = stmt.where(SettlementDetailRow.settlement_month == _normalize_period_yyyymm(settlement_month))
    if channel_id:
        stmt = stmt.where(SettlementDetailRow.channel_id == channel_id)
    if row_status:
        stmt = stmt.where(SettlementDetailRow.row_status == row_status)
    rows = db.scalars(stmt).all()
    if keyword:
        kw = keyword.strip().lower()
        rows = [x for x in rows if kw in (x.raw_game_name or "").lower() or kw in (x.game_name_snapshot or "").lower() or kw in (x.remark or "").lower()]
    stats = {
        "total": len(rows),
        "normal": sum(1 for x in rows if x.row_status == "normal"),
        "error": sum(1 for x in rows if x.row_status == "error"),
        "pending_confirm": sum(1 for x in rows if x.row_status == "pending_confirm"),
        "used_in_statement": sum(1 for x in rows if x.row_status == "used_in_statement"),
    }
    total = len(rows)
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)
    slice_rows = rows[(page - 1) * page_size : page * page_size]
    gids = [x.game_id for x in slice_rows if x.game_id]
    games = (
        {g.id: g.name for g in db.scalars(select(Game).where(Game.id.in_(gids))).all()} if gids else {}
    )
    cids = [x.channel_id for x in slice_rows]
    chs = (
        {c.id: c.name for c in db.scalars(select(Channel).where(Channel.id.in_(cids))).all()} if cids else {}
    )
    return {
        "stats": stats,
        "items": [
            {
                "id": x.id,
                "batch_id": x.batch_id,
                "settlement_month": x.settlement_month,
                "channel_id": x.channel_id,
                "channel_name": chs.get(x.channel_id, ""),
                "game_id": x.game_id,
                "game_name": games.get(x.game_id, "") if x.game_id else "",
                "raw_game_name": x.raw_game_name,
                "game_name_snapshot": x.game_name_snapshot,
                "gross_revenue": x.gross_revenue,
                "test_fee": x.test_fee,
                "coupon_fee": x.coupon_fee,
                "participation_amount": x.participation_amount,
                "revenue_share_ratio": x.revenue_share_ratio,
                "channel_fee_ratio": x.channel_fee_ratio,
                "settlement_amount": x.settlement_amount,
                "row_status": x.row_status,
                "error_message": x.error_message,
                "remark": x.remark,
                "monthly_statement_id": x.monthly_statement_id,
                "created_at": x.created_at,
                "updated_at": x.updated_at,
            }
            for x in slice_rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.patch("/settlement-details/{detail_id}")
def patch_settlement_detail(
    detail_id: int,
    payload: SettlementDetailPatchIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role(MONTHLY_SETTLEMENT_WRITE_ROLES)),
):
    row = db.get(SettlementDetailRow, detail_id)
    if not row:
        raise HTTPException(status_code=404, detail="明细不存在")
    if row.row_status == "used_in_statement":
        raise HTTPException(status_code=400, detail="已用于生成账单的明细不可编辑")
    if payload.game_id is not None:
        g = db.get(Game, payload.game_id)
        if not g:
            raise HTTPException(status_code=400, detail="游戏不存在")
        row.game_id = payload.game_id
        row.game_name_snapshot = g.name
    if payload.test_fee is not None:
        row.test_fee = _money2(payload.test_fee)
    if payload.coupon_fee is not None:
        row.coupon_fee = _money2(payload.coupon_fee)
    if payload.revenue_share_ratio is not None:
        row.revenue_share_ratio = payload.revenue_share_ratio
    if payload.channel_fee_ratio is not None:
        row.channel_fee_ratio = payload.channel_fee_ratio
    if payload.remark is not None:
        row.remark = (payload.remark or "").strip()[:500]
    if payload.row_status is not None:
        row.row_status = payload.row_status
    _recalc_settlement_detail_row(db, row, row.channel_id, row.game_id)
    row.updated_at = dt.datetime.now()
    write_system_audit(
        db,
        ctx["user"],
        "settlement_detail_patch",
        "settlement_detail",
        str(row.id),
        "修改结算明细",
    )
    db.commit()
    return {"id": row.id, "row_status": row.row_status, "settlement_amount": row.settlement_amount}


@app.post("/monthly-settlement-statements/generate")
def monthly_settlement_generate(
    payload: MonthlySettlementGenerateIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role(MONTHLY_SETTLEMENT_WRITE_ROLES)),
):
    month = _normalize_period_yyyymm(payload.settlement_month)
    ch = db.get(Channel, payload.channel_id)
    if not ch:
        raise HTTPException(status_code=400, detail="渠道不存在")
    existing = db.scalar(
        select(ChannelMonthlySettlementStatement).where(
            ChannelMonthlySettlementStatement.settlement_month == month,
            ChannelMonthlySettlementStatement.channel_id == payload.channel_id,
        )
    )
    if existing and not payload.overwrite:
        raise HTTPException(status_code=400, detail="该账期与渠道已存在账单，若需覆盖请传 overwrite=true")
    details = db.scalars(
        select(SettlementDetailRow).where(
            SettlementDetailRow.settlement_month == month,
            SettlementDetailRow.channel_id == payload.channel_id,
            SettlementDetailRow.monthly_statement_id.is_(None),
            SettlementDetailRow.row_status.in_(["normal", "pending_confirm"]),
            SettlementDetailRow.game_id.isnot(None),
        )
    ).all()
    if not details:
        raise HTTPException(status_code=400, detail="没有可用于生成的明细（需状态为正常/待确认且未绑定账单）")
    by_game: dict[int, list[SettlementDetailRow]] = {}
    for d in details:
        by_game.setdefault(d.game_id or 0, []).append(d)
    ratio_map: dict[int, tuple[Decimal, Decimal]] = {}
    for gid, drows in by_game.items():
        keys = {(x.revenue_share_ratio, x.channel_fee_ratio) for x in drows}
        if len(keys) > 1:
            raise HTTPException(
                status_code=400,
                detail=f"游戏 ID {gid} 在同一账期与渠道下存在不一致的分成/渠道费比例，请先修正明细",
            )
        r0 = drows[0]
        if r0.revenue_share_ratio is None or r0.channel_fee_ratio is None:
            raise HTTPException(status_code=400, detail=f"游戏 ID {gid} 比例不完整，无法生成账单")
        ratio_map[gid] = (r0.revenue_share_ratio, r0.channel_fee_ratio)

    if existing and payload.overwrite:
        old_id = existing.id
        linked = db.scalars(select(SettlementDetailRow).where(SettlementDetailRow.monthly_statement_id == old_id)).all()
        for dr in linked:
            dr.monthly_statement_id = None
            if dr.row_status == "used_in_statement":
                dr.row_status = "normal"
        db.execute(delete(ChannelMonthlySettlementItem).where(ChannelMonthlySettlementItem.statement_id == old_id))
        db.delete(existing)
        db.flush()

    contract_sample = None
    for gid in ratio_map:
        contract_sample = _pick_contract_for_month(db, payload.channel_id, gid, month)
        if contract_sample:
            break

    agg: dict[int, dict] = {}
    for d in details:
        gid = d.game_id or 0
        if gid not in agg:
            g = db.get(Game, gid)
            agg[gid] = {
                "game_name": g.name if g else d.game_name_snapshot,
                "gross_revenue": Decimal("0"),
                "test_fee": Decimal("0"),
                "coupon_fee": Decimal("0"),
                "participation_amount": Decimal("0"),
                "settlement_amount": Decimal("0"),
                "revenue_share_ratio": d.revenue_share_ratio or Decimal("0"),
                "channel_fee_ratio": d.channel_fee_ratio or Decimal("0"),
                "remark": "",
            }
        a = agg[gid]
        a["gross_revenue"] = _money2(a["gross_revenue"] + d.gross_revenue)
        a["test_fee"] = _money2(a["test_fee"] + d.test_fee)
        a["coupon_fee"] = _money2(a["coupon_fee"] + d.coupon_fee)
        a["participation_amount"] = _money2(a["participation_amount"] + d.participation_amount)
        a["settlement_amount"] = _money2(a["settlement_amount"] + d.settlement_amount)

    total_gross = _money2(sum((x["gross_revenue"] for x in agg.values()), start=Decimal("0")))
    total_test = _money2(sum((x["test_fee"] for x in agg.values()), start=Decimal("0")))
    total_coupon = _money2(sum((x["coupon_fee"] for x in agg.values()), start=Decimal("0")))
    total_part = _money2(sum((x["participation_amount"] for x in agg.values()), start=Decimal("0")))
    total_settle = _money2(sum((x["settlement_amount"] for x in agg.values()), start=Decimal("0")))

    st_no = _next_monthly_statement_no(db, month)
    c0 = contract_sample
    statement_title = f"{ch.name} 渠道月度结算对账单 {month}"
    if c0 and (c0.statement_title_template or "").strip():
        tmpl = (c0.statement_title_template or "").strip()
        try:
            statement_title = tmpl.format(channel_name=ch.name, settlement_month=month, month=month)
        except Exception:
            statement_title = tmpl.replace("{channel_name}", ch.name).replace("{settlement_month}", month)

    stmt = ChannelMonthlySettlementStatement(
        statement_no=st_no,
        settlement_month=month,
        channel_id=payload.channel_id,
        statement_title=statement_title,
        our_company_name=(c0.our_company_name if c0 else "") or "广州熊动科技有限公司",
        opposite_company_name=(c0.opposite_company_name if c0 else "") or ch.name,
        total_gross_revenue=total_gross,
        total_test_fee=total_test,
        total_coupon_fee=total_coupon,
        total_participation_amount=total_part,
        total_settlement_amount=total_settle,
        total_settlement_amount_cn=_rmb_upper_case(total_settle),
        statement_status="draft",
        remark="",
        our_company_address=(c0.our_company_address if c0 else "") or "",
        our_company_phone=(c0.our_company_phone if c0 else "") or "",
        our_tax_no=(c0.our_tax_no if c0 else "") or "",
        our_bank_name=(c0.our_bank_name if c0 else "") or "",
        our_bank_account=(c0.our_bank_account if c0 else "") or "",
        opposite_tax_no=(c0.opposite_tax_no if c0 else "") or "",
        opposite_bank_name=(c0.opposite_bank_name if c0 else "") or "",
        opposite_bank_account=(c0.opposite_bank_account if c0 else "") or "",
        created_by=ctx["user"],
    )
    db.add(stmt)
    db.flush()

    sort_no = 0
    for _gid, a in sorted(agg.items(), key=lambda kv: kv[1]["game_name"]):
        sort_no += 1
        db.add(
            ChannelMonthlySettlementItem(
                statement_id=stmt.id,
                sort_no=sort_no,
                settlement_month=month,
                game_name=a["game_name"],
                gross_revenue=a["gross_revenue"],
                test_fee=a["test_fee"],
                coupon_fee=a["coupon_fee"],
                participation_amount=a["participation_amount"],
                revenue_share_ratio=a["revenue_share_ratio"],
                channel_fee_ratio=a["channel_fee_ratio"],
                settlement_amount=a["settlement_amount"],
                remark="",
            )
        )
    for d in details:
        d.monthly_statement_id = stmt.id
        d.row_status = "used_in_statement"
        d.updated_at = dt.datetime.now()

    write_system_audit(
        db,
        ctx["user"],
        "monthly_settlement_generate",
        "monthly_settlement_statement",
        str(stmt.id),
        f"生成月度账单 {month} / {ch.name}",
    )
    db.commit()
    return {"id": stmt.id, "statement_no": stmt.statement_no}


@app.get("/monthly-settlement-statements")
def list_monthly_settlement_statements(
    settlement_month: Optional[str] = None,
    channel_id: Optional[int] = None,
    statement_status: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role(MONTHLY_SETTLEMENT_READ_ROLES)),
):
    stmt = select(ChannelMonthlySettlementStatement).order_by(ChannelMonthlySettlementStatement.id.desc())
    if settlement_month:
        stmt = stmt.where(ChannelMonthlySettlementStatement.settlement_month == _normalize_period_yyyymm(settlement_month))
    if channel_id:
        stmt = stmt.where(ChannelMonthlySettlementStatement.channel_id == channel_id)
    if statement_status:
        stmt = stmt.where(ChannelMonthlySettlementStatement.statement_status == statement_status)
    rows = db.scalars(stmt).all()
    total = len(rows)
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)
    slice_rows = rows[(page - 1) * page_size : page * page_size]
    _mcids = [x.channel_id for x in slice_rows]
    chs = (
        {c.id: c.name for c in db.scalars(select(Channel).where(Channel.id.in_(_mcids))).all()} if _mcids else {}
    )
    return {
        "items": [
            {
                "id": x.id,
                "statement_no": x.statement_no,
                "settlement_month": x.settlement_month,
                "channel_id": x.channel_id,
                "channel_name": chs.get(x.channel_id, ""),
                "statement_title": x.statement_title,
                "total_gross_revenue": x.total_gross_revenue,
                "total_test_fee": x.total_test_fee,
                "total_coupon_fee": x.total_coupon_fee,
                "total_participation_amount": x.total_participation_amount,
                "total_settlement_amount": x.total_settlement_amount,
                "total_settlement_amount_cn": x.total_settlement_amount_cn,
                "statement_status": x.statement_status,
                "confirmed_at": x.confirmed_at,
                "exported_at": x.exported_at,
                "paid_at": x.paid_at,
                "created_by": x.created_by,
                "created_at": x.created_at,
                "updated_at": x.updated_at,
            }
            for x in slice_rows
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@app.get("/monthly-settlement-statements/{statement_id}")
def get_monthly_settlement_statement(
    statement_id: int,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role(MONTHLY_SETTLEMENT_READ_ROLES)),
):
    statement = db.get(ChannelMonthlySettlementStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="账单不存在")
    channel = db.get(Channel, statement.channel_id)
    items = db.scalars(
        select(ChannelMonthlySettlementItem)
        .where(ChannelMonthlySettlementItem.statement_id == statement_id)
        .order_by(ChannelMonthlySettlementItem.sort_no.asc(), ChannelMonthlySettlementItem.id.asc())
    ).all()
    return {
        "id": statement.id,
        "statement_no": statement.statement_no,
        "settlement_month": statement.settlement_month,
        "channel_id": statement.channel_id,
        "channel_name": channel.name if channel else "",
        "statement_title": statement.statement_title,
        "our_company_name": statement.our_company_name,
        "our_company_address": statement.our_company_address,
        "our_company_phone": statement.our_company_phone,
        "our_tax_no": statement.our_tax_no,
        "our_bank_name": statement.our_bank_name,
        "our_bank_account": statement.our_bank_account,
        "opposite_company_name": statement.opposite_company_name,
        "opposite_tax_no": statement.opposite_tax_no,
        "opposite_bank_name": statement.opposite_bank_name,
        "opposite_bank_account": statement.opposite_bank_account,
        "total_gross_revenue": statement.total_gross_revenue,
        "total_test_fee": statement.total_test_fee,
        "total_coupon_fee": statement.total_coupon_fee,
        "total_participation_amount": statement.total_participation_amount,
        "total_settlement_amount": statement.total_settlement_amount,
        "total_settlement_amount_cn": statement.total_settlement_amount_cn,
        "statement_status": statement.statement_status,
        "confirmed_at": statement.confirmed_at,
        "exported_at": statement.exported_at,
        "paid_at": statement.paid_at,
        "remark": statement.remark,
        "created_by": statement.created_by,
        "created_at": statement.created_at,
        "updated_at": statement.updated_at,
        "items": [
            {
                "id": x.id,
                "sort_no": x.sort_no,
                "settlement_month": x.settlement_month,
                "game_name": x.game_name,
                "gross_revenue": x.gross_revenue,
                "test_fee": x.test_fee,
                "coupon_fee": x.coupon_fee,
                "participation_amount": x.participation_amount,
                "revenue_share_ratio": x.revenue_share_ratio,
                "channel_fee_ratio": x.channel_fee_ratio,
                "settlement_amount": x.settlement_amount,
                "remark": x.remark,
            }
            for x in items
        ],
    }


@app.patch("/monthly-settlement-statements/{statement_id}/status")
def patch_monthly_settlement_status(
    statement_id: int,
    payload: MonthlySettlementStatusPatchIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role(MONTHLY_SETTLEMENT_WRITE_ROLES)),
):
    statement = db.get(ChannelMonthlySettlementStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="账单不存在")
    now = dt.datetime.now()
    statement.statement_status = payload.statement_status
    if payload.statement_status == "confirmed":
        statement.confirmed_at = now
    if payload.statement_status == "exported":
        statement.exported_at = now
    if payload.statement_status == "paid":
        statement.paid_at = now
    statement.updated_at = now
    write_system_audit(
        db,
        ctx["user"],
        "monthly_settlement_status",
        "monthly_settlement_statement",
        str(statement.id),
        f"状态 → {payload.statement_status}",
    )
    db.commit()
    return {"id": statement.id, "statement_status": statement.statement_status}


@app.get("/monthly-settlement-statements/{statement_id}/export-excel")
def export_monthly_settlement_excel(
    statement_id: int,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role(MONTHLY_SETTLEMENT_READ_ROLES)),
):
    statement = db.get(ChannelMonthlySettlementStatement, statement_id)
    if not statement:
        raise HTTPException(status_code=404, detail="账单不存在")
    channel = db.get(Channel, statement.channel_id)
    ch_label = channel.name if channel else ""
    items = db.scalars(
        select(ChannelMonthlySettlementItem)
        .where(ChannelMonthlySettlementItem.statement_id == statement_id)
        .order_by(ChannelMonthlySettlementItem.sort_no.asc(), ChannelMonthlySettlementItem.id.asc())
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "对账单"
    row_ix = 1
    ws.merge_cells(start_row=row_ix, start_column=1, end_row=row_ix, end_column=9)
    tcell = ws.cell(
        row=row_ix,
        column=1,
        value=statement.statement_title or f"{ch_label} 渠道月度结算对账单 {statement.settlement_month}",
    )
    tcell.font = Font(bold=True, size=14)
    tcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_ix += 1
    ws.merge_cells(start_row=row_ix, start_column=1, end_row=row_ix, end_column=9)
    ws.cell(row=row_ix, column=1, value=f"账单编号：{statement.statement_no}    账期：{statement.settlement_month}")
    row_ix += 2
    headers = [
        "序号",
        "游戏",
        "流水",
        "测试费",
        "代金券",
        "可参与分成金额",
        "流水分成比例",
        "渠道费比例",
        "结算金额",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row_ix, column=c, value=h)
    row_ix += 1
    for it in items:
        ws.cell(row=row_ix, column=1, value=it.sort_no)
        ws.cell(row=row_ix, column=2, value=it.game_name)
        ws.cell(row=row_ix, column=3, value=float(it.gross_revenue))
        ws.cell(row=row_ix, column=4, value=float(it.test_fee))
        ws.cell(row=row_ix, column=5, value=float(it.coupon_fee))
        ws.cell(row=row_ix, column=6, value=float(it.participation_amount))
        ws.cell(row=row_ix, column=7, value=float(it.revenue_share_ratio))
        ws.cell(row=row_ix, column=8, value=float(it.channel_fee_ratio))
        ws.cell(row=row_ix, column=9, value=float(it.settlement_amount))
        row_ix += 1
    row_ix += 1
    ws.merge_cells(start_row=row_ix, start_column=1, end_row=row_ix, end_column=6)
    ws.cell(row=row_ix, column=1, value="合计（结算金额）")
    ws.cell(row=row_ix, column=7, value=float(statement.total_settlement_amount))
    row_ix += 1
    ws.merge_cells(start_row=row_ix, start_column=1, end_row=row_ix, end_column=9)
    ws.cell(row=row_ix, column=1, value=f"大写金额：{statement.total_settlement_amount_cn}")
    row_ix += 2
    ws.cell(row=row_ix, column=1, value="备注说明：")
    row_ix += 1
    ws.merge_cells(start_row=row_ix, start_column=1, end_row=row_ix, end_column=9)
    ws.cell(row=row_ix, column=1, value=statement.remark or "")
    row_ix += 2
    ws.cell(row=row_ix, column=1, value="甲方（我司）")
    ws.cell(row=row_ix, column=6, value="乙方（渠道）")
    row_ix += 1
    ws.cell(row=row_ix, column=1, value=f"名称：{statement.our_company_name}")
    ws.cell(row=row_ix, column=6, value=f"名称：{statement.opposite_company_name}")
    row_ix += 1
    ws.cell(row=row_ix, column=1, value=f"地址：{statement.our_company_address}")
    ws.cell(row=row_ix, column=6, value=f"税号：{statement.opposite_tax_no}")
    row_ix += 1
    ws.cell(row=row_ix, column=1, value=f"电话：{statement.our_company_phone}")
    ws.cell(row=row_ix, column=6, value=f"开户行：{statement.opposite_bank_name}")
    row_ix += 1
    ws.cell(row=row_ix, column=1, value=f"税号：{statement.our_tax_no}")
    ws.cell(row=row_ix, column=6, value=f"账号：{statement.opposite_bank_account}")
    row_ix += 1
    ws.cell(row=row_ix, column=1, value=f"开户行/账号：{statement.our_bank_name} {statement.our_bank_account}")
    row_ix += 3
    ws.merge_cells(start_row=row_ix, start_column=1, end_row=row_ix, end_column=4)
    ws.merge_cells(start_row=row_ix, start_column=6, end_row=row_ix, end_column=9)
    ws.cell(row=row_ix, column=1, value="甲方盖章：___________    日期：___________")
    ws.cell(row=row_ix, column=6, value="乙方盖章：___________    日期：___________")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"渠道月度结算对账单_{ch_label}_{statement.settlement_month}.xlsx"
    write_system_audit(
        db,
        ctx["user"],
        "monthly_settlement_export",
        "monthly_settlement_statement",
        str(statement.id),
        "导出 Excel",
    )
    db.commit()
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/settlement-contracts")
def list_settlement_contracts(
    channel_id: Optional[int] = None,
    game_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: dict = Depends(require_role(MONTHLY_SETTLEMENT_READ_ROLES)),
):
    stmt = select(ChannelGameContract).order_by(ChannelGameContract.id.desc())
    if channel_id:
        stmt = stmt.where(ChannelGameContract.channel_id == channel_id)
    if game_id:
        stmt = stmt.where(ChannelGameContract.game_id == game_id)
    rows = db.scalars(stmt).all()
    return rows


@app.post("/settlement-contracts")
def create_settlement_contract(
    payload: SettlementContractIn,
    db: Session = Depends(get_db),
    ctx: dict = Depends(require_role([Role.admin, Role.finance_manager, Role.finance])),
):
    row = ChannelGameContract(
        channel_id=payload.channel_id,
        game_id=payload.game_id,
        effective_start_date=payload.effective_start_date,
        effective_end_date=payload.effective_end_date,
        revenue_share_ratio=payload.revenue_share_ratio,
        channel_fee_ratio=payload.channel_fee_ratio,
        deduct_test_fee=payload.deduct_test_fee,
        deduct_coupon_fee=payload.deduct_coupon_fee,
        our_company_name=payload.our_company_name,
        our_company_address=payload.our_company_address,
        our_company_phone=payload.our_company_phone,
        our_tax_no=payload.our_tax_no,
        our_bank_name=payload.our_bank_name,
        our_bank_account=payload.our_bank_account,
        opposite_company_name=payload.opposite_company_name,
        opposite_tax_no=payload.opposite_tax_no,
        opposite_bank_name=payload.opposite_bank_name,
        opposite_bank_account=payload.opposite_bank_account,
        statement_title_template=payload.statement_title_template,
        remark=payload.remark,
        status=payload.status,
    )
    db.add(row)
    db.flush()
    write_system_audit(db, ctx["user"], "create_settlement_contract", "settlement_contract", str(row.id), "新增结算合同配置")
    db.commit()
    return {"id": row.id}
